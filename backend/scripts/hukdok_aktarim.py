#!/usr/bin/env python3
"""HUKDOK aktarımının ÇEKİRDEK yazma yolu (G064) — iskelet + kuru koşu.

Tam 68 sütunluk eşleme BU TURUN İŞİ DEĞİL (final export + CEVAP xlsx geldikten
sonra ayrı plan turu). Burada iskelet kurulur ve KANITLANIR: SistemNo anahtarlı
idempotent upsert, satır bazlı hata izolasyonu, kuru koşu, belge envanter
denkliği, kardeş-föy çelişki raporu.

    docker compose exec -T backend python scripts/hukdok_aktarim.py \\
        --input /app/data/teslim.xlsx --dry-run
    docker compose exec -T backend python scripts/hukdok_aktarim.py \\
        --input /app/data/teslim.xlsx --limit 50

İşletim modeli (FAZ F gereksinim belgesi §0): aktarım bir OLAY DEĞİL, TEKRAR
EDEN bir süreçtir — göz hastalıkları dilimi partiler hâlinde gelecek, ayrıca
dört düzeltme listesi yolda. "Aynı girdiyle iki kez koşulduğunda kayıt sayısı
değişmez" bu yüzden süs değil, işletim modelinin kendisi.

Pazarlıksız kurallar
--------------------
* **SistemNo idempotency anahtarıdır** (`case_foys`, G063). İkinci koşu satır
  ikilemez; değişmeyen alan için `case_history` satırı DA yazılmaz — yoksa
  "değişiklik yok" koşusu tarihçeyi şişirir ve idempotentlik ölçülemez hâle
  gelirdi.
* **Mevcut kartta YALNIZ UPDATE.** Kart DELETE+INSERT yasak; `case_parties`
  satırları toptan silinip yeniden yazılmaz. Sebep şemada: `case_documents.
  case_party_id` FK'sı `ondelete=SET NULL` (models.py:770) — toptan taraf
  silme belge-taraf bağını HATA VERMEDEN koparır. Koşu öncesi/sonrası belge
  envanteri (`services/belge_envanteri.py`) DENK çıkmak zorundadır; denk
  değilse koşu KENDİNİ GERİ ALIR (kapı commit'ten ÖNCE ölçer — hasar kalıcı
  olmadan durdurulur) ve NONZERO çıkar.
* **Satır hatası SAVEPOINT ile izole** (temizlik planı §8: `rollback` değil).
  Hatalı satır kısmen yazılmış olabilir (föy INSERT'i alan doğrulamasından
  ÖNCE gelir); savepoint o kısmı da geri alır. Parti DÜŞMEZ, satır rapora
  düşer. Yarım föy yazıp devam etmek, düzeltme listesiyle zaten geri gelecek
  bir satır için sessiz veri bozulması olurdu.
* **`None` = "bu teslimde yok", "boşalt" DEĞİL** (foy_map ile aynı sözleşme):
  partili teslimde eksik sütun mevcut değeri silmez. Alan boşaltmanın TEK
  yolu açık düzeltme talimatıdır (G112, `Düzeltme_Logu` sayfası): Yeni Değer
  `(boş)` VE `Sheet`'te o hücre gerçekten boş VE bizde dolu — ÜÇLÜ şart
  sağlanmadan boşaltma yok (`_bosaltma_talimatlari`). Aynı sayfanın
  "Gerekçe" sütunu değişen alanın `case_history.source` imzasına provenance
  olarak eklenir; imza `HUKDOK_TESLIM_` ile başlamaya devam eder.
* **Kart YARATILMAZ.** Bu ağırlıkla bir UPDATE dalgasıdır (eşleştirme köprüsü
  DosyaNo↔`klasor_no_2` %97,4); eşleşmeyen satır rapora düşer. Kart açmak
  ofis dosya numarasını SharePoint sayacından ATOMİK tahsis etmeyi gerektirir
  (CLAUDE.md belge akışı) — çevrimdışı bir aktarım scriptinin işi değildir.
* **İlerleme imleci YOK, gerekmiyor**: işlenmiş SistemNo'lar `case_foys`'ta
  yaşar; yarım kalan koşuyu tekrar başlatmak bedava (idempotentlik).
* `statement_timeout` koşu süresince açıkça yükseltilir (§8 madde 6): engine
  30 sn ile bağlanır, toplu yazma bunu meşru aşar.

YAZILAN kart alanları (tam eşleme turu 2026-08-19; G104 eki 2026-09-02)
-----------------------------------------------------------------------
`KART_ALANLARI` + `KART_TURETILEN` sözlükleri tek doğruluk kaynağıdır; hepsi
`kart_degerleri()`ndan geçer. Kabaca: sınıflandırma (`file_type`, `status`,
`subject`, `court`), esas (`esas_no` — tarihçe yolundan), tarihler
(`opening_date`, `acceptance_date`, `arsiv_tarihi`, `arabuluculuk_karar_tarihi`),
para (`islah_tutari`, `manevi_tazminat`, D4 ile türetilen `maddi_tazminat`, üç
`hukmedilen_*`), dosya numaraları (`hasar_dosya_no`, `hukuk_no`), süreç
(`dosya_son_durumu`, `bureau_type`, `arabuluculuk_no`), G044'ün tıbbi beşlisi
ve belgeleme olayı alanları (`olay_turu`, `hukumdeki_rol` — G103 kapalı
listelerine AD bazlı eşleme; tanınmayan değer YAZILMAZ, satır raporuna düşer;
G104).

Avukatlar AYRI yoldan gider: "Sorumlu Avukatlar" bir listedir, `case_lawyers`
satırlarına YALNIZ-EKLEME ile açılır; kartın tek kutusu (`responsible_lawyer_name`)
ancak föyde TEK isim varsa yazılır. Yazım teslimin aksansız hâli değil bizim
kayıtlı yazımımızdır (`avukat_haritasi_kur`).

**Bilinçli YAZILMAYANLAR** (gerekçeleri ölçümle, 2026-08-19): `service_type`
(bitmask semantiği kararlaşmadı) · `Ek Alt Kırılım*` (karşı tarafın kendi
uyarısı: dosya açılış etiketi, güncel değil) · `Para Birimi`/`MüvekkilNo`
(taşınmaz). `court` ve `sub_type` İÇERİK farkında yazılır, yalnız yazım
farkında dokunulmaz (`ICERIK_KARSILASTIRMALI_ALANLAR`).
Karar künyesi
(`karar_no`/`karar_tarihi`) BİLİNÇLİ YAZILMAZ ve BOŞALTILMAZ — o kolonların
tek yazma yolu `managers/stage_decisions.py`ın aşama fotoğrafıdır (G062);
buradan yazmak ikinci bir yazıcı doğururdu. Künye yalnız OKUNUR ve kardeş-föy
çelişki raporunu üretir; `Düzeltme_Logu`'ndaki künye boşaltma talimatı
uygulanmaz, satır raporuna düşer. Boşaltma `ICERIK_KARSILASTIRMALI_ALANLAR`
(`court`/`sub_type`) için de geçersizdir (yazım bizim; rapora düşer).
`cases.sistem_no`/`cases.tku_no` da yazılmaz (nihai tekilleştirme tam eşleme
turunun işi).

`scripts/import_excel_cases.py` KULLANILMAZ ve çağrılmaz (temizlik planı §8:
idempotent değil, hata yolunda sessiz veri kaybı, `-2` mükerrer üretimi).
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple, cast

from sqlalchemy import func, text
from sqlalchemy.exc import SQLAlchemyError

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# E402 (import'tan önce sys.path kurulumu) scripts/* için pyproject'te bilinçli
# olarak kapalıdır — script tek başına da koşabilmeli.
import models
from managers import case_manager, foy_map, seed_data, stage_decisions
from managers.reference_lists import tr_title, tr_upper
from party_check import normalize_party_key
from required_fields import AKTARIM_SOURCE_PREFIX
from services import belge_envanteri

logger = logging.getLogger("HukdokAktarim")

# ─── Çıkış kodları ───────────────────────────────────────────────────────────
# En ağırı kazanır. 2 belge koruma şartının ihlalidir: koşu geri alınmıştır.
CIKIS_TAMAM = 0
CIKIS_SATIR_HATASI = 1
CIKIS_ENVANTER = 2
CIKIS_GIRDI = 3

# Engine 30 sn statement_timeout ile bağlanır (database.py); toplu yazma bunu
# meşru aşar. Koşu süresince açıkça yükseltilir (§8 madde 6).
VARSAYILAN_TIMEOUT_MS = 600_000

DEGISTIREN = "hukdok_aktarim"

# ─── Kaynak sütunlar ─────────────────────────────────────────────────────────
# Aday adlar; ilk eşleşen kullanılır. Başlık karşılaştırması aksan ve boşluk
# duyarsızdır (_baslik_anahtari) — teslim paketleri arasında "Islah Tutarı" /
# "İslah Tutarı" gibi yazım farkları görülüyor. TAM 68 sütunluk eşleme bu turun
# işi DEĞİLDİR; buradaki küme yalnız çekirdeği besler.
SUTUN_ADAYLARI: Dict[str, Tuple[str, ...]] = {
    "sistem_no":    ("SistemNo", "Sistem No"),
    # "Klasör No": teslim paketinde TKU grup anahtarının GERÇEK başlığı budur
    # (SUTUN_SOZLUGU sayfası #4, KİMLİK, 8.152 dolu; değerler "TKU-784"). 100
    # föylük provada (19.08) föylerin TAMAMI tku_no'suz doğmuştu — aday listesi
    # yalnız "TKU*" varyantlarını tanıyordu. `dosya_no`nun "Klasör No.2"
    # adayıyla çakışmaz: başlık anahtarları KLASORNO ≠ KLASORNO2.
    "tku_no":       ("TKU", "TKU No", "TKU No.", "Klasör No"),
    "hasar_no":     ("Hasar No", "Hasar Numarası"),
    "dosya_no":     ("Dosya No", "DosyaNo", "Klasör No.2"),
    "arsiv_tarihi": ("Arşiv Tarihi",),
    "islah_tutari": ("Islah Tutarı", "İslah Tutarı"),
    "tibbi_olay":   ("Tıbbi Olay",),
    "karar_no":     ("Karar No",),
    "karar_tarihi": ("Karar Tarihi",),
    # --- tam eşleme turu (2026-08-19): kart alanlarının kaynak sütunları
    "ana_tur":                   ("Ana Tür",),
    "durum":                     ("Durum",),
    "dava_konusu":               ("Dava Konusu",),
    "yerel_mahkeme":             ("Yerel Mahkeme",),
    "esas":                      ("Esas",),
    "dava_tarihi":               ("Dava Tarihi",),
    "is_kabul_tarihi":           ("İş Kabul Tarihi",),
    "dava_degeri":               ("Dava Değeri TL", "Dava Değeri"),
    "manevi_dava_degeri":        ("Manevi Dava Değeri TL", "Manevi Dava Değeri"),
    "hukmedilen_maddi":          ("Hükmedilen Maddi",),
    "hukmedilen_manevi":         ("Hükmedilen Manevi",),
    "hukmedilen_toplam":         ("Hükmedilen Toplam",),
    "son_durum":                 ("Son Durum",),
    "buro_ozel_turu":            ("Buro Özel Türü",),
    "hukuk_no":                  ("Hukuk No",),
    "arabuluculuk_no":           ("Arabuluculuk Numarası",),
    "arabuluculuk_karar_tarihi": ("Arabuluculuk Karar Tarihi",),
    "istinaf_basvuran":          ("İstinaf Mahkemesi Başvuran Taraf",),
    "uzmanlik_alani":            ("Dava Türü Alt Kırılımı", "Uzmanlık Alanı"),
    "avukatlar":                 ("Sorumlu Avukatlar", "Sorumlu Avukat"),
    "muvekkil":                  ("Müvekkil",),
    "karsi_taraf":               ("Karşı Taraf",),
    "sigortali":                 ("Sigortalı",),
    "davali_idare":              ("Davalı İdare",),
    "taraf_sifati":              ("Taraf Sıfatı",),
    "tibbi_surec":               ("Tıbbi Süreç",),
    "iddia_edilen_kusur":        ("İddia Edilen Kusur",),
    "hastada_olusan_zarar":      ("Hastada Oluşan Zarar",),
    "uygulanan_yontem":          ("Uygulanan Yöntem",),
    # --- belgeleme olayı alanları (G104, 25.08 belgesi §5): başlıklar karşı
    # tarafla yazılı olarak henüz SABİTLENMEDİ — aday desen toleranslıdır
    # (_baslik_anahtari aksan/büyük-küçük/boşluk farklarını zaten yutar) ve
    # başlık teslimde YOKSA alan "bu teslimde yok" sayılır: eski paketlerle
    # koşu davranış değiştirmez (None sözleşmesi).
    "olay_turu":                 ("Olay Türü",),
    "hukumdeki_rol":             ("Hükümdeki Rol",),
}

# Satırın kimliği: bu sütun yoksa dosya bu script için okunamaz.
ZORUNLU_SUTUNLAR = ("sistem_no",)

# Kardeş föylerde uyuşmazlığı raporlanan künye alanları (yazılmaz, okunur).
KUNYE_ALANLARI = ("karar_no", "karar_tarihi")

# D5 — yer tutucu değerler NULL'a çevrilir (metin biçimli tarihler dahil);
# tarih ve sayı yollarının ORTAK sözlüğü, "-" bir alanda yer tutucuysa
# diğerinde de yer tutucudur.
YER_TUTUCULAR = frozenset({"-", "--", "—", "?", "YOK", "BELİRSİZ", "BOŞ", "N/A", "NA"})
YER_TUTUCU_YIL = 1900

_TARIH_BICIMLERI = ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y")

# Çok değerli dosya numarası ayracı (`klasor_no_2`: "624.001.00;3.137.00").
# Numaraların KENDİSİ nokta içerir — ayraç olarak yalnız ';' ve satır sonu
# kabul edilir; virgül/nokta eklemek numaraları ortadan bölerdi.
_AYRAC = re.compile(r"[;\r\n]+")

# Kart alanlarının kolon sınırları modelden okunur, elle tekrarlanmaz
# (foy_map._LIMITS gerekçesi: şema büyürse kod kendiliğinden uyar).
_KART_LIMITLERI: Dict[str, int] = {}
for _kolon in models.Case.__table__.columns:
    _uzunluk = getattr(_kolon.type, "length", None)
    if _uzunluk:
        _KART_LIMITLERI[_kolon.name] = _uzunluk


class AktarimHatasi(Exception):
    """Koşunun tamamını durduran girdi hatası (dosya/sayfa/başlık)."""


class SatirHatasi(Exception):
    """TEK satırı düşüren VERİ hatası — parti devam eder, satır rapora düşer.

    İnsan müdahalesi gerektirir (düzeltme listesi konusu); koşuyu NONZERO
    yapar.
    """


class SatirAtlandi(SatirHatasi):
    """Satır BEKLENEN sebeple işlenmedi (kart yok) — koşuyu kırmızıya çekmez.

    Eşleştirme köprüsü %97,4; kalan satırların kartı henüz yok ve bu turda
    kart YARATILMIYOR. Beklenen bir sonucu "hata" saymak, gerçek hataları
    gürültüde boğardı — ama satır yine de rapora düşer.
    """


class AlanHatasi(Exception):
    """TEK alanın yazımını düşüren DEĞER hatası — satır işlenmeye DEVAM eder.

    `SatirHatasi`nın alan-düzeyi kardeşi (G104): kapalı listede karşılığı
    olmayan değer föyün DİĞER alanlarını düşürmez; yalnız bu alan yazılmaz,
    sebep satır raporuna HATA olarak düşer (tahmin yasağı — karşı tarafın
    düzeltme listesi konusudur, koşu NONZERO çıkar). SAVEPOINT gerekmez:
    dönüşüm yazımdan ÖNCE koşar, geri alınacak bir şey yoktur. Bilinçli olarak
    `SatirHatasi`ndan TÜREMEZ — türeseydi satır işleme/ön geçiş yolları bunu
    "satır düştü" sanırdı; `kart_degerleri` alan düzeyinde yakalar.
    """


@dataclass
class HamSatir:
    satir_no: int                      # xlsx'teki 1 tabanlı satır numarası
    degerler: Dict[str, Any]


@dataclass
class RaporSatiri:
    satir_no: int
    sistem_no: str
    dosya_no: str
    tur: str                           # HATA | ATLANDI
    sebep: str


@dataclass
class Celiski:
    kume: str                          # KART | TKU
    kume_anahtari: str
    alan: str
    degerler: str                      # "SSTMN-1=2018/143 | SSTMN-2=2016/768"


@dataclass
class AktarimSonucu:
    okunan: int = 0
    islenen: int = 0
    foy_yeni: int = 0
    foy_guncellenen: int = 0
    alan_degisikligi: int = 0
    bosaltilan: int = 0                # alan_degisikligi'nin alt kümesi (G112 açık boşaltma)
    avukat_eklenen: int = 0
    taraf_eklenen: int = 0
    asama_eklenen: int = 0
    onceki_esas_eklenen: int = 0
    havuz_disi_durum: int = 0
    atlanan: int = 0
    dry_run: bool = False
    yazildi: bool = False              # commit edildi mi?
    kaynak_imzasi: str = ""
    rapor_satirlari: List[RaporSatiri] = field(default_factory=list)
    celiskiler: List[Celiski] = field(default_factory=list)
    # KART kimlikleri — sayaç DEĞİL küme: bir kartın iki föyü de alan
    # değiştirirse bu "iki kart" değildir. Özet satırı "N kart" yazıyor;
    # sayaç sürümü 100 föylük provada 52 kartı "58 kart" diye raporladı.
    degisen_kartlar: Set[int] = field(default_factory=set)
    envanter_once: Optional[belge_envanteri.BelgeEnvanteri] = None
    envanter_sonra: Optional[belge_envanteri.BelgeEnvanteri] = None
    envanter_farki: Dict[str, Tuple[Any, Any]] = field(default_factory=dict)
    raporlar: List[Path] = field(default_factory=list)

    @property
    def kart_degisen(self) -> int:
        """Alanı değişen BENZERSİZ kart sayısı (özet satırının "N kart"ı)."""
        return len(self.degisen_kartlar)

    @property
    def hatalar(self) -> List[RaporSatiri]:
        return [r for r in self.rapor_satirlari if r.tur == "HATA"]

    @property
    def cikis_kodu(self) -> int:
        if self.envanter_farki:
            return CIKIS_ENVANTER
        return CIKIS_SATIR_HATASI if self.hatalar else CIKIS_TAMAM


# ═══════════════════════════════════════════════════════════════════════════
# Normalizasyon
# ═══════════════════════════════════════════════════════════════════════════

def _baslik_anahtari(deger: Any) -> str:
    """Başlık karşılaştırma anahtarı: TR büyük harf + aksan sadeleştirme +
    yalnız harf/rakam. "Arşiv Tarihi" ile "ARSIV TARIHI" aynı sütundur."""
    buyuk = tr_upper(str(deger or ""))
    ayrik = unicodedata.normalize("NFD", buyuk)
    sade = "".join(ch for ch in ayrik if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]", "", sade)


def _metin(deger: Any) -> Optional[str]:
    """Boşluk-normalize metin; boş hücre None ("bu teslimde yok")."""
    if deger is None:
        return None
    if isinstance(deger, (datetime, date)):
        return deger.isoformat()
    metin = " ".join(str(deger).split())
    return metin or None


def _kirp(deger: Optional[str], kolon: str) -> Optional[str]:
    """Kimlik OLMAYAN metni kolon sınırına kırpar (deneme-düzeyi → WARNING)."""
    if deger is None:
        return None
    sinir = _KART_LIMITLERI.get(kolon)
    if sinir and len(deger) > sinir:
        logger.warning(f"cases.{kolon} kolon sınırına kırpıldı (>{sinir}): {deger[:60]!r}…")
        return deger[:sinir]
    return deger


def _tarih(deger: Any, alan: str) -> Optional[date]:
    """D5 — yer tutucu tarihler NULL; çözümlenemeyen tarih SATIRI DÜŞÜRÜR.

    Yer tutucular (01.01.1900, metin biçimliler, gelecek tarihler) bilinen ve
    ölçülmüş sınıflardır; sessizce NULL'lanırlar. Çözümlenemeyen bir tarih ise
    bilinmeyen bir kusurdur: sessizce NULL'lamak onu karşı tarafın düzeltme
    listesinden gizlerdi.
    """
    if deger is None:
        return None
    if isinstance(deger, datetime):
        deger = deger.date()
    if isinstance(deger, date):
        return _tarih_suz(deger, alan)

    ham = " ".join(str(deger).split())
    if not ham:
        return None
    if tr_upper(ham) in YER_TUTUCULAR:
        logger.warning(f"{alan}: yer tutucu tarih NULL'landı ({ham!r})")
        return None
    for bicim in _TARIH_BICIMLERI:
        try:
            return _tarih_suz(datetime.strptime(ham, bicim).date(), alan)
        except ValueError:
            continue
    # Saat ekli METİN hücreler: pakette çoğu tarih hücresi gerçek tarih tipinde
    # gelir, ama METİN olarak yazılmış olanlar "2024-07-16 00:00:00" biçiminde
    # okunur (tam eşleme koşusunda 6 satır bu yüzden DÜŞMÜŞTÜ). Tarih kısmı
    # ayrıştırılabiliyorsa satırı düşürmek için sebep yok.
    try:
        return _tarih_suz(datetime.fromisoformat(ham).date(), alan)
    except ValueError:
        pass
    raise SatirHatasi(f"{alan} çözümlenemedi: {ham!r}")


def _tarih_suz(deger: date, alan: str) -> Optional[date]:
    """D5'in tarih-nesnesi tarafı: 1900 ve öncesi + gelecek tarih = yer tutucu."""
    if deger.year <= YER_TUTUCU_YIL:
        logger.warning(f"{alan}: yer tutucu tarih NULL'landı ({deger.isoformat()})")
        return None
    if deger > date.today():
        logger.warning(f"{alan}: gelecek tarih NULL'landı ({deger.isoformat()})")
        return None
    return deger


def _sayi(deger: Any, alan: str) -> Optional[Decimal]:
    """TR biçimli parasal değer → Decimal. Çözümlenemeyen SATIRI DÜŞÜRÜR.

    "12.345,67" (nokta binlik, virgül ondalık) ile "12345.67" birlikte
    yaşıyor; ayrım virgülün varlığıyla ve binlik kalıbıyla yapılır.
    """
    if deger is None:
        return None
    if isinstance(deger, bool):                       # openpyxl TRUE/FALSE hücresi
        raise SatirHatasi(f"{alan} sayı değil: {deger!r}")
    if isinstance(deger, (int, float, Decimal)):
        return _sayi_suz(Decimal(str(deger)), alan)

    ham = " ".join(str(deger).split())
    if not ham:
        return None
    temiz = ham.replace("₺", "").replace("TL", "").replace(" ", "").strip()
    if not temiz or tr_upper(temiz) in YER_TUTUCULAR:
        return None
    if "," in temiz:
        temiz = temiz.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", temiz):    # 12.345 → binlik ayracı
        temiz = temiz.replace(".", "")
    try:
        return _sayi_suz(Decimal(temiz), alan)
    except InvalidOperation:
        raise SatirHatasi(f"{alan} çözümlenemedi: {ham!r}") from None


def _sayi_suz(deger: Decimal, alan: str) -> Decimal:
    if deger < 0:
        raise SatirHatasi(f"{alan} negatif olamaz: {deger}")
    return deger


def _tarih_yumusak(deger: Any, alan: str) -> str:
    """Çelişki raporu için tarih anahtarı — ÇÖZÜLEMEZSE ham metin.

    Rapor "iki kardeş föy aynı şeyi mi söylüyor?" sorusunu sorar; çözümlenemeyen
    bir künye tarihi de karşılaştırılabilir bir cevaptır (ve satırı düşürmez —
    künye alanları YAZILMIYOR).
    """
    try:
        cozulen = _tarih(deger, alan)
    except SatirHatasi:
        return _metin(deger) or ""
    return cozulen.isoformat() if cozulen else ""


def _esleme(sozluk: Dict[str, str]) -> Callable[[Any, str], Optional[str]]:
    """Kapalı havuz eşlemesi — tanınmayan değer YAZILMAZ (None döner).

    Tahmin yasağının dönüştürücü hâli: teslimin serbest metnini bizim listemize
    zorlamak yerine, eşlemesi olmayanı boş bırakırız. Örnek: "İstinaf Başvuran
    Taraf" sütununda 17 farklı yazım var (DAVACI · DAVALI-DAVACI · SANIK
    MÜDAFİ · bir hastane adı); bizim kapalı listemiz üç değerli.
    """
    def donustur(deger: Any, alan: str) -> Optional[str]:
        metin = _metin(deger)
        if metin is None:
            return None
        # Anahtar başlık eşlemesiyle AYNI normalizasyondan geçer (TR büyük harf
        # + aksan sadeleştirme + yalnız harf/rakam): "İDARE"→IDARE,
        # "DAVACI/ DAVALI" ile "DAVACI-DAVALI" tek anahtarda buluşur.
        return sozluk.get(_baslik_anahtari(metin))
    return donustur


# Ana Tür → cases.file_type. Değerler DB'den doğrulandı (2026-08-19): 14.345
# kartın file_type havuzu Hukuk/İcra/Tahkim/İdare/Arabuluculuk/Ceza/Vergi/
# Danışmanlık/Savcılık. Teslimdeki 8 tür bunların dokuzuna birebir oturuyor.
ANA_TUR_ESLEMESI = {
    "HUKUK": "Hukuk", "IDARE": "İdare", "CEZA": "Ceza",
    "ARABULUCULUK": "Arabuluculuk", "ICRA": "İcra", "SAVCILIK": "Savcılık",
    "TAHKIM": "Tahkim", "DANISMANLIK": "Danışmanlık",
}

# Durum → cases.status. Bizim havuz iki değerli (DERDEST/MAHZEN), teslimin de.
DURUM_ESLEMESI = {"AKTIF": "DERDEST", "ARSIV": "MAHZEN"}

# İstinaf başvuran taraf → G044'ün kapalı listesi (Davacı/Davalı/Her İki Taraf).
# Birleşik yazımların hepsi "Her İki Taraf"tır; listede olmayan 8 satır
# (DİĞER DAVALI, SANIK MÜDAFİ, FERİ MÜDAHİL, bir hastane adı) YAZILMAZ.
ISTINAF_BASVURAN_ESLEMESI = {
    "DAVACI": "Davacı", "DAVALI": "Davalı",
    "DAVALIDAVACI": "Her İki Taraf", "DAVACIDAVALI": "Her İki Taraf",
}


def _metin_alan(deger: Any, alan: str) -> Optional[str]:
    return _kirp(_metin(deger), alan)


# ─── Avukat adı eşlemesi ─────────────────────────────────────────────────────
# Teslim avukat adlarını AKSANSIZ yazıyor ("Tugce Ungor Yanık"), üstelik
# "Sorumlu Avukatlar" bir LİSTEDİR (7.313 föyde tek isim, 1.085'inde çoklu);
# bizde `cases.responsible_lawyer_name` TEK kutu, çoklu iş `case_lawyers`
# tablosunun. Ölçüm (2026-08-19): teslimde 12 farklı ad var; 8'inin doğru
# yazımı zaten kartlarımızda ("Tuğçe Üngör Yanık"), 4'ü bizde hiç geçmiyor.
#
# Öncelik: (1) kartlardaki mevcut yazım — aksanlar doğru, (2) resmî `lawyers`
# listesi (BÜYÜK HARF saklanıyor → tr_title), (3) teslimin ham adı (tr_title).
# Harita koşu başında DB'den kurulur; dönüştürücüler saf kalsın diye modül
# düzeyinde yaşar ve koşu sonunda temizlenir.
_AVUKAT_HARITASI: Dict[str, str] = {}


def avukat_haritasi_kur(db) -> Dict[str, str]:
    """{normalize ad: gösterilecek yazım} — koşu başında BİR kez okunur."""
    _AVUKAT_HARITASI.clear()
    for (ad,) in db.query(models.Lawyer.name).filter(models.Lawyer.name.isnot(None)):
        if ad and ad.strip():
            _AVUKAT_HARITASI.setdefault(_baslik_anahtari(ad), tr_title(ad))
    # Kart yazımları resmî listeyi EZER: orada aksanlar korunmuş.
    sorgu = db.query(models.Case.responsible_lawyer_name).filter(
        models.Case.responsible_lawyer_name.isnot(None),
        models.Case.responsible_lawyer_name != "",
    ).distinct()
    for (ad,) in sorgu:
        temiz = (ad or "").strip()
        if temiz and ";" not in temiz:      # "A;B" birleşik kayıtlar ad değildir
            _AVUKAT_HARITASI[_baslik_anahtari(temiz)] = temiz
    return _AVUKAT_HARITASI


def _avukat_adlari(deger: Any) -> List[str]:
    """Teslimin virgüllü listesini bizim yazımımızla ad listesine çevirir."""
    ham = _metin(deger)
    if not ham:
        return []
    adlar: List[str] = []
    for parca in ham.split(","):
        temiz = " ".join(parca.split())
        if not temiz:
            continue
        ad = _AVUKAT_HARITASI.get(_baslik_anahtari(temiz)) or tr_title(temiz)
        if ad not in adlar:
            adlar.append(ad)
    return adlar


def _tek_avukat(deger: Any, alan: str) -> Optional[str]:
    """`responsible_lawyer_name` YALNIZ tek isimli föylerde yazılır.

    Çoklu föyde "sorumlu" hangisi belli değildir; ilkini seçmek uydurma olurdu
    (K1 kararının aynısı: yanlış veri boş veriden pahalıdır). Çoklu listenin
    tamamı `case_lawyers` satırlarına yazılır — bilgi kaybolmaz.
    """
    adlar = _avukat_adlari(deger)
    return _kirp(adlar[0], alan) if len(adlar) == 1 else None


def _baslik_bicimli(deger: Any, alan: str) -> Optional[str]:
    """Teslimin BÜYÜK HARF metnini bizim saklama biçimimize çevirir.

    Uzmanlık alanı teslimde "ÇOCUK SAĞLIĞI VE HASTALIKLARI" gibi gelir; bizim
    referans listelerimizin saklama formatı `tr_title` ("Çocuk Sağlığı Ve
    Hastalıkları"). Ham hâliyle yazmak 7.039 kartta yalnız yazımı bozardı —
    içerik zaten aynı. Eşleme SÖZLÜĞÜ değildir (77 ham değer ↔ 44 seed işi
    duruyor); yalnız biçim düzeltir.
    """
    metin = _metin(deger)
    return _kirp(tr_title(metin), alan) if metin else None


# ─── Belgeleme olayı alanları (G104) ─────────────────────────────────────────
# Değer eşlemesi AD bazlıdır ve KAPALI listeye karşı yapılır; adların tek
# doğruluk kaynağı G103 seed sabitleridir (`seed_data.EVENT_TYPES` /
# `JUDGMENT_ROLES`) — burada tekrarlanmaz, liste değişirse eşleme kendiliğinden
# uyar. `_esleme`den farkı RAPORLAMADIR: İstinaf eşlemesinde tanınmayan değer
# sessizce boş kalır (17 yazım ölçülüp bilinçli kabul edildi); bu iki alanda
# değer havuzu henüz ölçülmedi — tanınmayan değer `AlanHatasi` ile satır
# raporuna düşer, sessiz boşluk kusuru düzeltme listesinden gizlerdi.
_OLAY_TURU_ADLARI: Dict[str, str] = dict(seed_data.EVENT_TYPES)
OLAY_TURU_ESLEMESI: Dict[str, str] = {
    _baslik_anahtari(ad): ad for ad in _OLAY_TURU_ADLARI.values()
}
HUKUMDEKI_ROL_ESLEMESI: Dict[str, str] = {
    _baslik_anahtari(ad): ad for _kod, ad in seed_data.JUDGMENT_ROLES
}
# 31.08 teslim kuralı: ` ; ` ile birlikte gelen {Tıbbi Olay, Belgeleme Olayı}
# kapalı listenin KARMA değerine normalize edilir (kart alanı TEK SLOT; karma
# durum açık değerle taşınır, yazım birebir kapalı liste adıdır — G103).
_OLAY_TURU_KARMA_PARCALARI = frozenset(
    {_OLAY_TURU_ADLARI["TIBBI"], _OLAY_TURU_ADLARI["BELGELEME"]}
)


def _kapali_liste_parcalari(deger: Any, alan: str, harita: Dict[str, str]) -> List[str]:
    """Hücreyi ` ; ` ayracıyla parçalar, her parçayı kapalı liste ADINA çözer.

    Dönen liste mükerrersiz KANONİK adlardır (bizim yazımımız; karşılaştırma
    `_baslik_anahtari` ile — aksan/büyük-küçük/noktalama toleranslı). Yer
    tutucu parçalar D5'in ORTAK sözlüğüyle NULL sayılır (tarih yoluyla aynı
    davranış: "-" bir alanda yer tutucuysa diğerinde de öyledir). Tanınmayan
    parça `AlanHatasi` atar — hücrenin tanınan kısmını yazmak tahmin olurdu.
    """
    ham = _metin(deger)
    if ham is None:
        return []
    adlar: List[str] = []
    for parca in _AYRAC.split(ham):
        temiz = " ".join(parca.split())
        if not temiz:
            continue
        if tr_upper(temiz) in YER_TUTUCULAR:
            logger.warning(f"{alan}: yer tutucu değer NULL'landı ({temiz!r})")
            continue
        ad = harita.get(_baslik_anahtari(temiz))
        if ad is None:
            raise AlanHatasi(f"kapalı listede karşılığı yok: {temiz!r}")
        if ad not in adlar:
            adlar.append(ad)
    return adlar


def _olay_turu(deger: Any, alan: str) -> Optional[str]:
    """`cases.olay_turu` — tek değer ya da {Tıbbi, Belgeleme} → KARMA.

    Başka çok değerli kombinasyon (örn. tekil + KARMA) TANIMSIZDIR: normalize
    etmek tahmin olurdu; alan yazılmaz, satır rapora düşer. Aynı değerin
    mükerrer yazımı ("Tıbbi Olay ; Tıbbi Olay") tek değerdir — belirsizlik yok.
    """
    adlar = _kapali_liste_parcalari(deger, alan, OLAY_TURU_ESLEMESI)
    if not adlar:
        return None
    if len(adlar) == 1:
        return adlar[0]
    if set(adlar) == _OLAY_TURU_KARMA_PARCALARI:
        return _OLAY_TURU_ADLARI["KARMA"]
    raise AlanHatasi(f"çok değerli hücre tanımsız kombinasyon: {' ; '.join(adlar)}")


def _hukumdeki_rol(deger: Any, alan: str) -> Optional[str]:
    """`cases.hukumdeki_rol` — çok değer TANIMSIZDIR (SÖZLEŞME, 31.08).

    Rol, belgeleme olgusunun GÜNCEL kademedeki hükümde oynadığı roldür; iki
    rol aynı anda oynanmaz — çok değerli hücre yazılmaz, satır rapora düşer.
    """
    adlar = _kapali_liste_parcalari(deger, alan, HUKUMDEKI_ROL_ESLEMESI)
    if not adlar:
        return None
    if len(adlar) > 1:
        raise AlanHatasi(f"çok değerli hücre tanımsız: {' ; '.join(adlar)}")
    return adlar[0]


# Kart alanı → (kaynak sütun anahtarı, dönüştürücü)
KART_ALANLARI: Dict[str, Tuple[str, Callable[[Any, str], Any]]] = {
    # --- kimlik/sınıflandırma
    "file_type":    ("ana_tur", _esleme(ANA_TUR_ESLEMESI)),
    "status":       ("durum", _esleme(DURUM_ESLEMESI)),
    "subject":      ("dava_konusu", _metin_alan),
    "court":        ("yerel_mahkeme", _metin_alan),          # İÇERİK modu — aşağı bak
    "sub_type":     ("uzmanlik_alani", _baslik_bicimli),     # İÇERİK modu
    "responsible_lawyer_name": ("avukatlar", _tek_avukat),  # yalnız tek isimli föy
    "esas_no":      ("esas", _metin_alan),                   # ÖZEL: esas tarihçesi
    # --- tarihler
    "opening_date":    ("dava_tarihi", _tarih),
    "acceptance_date": ("is_kabul_tarihi", _tarih),
    "arsiv_tarihi":    ("arsiv_tarihi", _tarih),
    # --- para
    "islah_tutari":     ("islah_tutari", _sayi),
    "manevi_tazminat":  ("manevi_dava_degeri", _sayi),
    "hukmedilen_maddi":  ("hukmedilen_maddi", _sayi),
    "hukmedilen_manevi": ("hukmedilen_manevi", _sayi),
    "hukmedilen_toplam": ("hukmedilen_toplam", _sayi),
    # --- takip/dosya numaraları
    # DB'de bir de `last_status` kolonu var ama MODELDE YOK (eski kalıntı);
    # takip panelinin yazdığı alan `dosya_son_durumu`dur (14.220 kartta dolu).
    "dosya_son_durumu": ("son_durum", _metin_alan),
    "bureau_type":    ("buro_ozel_turu", _metin_alan),
    "hasar_dosya_no": ("hasar_no", _metin_alan),
    "hukuk_no":       ("hukuk_no", _metin_alan),
    # --- süreç
    # `istinaf_basvuran_taraf` BURADA YOK: `_PHOTO_COLUMNS["ISTINAF"]`in
    # kolonu, yani tek yazma yolu aşama fotoğrafı (G062). Kart tarafından da
    # yazılsaydı iki yazıcı olurdu — ölçüldü: aşama fotoğrafı boş basvuran_taraf
    # ile kolonu siliyor, kart yolu 68 sütundan geri yazıyordu; ikinci koşu
    # 5 kartta salınıyordu. 68 sütundaki değer artık AŞAMA SATIRINA besleniyor.
    "arabuluculuk_no":           ("arabuluculuk_no", _metin_alan),
    "arabuluculuk_karar_tarihi": ("arabuluculuk_karar_tarihi", _tarih),
    # --- tıbbi beşli (G044)
    "tibbi_surec":          ("tibbi_surec", _metin_alan),
    "tibbi_olay":           ("tibbi_olay", _metin_alan),
    "iddia_edilen_kusur":   ("iddia_edilen_kusur", _metin_alan),
    "hastada_olusan_zarar": ("hastada_olusan_zarar", _metin_alan),
    "uygulanan_yontem":     ("uygulanan_yontem", _metin_alan),
    # --- belgeleme olayı alanları (G104): kapalı liste ADLARI (G103 şeması).
    # Dolu hücre kuralı METİN alanlarıyla aynı sınıf (varsayılan üzerine
    # yazma); İÇERİK modu GEREKSİZ — dönüştürücü çıktısı zaten kanonik ad,
    # yalnız-yazım farkı bu alanlarda oluşamaz.
    "olay_turu":            ("olay_turu", _olay_turu),
    "hukumdeki_rol":        ("hukumdeki_rol", _hukumdeki_rol),
}

# İÇERİK farkı varsa yazılan, YAZIM farkı varsa dokunulmayan alanlar.
#
# Ölçüm (2026-08-19, 7.932 eşleşen föy): `court`ta 562 farkın 480'i yalnız
# yazım (BÜYÜK HARF / eksik nokta), 82'si GERÇEK başka mahkeme ("İzmir 4.
# İdare" ↔ "İzmir 15. Asliye Hukuk"). `sub_type`ta 7.390 farkın 7.039'u yazım,
# 351'i gerçekten başka uzmanlık ("Göğüs Cerrahisi" ↔ "Genel Cerrahi").
#
# Kural: içerik teslimin (kaynak orada, bilgi daha güncel), yazım bizim
# (G067-G070 mahkeme adı kimliği + referans listelerinin `tr_title` formatı).
# Karşılaştırma `_baslik_anahtari` ile: aksan, büyük/küçük harf ve noktalama
# yok sayılır.
ICERIK_KARSILASTIRMALI_ALANLAR = frozenset({"court", "sub_type"})

# Türetilen alanlar: değeri TEK sütundan gelmeyenler.
KART_TURETILEN: Dict[str, Callable[[Dict[str, Any], str], Any]] = {}


def _maddi_tazminat(degerler: Dict[str, Any], alan: str) -> Optional[Decimal]:
    """D4 — maddi = Dava Değeri − Manevi; NEGATİFSE hesaplama yapılmaz.

    Şartname §2/D4: 98 satırda manevi > dava değeri; orada maddi NULL kalır
    (uydurma sayı üretmek yerine boş). Manevi = Dava Değeri olan satırlarda
    maddi = 0 DOĞRUDUR ve yazılır (NULL ≠ 0 kuralı).
    """
    dava_degeri = _sayi(degerler.get("dava_degeri"), "dava_degeri")
    if dava_degeri is None:
        return None
    manevi = _sayi(degerler.get("manevi_dava_degeri"), "manevi_dava_degeri") or Decimal(0)
    fark = dava_degeri - manevi
    return None if fark < 0 else fark


KART_TURETILEN["maddi_tazminat"] = _maddi_tazminat


# ═══════════════════════════════════════════════════════════════════════════
# Girdi okuma
# ═══════════════════════════════════════════════════════════════════════════

def xlsx_oku(yol: Path, *, sheet: Optional[str] = None,
             limit: Optional[int] = None) -> Tuple[List[HamSatir], Dict[str, str]]:
    """Teslim paketini okur → (satırlar, {alan: bulunan başlık}).

    Gerçek teslim paketi REPOYA GİRMEZ (A.2 dersi: gerçek müvekkil verisi
    OneDrive/repo dışında kalır) — dosya yalnız çalışma zamanında `--input`
    ile okunur; testler openpyxl ile SENTETİK mini paket üretir.
    """
    import openpyxl

    if not yol.exists():
        raise AktarimHatasi(f"Girdi dosyası yok: {yol}")

    wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    try:
        if sheet is not None and sheet not in wb.sheetnames:
            raise AktarimHatasi(f"Sayfa yok: {sheet!r} (mevcut: {', '.join(wb.sheetnames)})")
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        satir_akisi = ws.iter_rows(values_only=True)
        baslik_satiri = next(satir_akisi, None)
        if baslik_satiri is None:
            raise AktarimHatasi(f"Girdi boş: {yol}")

        indeksler, bulunanlar = _sutun_indeksleri(baslik_satiri)
        eksik = [a for a in ZORUNLU_SUTUNLAR if a not in indeksler]
        if eksik:
            raise AktarimHatasi(
                f"Zorunlu sütun(lar) bulunamadı: {', '.join(eksik)} — "
                f"okunan başlıklar: {', '.join(str(b) for b in baslik_satiri if b)}"
            )

        satirlar: List[HamSatir] = []
        for sira, ham in enumerate(satir_akisi, start=2):
            if limit is not None and len(satirlar) >= limit:
                break                         # sınır ÖNCE bakılır: --limit 0 = hiç satır
            if ham is None or all(_metin(h) is None for h in ham):
                continue                      # tamamen boş satır: sessizce atla
            satirlar.append(HamSatir(
                satir_no=sira,
                degerler={
                    alan: (ham[i] if i < len(ham) else None)
                    for alan, i in indeksler.items()
                },
            ))
        return satirlar, bulunanlar
    finally:
        wb.close()


def _sutun_indeksleri(baslik_satiri: Sequence[Any]) -> Tuple[Dict[str, int], Dict[str, str]]:
    """{alan: sütun indeksi} + {alan: dosyada bulunan başlık}."""
    dosyadaki = {}
    for i, ham in enumerate(baslik_satiri):
        anahtar = _baslik_anahtari(ham)
        if anahtar and anahtar not in dosyadaki:
            dosyadaki[anahtar] = (i, str(ham).strip())

    indeksler: Dict[str, int] = {}
    bulunanlar: Dict[str, str] = {}
    for alan, adaylar in SUTUN_ADAYLARI.items():
        for aday in adaylar:
            eslesme = dosyadaki.get(_baslik_anahtari(aday))
            if eslesme:
                indeksler[alan], bulunanlar[alan] = eslesme[0], eslesme[1]
                break
    return indeksler, bulunanlar


# ═══════════════════════════════════════════════════════════════════════════
# Kart eşleştirme
# ═══════════════════════════════════════════════════════════════════════════

def _dosya_no_haritasi(db) -> Dict[str, List[int]]:
    """{normalize dosya no parçası: [kart id]} — köprü TEK sorguda okunur.

    Satır başına SELECT yapmak 8.409 satırda tam tarama demekti (`klasor_no_2`
    trigram index'i G042'de düştü). Mükerrer anahtar OLAĞANDIR (14.317 dolu /
    14.204 distinct → 112 grup); liste hâlinde tutulur, çok eşleşen satır
    "belirsiz eşleşme" ile rapora düşer — yanlış karta yazmaktansa.
    """
    harita: Dict[str, List[int]] = {}
    sorgu = (
        db.query(models.Case.id, models.Case.klasor_no_2)
        .filter(models.Case.klasor_no_2.isnot(None), models.Case.deleted_at.is_(None))
        .order_by(models.Case.id)
    )
    for case_id, klasor in sorgu.yield_per(1000):
        for parca in _dosya_no_parcalari(klasor):
            harita.setdefault(parca, []).append(case_id)
    return harita


def _dosya_no_parcalari(deger: Any) -> List[str]:
    """Çok değerli dosya numarasını parçalara ayırır (mükerrersiz, sıralı).

    `klasor_no_2` ÇOK DEĞERLİDİR: eski dosya numaraları ';' ile birleşik
    tutulur ("624.001.00;3.137.00"). Lokal prod kopyasında ölçüm (2026-08-19):
    14.317 dolu kaydın **1.267'si** çok değerli — ve bunlar tam da birleşik
    kartlar, yani bu görevin çekirdek vakası (1.211 kart 2+ föyü birleşik
    taşıyor). Tam-değer eşleşmesi o kartların hepsini ISKALARDI. Aynı ayrıştırma
    teslim tarafına da uygulanır (simetri): paket de birleşik yazabilir.
    """
    parcalar: List[str] = []
    for ham in _AYRAC.split(_metin(deger) or ""):
        anahtar = _eslesme_anahtari(ham)
        if anahtar and anahtar not in parcalar:
            parcalar.append(anahtar)
    return parcalar


def _eslesme_anahtari(deger: Any) -> str:
    """Dosya No ↔ klasor_no_2 karşılaştırma anahtarı (boşluk + harf duyarsız)."""
    return tr_upper(_metin(deger) or "")


# ═══════════════════════════════════════════════════════════════════════════
# Düzeltme_Logu (G112) — gerekçe provenance + açık boşaltma talimatı
# ═══════════════════════════════════════════════════════════════════════════
# Veri ekibinin hücre değişiklik günlüğü: her satır bir (SistemNo, sütun)
# düzeltmesidir — eski/yeni değer, GEREKÇE ve tarih. Bugüne dek okunmuyordu;
# `case_history`ye yalnız "eski→yeni" düşüyordu. Sütun adı → kart alanı
# çevirisi KART_ALANLARI'nın ters haritasıdır (yeni sözlük İCAT EDİLMEZ):
# bilinmeyen sütun adı DEBUG ile yok sayılır — sayfa 20k satır, rapora
# düşürmek gürültü olurdu.
DUZELTME_SAYFASI = "Düzeltme_Logu"
#: Yeni Değer hücresinde AÇIK boşaltma işareti (üçlü şartın birinci ayağı).
BOSALTMA_ISARETI = "(boş)"

DUZELTME_SUTUNLARI: Dict[str, Tuple[str, ...]] = {
    "sistem_no": ("SistemNo", "Sistem No"),
    # Değişen sütunun ADI — karşı tarafla başlık henüz yazılı sabitlenmedi
    # (G114 sözleşmesi); aday desen toleranslı, `_baslik_anahtari` yazım
    # farklarını yutar.
    "sutun":     ("Sütun", "Sütun Adı", "Alan", "Alan Adı", "Kolon", "Değişen Sütun", "Değişen Alan"),
    "eski":      ("Eski Değer",),
    "yeni":      ("Yeni Değer",),
    "gerekce":   ("Gerekçe",),
    "tarih":     ("Tarih",),
}
_DUZELTME_ZORUNLU = ("sistem_no", "sutun", "yeni")
# "Tarih" hücresi saatli metin de gelebilir ("02.08.2026 14:05"); sıralama için.
_DUZELTME_TARIH_BICIMLERI: Tuple[str, ...] = tuple(
    f"{b}{ek}" for b in _TARIH_BICIMLERI for ek in ("", " %H:%M", " %H:%M:%S")
)

# Künye/aşama kolonları: `Düzeltme_Logu` bunları ADLANDIRABİLİR ama tek yazma
# yolu stage_decisions'ın fotoğrafıdır (G062) — boşaltma talimatı UYGULANMAZ,
# satır raporuna düşer (tahmin yasağının boşaltma tarafı: sessizce yutulmaz).
_DUZELTME_KUNYE_KAYNAKLARI: Dict[str, str] = {
    "karar_no": "karar_no",
    "karar_tarihi": "karar_tarihi",
    "istinaf_basvuran": "istinaf_basvuran_taraf",
}


def _duzeltme_alan_haritasi() -> Dict[str, str]:
    """{başlık anahtarı: kart alanı} — KART_ALANLARI'nın SUTUN_ADAYLARI üzerinden ters haritası."""
    harita: Dict[str, str] = {}
    for alan, (kaynak, _donustur) in KART_ALANLARI.items():
        for aday in SUTUN_ADAYLARI.get(kaynak, ()):
            harita.setdefault(_baslik_anahtari(aday), alan)
    for kaynak, alan in _DUZELTME_KUNYE_KAYNAKLARI.items():
        for aday in SUTUN_ADAYLARI.get(kaynak, ()):
            harita.setdefault(_baslik_anahtari(aday), alan)
    return harita


DUZELTME_ALAN_HARITASI: Dict[str, str] = _duzeltme_alan_haritasi()
#: Boşaltılması YASAK alanlar: künye (tek yazıcı stage_decisions) + içerik modu.
BOSALTMA_DISI_ALANLAR: frozenset = frozenset(_DUZELTME_KUNYE_KAYNAKLARI.values()) | ICERIK_KARSILASTIRMALI_ALANLAR

# `case_history.source` sınırı modelden okunur (imza + gerekçe buraya sığmalı).
_SOURCE_SINIRI: int = models.CaseHistory.source.property.columns[0].type.length or 300


@dataclass
class DuzeltmeKaydi:
    satir_no: int                      # Düzeltme_Logu'ndaki 1 tabanlı satır
    sistem_no: str
    alan: str                          # kart alanı (KART_ALANLARI anahtarı) ya da künye adı
    yeni: Optional[str]                # ham "Yeni Değer" metni
    bosalt: bool                       # Yeni Değer == (boş)
    gerekce: Optional[str]
    tarih: Optional[datetime]


DuzeltmeHaritasi = Dict[Tuple[str, str], DuzeltmeKaydi]


def _bosalt_isareti_mi(yeni: Optional[str]) -> bool:
    """Yeni Değer `(boş)` mu? Parantez ANLAMLIDIR: çıplak "boş" bir metin değeridir,
    boşaltma talimatı değil (`_baslik_anahtari` noktalamayı attığı için burada kullanılmaz);
    büyük/küçük harf ve iç boşluk toleranslı ("( BOŞ )")."""
    if not yeni:
        return False
    return re.sub(r"\s+", "", tr_upper(yeni)) == re.sub(r"\s+", "", tr_upper(BOSALTMA_ISARETI))


def _duzeltme_tarihi(deger: Any) -> Optional[datetime]:
    """"Tarih" hücresi → datetime; çözülemezse None (en eski sayılır, uyarı yok)."""
    if deger is None:
        return None
    if isinstance(deger, datetime):
        return deger
    if isinstance(deger, date):
        return datetime(deger.year, deger.month, deger.day)
    ham = " ".join(str(deger).split())
    if not ham:
        return None
    try:
        return datetime.fromisoformat(ham)
    except ValueError:
        pass
    for bicim in _DUZELTME_TARIH_BICIMLERI:
        try:
            return datetime.strptime(ham, bicim)
        except ValueError:
            continue
    return None


def duzeltme_logunu_oku(yol: Path, *, sheet: str = DUZELTME_SAYFASI) -> DuzeltmeHaritasi:
    """`Düzeltme_Logu` sayfasını okur → {(SistemNo, kart alanı): en yeni kayıt}.

    Sayfa yoksa BOŞ sözlük (hata değil — eski paketlerde yok). Zorunlu başlık
    (SistemNo / sütun adı / Yeni Değer) yoksa sayfa uygulanamaz: TEK WARNING +
    boş (provenance yan üründür, koşuyu durdurmaz; boşaltma talimatı da
    okunamadığı için uygulanmaz — güvenli taraf). Bilinmeyen sütun adı DEBUG
    ile yok sayılır. Aynı (SistemNo, alan) için birden çok satır → en yeni
    `Tarih` kazanır (tarihsiz satır en eski sayılır; eşitlikte sonraki satır).
    """
    import openpyxl

    wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            logger.info(f"{sheet!r} sayfası yok — düzeltme logu okunmadı")
            return {}
        ws = wb[sheet]
        akis = ws.iter_rows(values_only=True)
        baslik = next(akis, None)
        if baslik is None:
            return {}
        dosyadaki: Dict[str, int] = {}
        for i, ham in enumerate(baslik):
            anahtar = _baslik_anahtari(ham)
            if anahtar and anahtar not in dosyadaki:
                dosyadaki[anahtar] = i
        indeksler: Dict[str, int] = {}
        for alan, adaylar in DUZELTME_SUTUNLARI.items():
            for aday in adaylar:
                if _baslik_anahtari(aday) in dosyadaki:
                    indeksler[alan] = dosyadaki[_baslik_anahtari(aday)]
                    break
        eksik = [a for a in _DUZELTME_ZORUNLU if a not in indeksler]
        if eksik:
            logger.warning(
                f"{sheet}: zorunlu başlık(lar) yok ({', '.join(eksik)}) — sayfa okunmadı; "
                f"okunan başlıklar: {', '.join(str(b) for b in baslik if b)}"
            )
            return {}

        def hucre(ham: Sequence[Any], alan: str) -> Any:
            i = indeksler.get(alan)
            return ham[i] if i is not None and i < len(ham) else None

        kayitlar: DuzeltmeHaritasi = {}
        bilinmeyen = 0
        for sira, ham in enumerate(akis, start=2):
            if ham is None:
                continue
            sistem_no = _metin(hucre(ham, "sistem_no"))
            sutun = _metin(hucre(ham, "sutun"))
            if not sistem_no or not sutun:
                continue
            alan = DUZELTME_ALAN_HARITASI.get(_baslik_anahtari(sutun))
            if alan is None:
                bilinmeyen += 1
                logger.debug(f"{sheet} satır {sira}: bilinmeyen sütun adı {sutun!r} — yok sayıldı")
                continue
            yeni = _metin(hucre(ham, "yeni"))
            kayit = DuzeltmeKaydi(
                satir_no=sira, sistem_no=sistem_no, alan=alan, yeni=yeni,
                bosalt=_bosalt_isareti_mi(yeni),
                gerekce=_metin(hucre(ham, "gerekce")),
                tarih=_duzeltme_tarihi(hucre(ham, "tarih")),
            )
            mevcut = kayitlar.get((sistem_no, alan))
            if mevcut is None or (kayit.tarih or datetime.min) >= (mevcut.tarih or datetime.min):
                kayitlar[(sistem_no, alan)] = kayit
        logger.info(
            f"{sheet}: {len(kayitlar)} (SistemNo, alan) kaydı okundu"
            f"{f', {bilinmeyen} satır bilinmeyen sütun adıyla yok sayıldı' if bilinmeyen else ''}"
        )
        return kayitlar
    finally:
        wb.close()


def _bosaltma_talimatlari(
    satir: HamSatir, sistem_no: str, duzeltmeler: Optional[DuzeltmeHaritasi],
) -> Tuple[List[Tuple[str, DuzeltmeKaydi]], List[Tuple[str, str]]]:
    """Föyün `(boş)` talimatlarını ikiye ayırır: (uygulanabilir, reddedilen).

    Uygulanabilir = üçlü şartın ilk iki ayağı: log `(boş)` diyor VE `Sheet`'te
    o sütun VAR ve hücre GERÇEKTEN boş. Sütunun paketten büsbütün eksik olması
    "hücre boş" DEĞİLDİR (partili teslim: `Sheet` teyit edemiyor → boşaltma
    yok). Üçüncü ayak (bizde dolu) yazma anında karta bakılarak ölçülür.
    Reddedilen = künye ya da içerik-modu alanı: talimat uygulanmaz, sebep
    satır raporuna düşer (`(alan, sebep)`).
    """
    if not duzeltmeler:
        return [], []
    uygulanabilir: List[Tuple[str, DuzeltmeKaydi]] = []
    reddedilen: List[Tuple[str, str]] = []
    for (s_no, alan), kayit in duzeltmeler.items():
        if s_no != sistem_no or not kayit.bosalt:
            continue
        if alan in BOSALTMA_DISI_ALANLAR:
            neden = ("karar künyesi tek yazıcı stage_decisions" if alan not in ICERIK_KARSILASTIRMALI_ALANLAR
                     else "içerik-karşılaştırmalı alan (yazım bizim)")
            reddedilen.append((alan, f"boşaltılmadı — {neden} (Düzeltme_Logu satır {kayit.satir_no})"))
            continue
        kaynak = KART_ALANLARI[alan][0]
        if kaynak not in satir.degerler or _metin(satir.degerler.get(kaynak)) is not None:
            logger.debug(
                f"{sistem_no} {alan}: Düzeltme_Logu (boş) diyor ama Sheet hücresi "
                f"{'yok' if kaynak not in satir.degerler else 'dolu'} — boşaltma yok"
            )
            continue
        uygulanabilir.append((alan, kayit))
    return uygulanabilir, reddedilen


def _provenance_imzasi(source: str, kayit: Optional[DuzeltmeKaydi], *, bosaltma: bool = False) -> str:
    """`case_history.source`: imza (`HUKDOK_TESLIM_*`, KORUNUR) + Düzeltme_Logu gerekçesi.

    Gerekçesiz değişiklikte imza eski biçimiyle kalır (log satırı yoksa hiçbir
    ek yok). `(boş)` satırının gerekçesi yalnız boşaltma kaydına eklenir; aynı
    satır bir DEĞER yazımına gerekçe olamaz (çelişkili talimat — imza sade kalır).
    """
    if kayit is None:
        return source
    if bosaltma:
        ek = f"boşaltıldı: {kayit.gerekce or '-'}"
    elif kayit.bosalt or not kayit.gerekce:
        return source
    else:
        ek = f"gerekçe: {kayit.gerekce}"
    return f"{source} · {ek}"[:_SOURCE_SINIRI]


# ═══════════════════════════════════════════════════════════════════════════
# Satır işleme
# ═══════════════════════════════════════════════════════════════════════════

def kart_degerleri(satir: HamSatir,
                   atlanan_alanlar: Optional[List[Tuple[str, str]]] = None) -> Dict[str, Any]:
    """Satırın kart alanlarına çevrilmiş NİHAİ değerleri — TEK dönüşüm noktası.

    Hem yazma yolu hem çelişki ön geçişi buradan okur: ikisi farklı yerde
    dönüştürseydi "çelişkili" sayılan değerle yazılan değer ayrışabilirdi.
    `None` (bu teslimde yok) sözlüğe GİRMEZ — "boşalt" anlamı yok.

    `AlanHatasi` atan dönüştürücü (kapalı liste alanları, G104) satırı
    DÜŞÜRMEZ: alan sözlüğe girmez, `(alan, sebep)` çifti `atlanan_alanlar`a
    eklenir (verilmişse) ve yazma yolu satır raporuna düşürür. Çelişki ön
    geçişi toplamadan çağırır — aynı kusur rapora İKİ kez düşmesin.
    """
    degerler: Dict[str, Any] = {}
    for alan, (kaynak, donustur) in KART_ALANLARI.items():
        try:
            deger = donustur(satir.degerler.get(kaynak), alan)
        except AlanHatasi as exc:
            if atlanan_alanlar is not None:
                atlanan_alanlar.append((alan, str(exc)))
            continue
        if deger is not None:
            degerler[alan] = deger
    for alan, hesapla in KART_TURETILEN.items():
        deger = hesapla(satir.degerler, alan)
        if deger is not None:
            degerler[alan] = deger
    return degerler


def _kart_id_tahmini(db, satir: HamSatir, foy_haritasi: Dict[str, int],
                     dosya_haritasi: Dict[str, List[int]],
                     sistem_no: str) -> Optional[int]:
    """`_kart_coz`un ÖN GEÇİŞ ikizi: istisna atmaz, None döner.

    İKİNCİ ANAHTARI DA UYGULAR — yoksa asıl döngü satırı bir karta yazarken ön
    geçiş o satırı hiç saymaz ve "kardeş föyler uzlaşıyor mu" sorusu eksik
    veriyle cevaplanırdı.
    """
    case_id = foy_haritasi.get(sistem_no)
    if case_id is not None:
        return case_id
    adaylar: List[int] = []
    for parca in _dosya_no_parcalari(satir.degerler.get("dosya_no")):
        for aday in dosya_haritasi.get(parca) or []:
            if aday not in adaylar:
                adaylar.append(aday)
    if len(adaylar) == 1:
        return adaylar[0]
    return _ikinci_anahtarla_coz(db, satir, adaylar) if adaylar else None


def kart_alan_celiskileri(
    db, satirlar: Sequence[HamSatir], *, foy_haritasi: Dict[str, int],
    dosya_haritasi: Dict[str, List[int]],
    duzeltmeler: Optional[DuzeltmeHaritasi] = None,
) -> Tuple[Dict[int, Set[str]], List[Celiski]]:
    """Aynı kartın föyleri bir KART alanında çelişiyorsa o alan YAZILMAZ.

    **Neden (19.08 provasının bulgusu):** kart alanları TEK SLOT'tur, föy ise
    kart başına birden çok. Satır satır yazınca kartta kalan değer "en son
    işlenen föy"ünkidir; ikinci koşuda satır sırası değişince başka föy
    kazanır ve alan koşudan koşuya SALINIR. 100 föylük provada kart#195
    (12 föy; 10'u 2023-07-12, id-9902 2017-06-07, id-9908 2017-07-12) ve
    kart#12954 yüzünden 2. ve 3. koşu her seferinde 6 "değişiklik" üretti —
    oysa ne bizde ne teslimde bir şey değişmişti. Bu, G064'ün kabul kriterini
    ("aynı girdiyle ikinci koşu 0 değişiklik") doğrudan ihlal ediyordu.

    Çözüm tahmin ETMEK değil, tahmini REDDETMEK: çelişen alan yazılmaz, çelişki
    kardeş-föy raporuna düşer (KUNYE_ALANLARI ile aynı desen — künye de
    yazılmaz, raporlanır). Föyler arası gerçekten farklı olan değerlerin kalıcı
    evi FAZ F §1.5'tir (`case_parties`/föy düzeyi alanlar); o ev açılana kadar
    doğru davranış, kartta hangi föyün kazandığını kur'aya bırakmamaktır.

    Açık boşaltma (G112) da bir DEĞERDİR: bir föy `(boş)` derken kardeşi değer
    taşıyorsa alan çelişkilidir — aksi hâlde satır sırasına göre kart bir
    koşuda boşalır, diğerinde dolar (idempotency ihlali). `BOSALTMA_ISARETI`
    uzlaşı kümesine üye olarak girer.

    Döner: ({case_id: {çelişen alan}}, [rapor satırı]).
    """
    degerler: Dict[Tuple[int, str], List[Tuple[str, Any]]] = {}
    for satir in satirlar:
        sistem_no = _metin(satir.degerler.get("sistem_no"))
        if not sistem_no:
            continue
        case_id = _kart_id_tahmini(db, satir, foy_haritasi, dosya_haritasi, sistem_no)
        if case_id is None:
            continue
        try:
            satir_degerleri = kart_degerleri(satir)
        except SatirHatasi:
            continue              # bozuk satır zaten düşecek; uzlaşıyı kirletmesin
        for alan, deger in satir_degerleri.items():
            degerler.setdefault((case_id, alan), []).append((sistem_no, deger))
        for alan, _kayit in _bosaltma_talimatlari(satir, sistem_no, duzeltmeler)[0]:
            if alan not in satir_degerleri:
                degerler.setdefault((case_id, alan), []).append((sistem_no, BOSALTMA_ISARETI))

    celiskili: Dict[int, Set[str]] = {}
    celiskiler: List[Celiski] = []
    for (case_id, alan), uyeler in sorted(degerler.items()):
        if len({d for _, d in uyeler}) < 2:
            continue
        celiskili.setdefault(case_id, set()).add(alan)
        celiskiler.append(Celiski(
            kume="KART", kume_anahtari=str(case_id), alan=alan,
            degerler=" | ".join(f"{s}={d}" for s, d in sorted(uyeler)),
        ))
    return celiskili, celiskiler


def _esas_uyuyor(case: models.Case, esas: Optional[str]) -> bool:
    """Kartın esası föyünkiyle aynı mı? (`;` ile birleşik yazılmışlar dahil)

    Bizdeki 672 kayıtta esas kirli ("2021/588;2026/4", "2025768;2025/768") —
    tam-değer karşılaştırması bu kartları eşleştiremezdi.
    """
    if not esas or not case.esas_no:
        return False
    anahtar = _baslik_anahtari(esas)
    return any(_baslik_anahtari(p) == anahtar for p in _AYRAC.split(case.esas_no))


def _ikinci_anahtarla_coz(db, satir: HamSatir, adaylar: List[int]) -> Optional[int]:
    """Dosya No birden çok karta düşünce ESAS + DOSYA TÜRÜ ile ayırır.

    Belirsizliğin tipik hâli bizim ikiz kartlarımızdır: aynı klasör numarası
    hem dava hem arabuluculuk kartında yazılı (örn. `1056.003.00` → HUKUK
    kartı + ARABU kartı). Föyün Ana Tür'ü ve esası hangisini kastettiğini
    söyler. Ölçüm (2026-08-19, 257 belirsiz satır): esas+tür 179'unu, yalnız
    esas 12'sini, yalnız tür 32'sini ayırıyor; 34'ü insanda kalıyor.

    Sıra ÖNEMLİ: önce ikisi birden, sonra esas, sonra tür. Tek bir kriterle
    "tek aday kaldı" demek, diğer kriterin çeliştiği bir kartı seçmek olabilir.
    Hiçbiri tek adaya inmiyorsa None döner — satır rapora düşer, tahmin YOK.
    """
    kartlar = [k for k in (db.get(models.Case, aday) for aday in adaylar)
               if k is not None and k.deleted_at is None]
    if len(kartlar) == 1:
        return cast(int, kartlar[0].id)
    if not kartlar:
        return None

    esas = _metin(satir.degerler.get("esas"))
    tur = _esleme(ANA_TUR_ESLEMESI)(satir.degerler.get("ana_tur"), "file_type")
    esas_uyan = [k for k in kartlar if _esas_uyuyor(k, esas)]
    tur_uyan = [k for k in kartlar if tur and k.file_type == tur]

    for aday_kume in ([k for k in esas_uyan if k in tur_uyan], esas_uyan, tur_uyan):
        if len(aday_kume) == 1:
            secilen = aday_kume[0]
            logger.info(
                f"Belirsiz eşleşme ikinci anahtarla çözüldü: "
                f"{_metin(satir.degerler.get('sistem_no'))} → kart {secilen.id} "
                f"(esas={esas!r}, tür={tur!r})"
            )
            return cast(int, secilen.id)
    return None


def _kart_coz(db, satir: HamSatir, foy_haritasi: Dict[str, int],
              dosya_haritasi: Dict[str, List[int]], sistem_no: str) -> models.Case:
    """Satırın kartını bulur. Bulunamazsa SatirHatasi — kart YARATILMAZ."""
    case_id = foy_haritasi.get(sistem_no)
    if case_id is None:
        parcalar = _dosya_no_parcalari(satir.degerler.get("dosya_no"))
        if not parcalar:
            raise SatirHatasi("Dosya No boş ve föy kaydı yok — kart eşleştirilemedi")
        gosterim = "/".join(parcalar)
        adaylar: List[int] = []
        for parca in parcalar:
            for aday in dosya_haritasi.get(parca) or []:
                if aday not in adaylar:
                    adaylar.append(aday)
        if not adaylar:
            raise SatirAtlandi(
                f"Kart bulunamadı (Dosya No {gosterim!r} klasor_no_2 ile eşleşmiyor)"
            )
        if len(adaylar) > 1:
            case_id = _ikinci_anahtarla_coz(db, satir, adaylar)
            if case_id is None:
                raise SatirHatasi(
                    f"Belirsiz eşleşme: Dosya No {gosterim!r} {len(adaylar)} kartla eşleşiyor "
                    f"({', '.join(str(a) for a in adaylar[:5])}) — esas/tür de ayırmadı"
                )
        else:
            case_id = adaylar[0]

    case = db.get(models.Case, case_id)
    if case is None or case.deleted_at is not None:
        raise SatirHatasi(f"Kart {case_id} yok ya da silinmiş")
    return case


def _kart_alanlarini_yaz(db, case: models.Case, satir: HamSatir,
                         source: str,
                         celiskili_alanlar: Set[str] = frozenset(),
                         atlanan_alanlar: Optional[List[Tuple[str, str]]] = None,
                         *, sistem_no: Optional[str] = None,
                         duzeltmeler: Optional[DuzeltmeHaritasi] = None,
                         bosaltilanlar: Optional[List[str]] = None) -> List[str]:
    """DAR alan kümesini kartın ÜZERİNE yazar (UPDATE-in-place); değişenleri döner.

    Değişmeyen alan için ne UPDATE ne `case_history` satırı üretilir — ikinci
    koşunun "0 değişiklik" kabul kriteri buna dayanır. `None` gelen alan
    KORUNUR (partili teslimde eksik sütun mevcut değeri silmez).
    `celiskili_alanlar` kardeş föylerin uzlaşamadığı alanlardır: dönüşüm YİNE
    koşar (bozuk değer satırı düşürmeye devam etsin diye) ama yazım atlanır —
    gerekçe `kart_alan_celiskileri` docstring'inde. `atlanan_alanlar`
    `kart_degerleri`ye aynen geçer (G104 alan-düzeyi rapor toplayıcısı).

    G112: `duzeltmeler` verilirse (a) değişen alanın `case_history.source`
    imzasına `Düzeltme_Logu` gerekçesi eklenir (`_provenance_imzasi`), (b)
    `(boş)` talimatı ÜÇLÜ şartla uygulanır — log `(boş)` + `Sheet` hücresi
    gerçekten boş (`_bosaltma_talimatlari`) + bizde dolu — alan NULL yazılır,
    tarihçeye "boşaltıldı: <gerekçe>" düşer, alan adı `bosaltilanlar`a eklenir
    (dönüş listesinde de yer alır). Bizde zaten boşsa hiçbir şey yazılmaz
    (ikinci koşu 0 değişiklik). Künye/içerik alanı talimatı `atlanan_alanlar`a
    düşer, uygulanmaz. `esas_no` boşaltması da tek yoldan (`sync_current_esas`
    boş değerle kolonu temizler, tarihçe satırları kalır).
    """
    degisenler: List[str] = []
    degerler = kart_degerleri(satir, atlanan_alanlar)
    for alan, yeni in degerler.items():
        if alan in celiskili_alanlar:
            continue
        eski = getattr(case, alan)
        if eski == yeni:
            continue
        if (alan in ICERIK_KARSILASTIRMALI_ALANLAR
                and _baslik_anahtari(eski) == _baslik_anahtari(yeni)):
            continue                      # yalnız yazım farkı — bizimki kalır
        if alan == "esas_no":
            # Türetilmiş alan: kolon + tarihçe TEK yoldan (G045). Buradan
            # setattr etmek `case_esas_numbers`ı bypass edip ikinci doğruluk
            # kaynağı doğururdu; eski esas da kayıtta kalmalı.
            case_manager.sync_current_esas(
                db, case, yeni,
                court=degerler.get("court") or case.court,
                source=source,
            )
        else:
            setattr(case, alan, yeni)
        kayit = duzeltmeler.get((sistem_no or "", alan)) if duzeltmeler else None
        db.add(models.CaseHistory(
            case_id=case.id, field_name=alan,
            old_value=_gecmis_metni(eski), new_value=_gecmis_metni(yeni),
            changed_by=DEGISTIREN, source=_provenance_imzasi(source, kayit),
        ))
        degisenler.append(alan)

    uygulanabilir, reddedilen = _bosaltma_talimatlari(satir, sistem_no or "", duzeltmeler)
    if atlanan_alanlar is not None:
        atlanan_alanlar.extend(reddedilen)
    for alan, kayit in uygulanabilir:
        if alan in celiskili_alanlar or alan in degerler:
            continue
        eski = getattr(case, alan)
        if eski is None or eski == "":
            continue                      # üçüncü ayak: bizde zaten boş — idempotent
        if alan == "esas_no":
            case_manager.sync_current_esas(db, case, None, source=source)
        else:
            setattr(case, alan, None)
        db.add(models.CaseHistory(
            case_id=case.id, field_name=alan,
            old_value=_gecmis_metni(eski), new_value=None,
            changed_by=DEGISTIREN, source=_provenance_imzasi(source, kayit, bosaltma=True),
        ))
        degisenler.append(alan)
        if bosaltilanlar is not None:
            bosaltilanlar.append(alan)
        logger.info(f"{sistem_no} {alan} boşaltıldı (Düzeltme_Logu satır {kayit.satir_no})")
    return degisenler


# Taraf sütunu → (party_type, varsayılan rol). Rol, föyün "Taraf Sıfatı"
# sütunundan gelir; o boşsa buradaki varsayılan yazılır.
TARAF_SUTUNLARI: Tuple[Tuple[str, str, str], ...] = (
    ("muvekkil",     "CLIENT", "Müvekkil"),
    ("karsi_taraf",  "COUNTER", "Karşı Taraf"),
    # D1 (şartname §2): Sigortalı taraf kaydı olur, `THIRD`, rol adı
    # "Sigortalı"; KARŞI TARAF DEĞİLDİR — çıkar çatışması kontrolünden hariç,
    # aramaya dahil.
    ("sigortali",    "THIRD", "Sigortalı"),
    ("davali_idare", "THIRD", "Davalı İdare"),
)


# ─── Karar aşamaları (Karar_Asamalari sayfası) ───────────────────────────────
# Teslimin 18.08'de eklediği aşama katmanı: föy başına 1-5 satır, her satır bir
# yargı aşaması (Yerel → İstinaf → Temyiz → Karar Düzeltme). 68 sütunluk ana
# sayfadaki tek-slot künyenin YERİNE GEÇMEZ, yanına gelir — ve bizim
# `case_stage_decisions` tablomuzun (G062) birebir karşılığıdır.
ASAMA_SAYFASI = "Karar_Asamalari"

ASAMA_SUTUNLARI: Dict[str, Tuple[str, ...]] = {
    "sistem_no":    ("SistemNo",),
    "asama_no":     ("AsamaNo", "Aşama No"),
    "asama":        ("Aşama",),
    "mahkeme":      ("Mahkeme",),
    "esas_no":      ("Esas No",),
    "karar_no":     ("Karar No",),
    "karar_tarihi": ("Karar Tarihi",),
    "karar_durumu": ("Karar Durumu",),
    "teblig_tarihi": ("Tebliğ Tarihi",),
    "basvuran_taraf": ("Başvuran Taraf",),
    "guven":        ("Güven",),
    "aciklama":     ("Açıklama",),
}

# Teslimin aşama etiketi → bizim `DECISION_STAGES`. "Önceki" bir KARAR aşaması
# değildir (yalnız görevsizlik öncesi esas numarası) — esas tarihçesine gider.
ASAMA_ESLEMESI = {
    "YEREL": "YEREL", "ISTINAF": "ISTINAF", "TEMYIZ": "TEMYIZ",
    "KARARDUZELTME": "KARAR_DUZELTME",
}
ASAMA_ONCEKI = "ONCEKI"

# Güven → `dogrulama_durumu`. Teslim "KESİN" derken kaynağı "68-türetme"dir:
# UYAP'tan ya da belgeden okunmuş değil, kendi 68 sütunundan TÜRETİLMİŞ.
# Damgayı olduğundan güçlü göstermek tahmin yasağının ihlali olurdu.
GUVEN_ESLEMESI = {"KESIN": "TURETILDI", "BELIRSIZ": "BELIRSIZ"}


def _taraf_adlari(deger: Any) -> List[str]:
    """`;` ile birleşik taraf listesini adlara böler (3.201 föyde çoklu)."""
    ham = _metin(deger)
    if not ham:
        return []
    adlar: List[str] = []
    for parca in _AYRAC.split(ham):
        ad = " ".join(parca.split())
        if ad and ad not in adlar:
            adlar.append(ad)
    return adlar


def _taraflari_yaz(db, case: models.Case, satir: HamSatir, source: str) -> List[str]:
    """Föyün taraflarını `case_parties`e YALNIZ-EKLEME ile işler.

    **Silme ve toptan yeniden yazma YASAK** (18.08 belge koruma şartı):
    `case_documents.case_party_id` FK'sı `ondelete=SET NULL` — mevcut taraf
    satırlarını silip yeniden yazan bir aktarım belge-taraf bağını HATA
    VERMEDEN koparırdı. Bu yüzden mevcut satıra DOKUNULMAZ (rolü bile
    güncellenmez); yalnız kartta hiç olmayan ad eklenir.

    Ad eşleşmesi `party_check.normalize_party_key` ile: "X A.Ş." ile "X Anonim
    Şirketi" aynı anahtara düşer, kelime sırası önemsizdir. Ölçüm (2026-08-19):
    föy müvekkillerinin %99,6'sı kartta zaten taraf olarak vardı — bu fonksiyon
    çoğunlukla eksik kalanı tamamlar.
    """
    sifat = _metin(satir.degerler.get("taraf_sifati"))
    eklenen: List[str] = []
    mevcut = {
        normalize_party_key(p.name): p
        for p in db.query(models.CaseParty).filter(models.CaseParty.case_id == case.id)
    }
    for kaynak, party_type, varsayilan_rol in TARAF_SUTUNLARI:
        rol = (sifat or varsayilan_rol) if party_type == "CLIENT" else varsayilan_rol
        for ad in _taraf_adlari(satir.degerler.get(kaynak)):
            anahtar = normalize_party_key(ad)
            if not anahtar or anahtar in mevcut:
                continue
            db.add(models.CaseParty(
                case_id=case.id, name=ad, role=rol, party_type=party_type,
                client_id=None,
            ))
            db.add(models.CaseHistory(
                case_id=case.id, field_name="taraf", old_value=None,
                new_value=f"{ad} ({rol})", changed_by=DEGISTIREN, source=source,
            ))
            mevcut[anahtar] = None            # aynı satırda ikilenmesin
            eklenen.append(ad)
    return eklenen


def _avukatlari_yaz(db, case: models.Case, satir: HamSatir, source: str) -> List[str]:
    """Föyün avukat listesini `case_lawyers`e YALNIZ-EKLEME ile işler.

    Silme YOK (belge koruma şartının kardeşi: toptan silip yeniden yazmak
    bağları koparır) ve mükerrer YOK — normalize ada göre bakılır, ikinci koşu
    hiçbir satır eklemez. `lawyer_id` resmî listede karşılığı varsa bağlanır;
    yoksa satır yine yazılır (ad kaybolmasın), bağ boş kalır.
    """
    adlar = _avukat_adlari(satir.degerler.get("avukatlar"))
    if not adlar:
        return []
    mevcut = {
        _baslik_anahtari(satir_.name): satir_
        for satir_ in db.query(models.CaseLawyer).filter(
            models.CaseLawyer.case_id == case.id
        )
    }
    eklenen: List[str] = []
    for ad in adlar:
        if _baslik_anahtari(ad) in mevcut:
            continue
        resmi = db.query(models.Lawyer.id).filter(
            func.upper(models.Lawyer.name) == tr_upper(ad)
        ).first()
        db.add(models.CaseLawyer(
            case_id=case.id, name=ad, lawyer_id=resmi[0] if resmi else None,
        ))
        db.add(models.CaseHistory(
            case_id=case.id, field_name="avukat", old_value=None, new_value=ad,
            changed_by=DEGISTIREN, source=source,
        ))
        mevcut[_baslik_anahtari(ad)] = None      # aynı satırda ikilenmesin
        eklenen.append(ad)
    return eklenen


def _gecmis_metni(deger: Any) -> Optional[str]:
    if deger is None:
        return None
    if isinstance(deger, (date, datetime)):
        return deger.isoformat()
    return str(deger)


def _satiri_isle(db, satir: HamSatir, *, foy_haritasi: Dict[str, int],
                 dosya_haritasi: Dict[str, List[int]], source: str,
                 foy_source: str, sonuc: AktarimSonucu,
                 kart_celiskileri: Optional[Dict[int, Set[str]]] = None,
                 duzeltmeler: Optional[DuzeltmeHaritasi] = None) -> int:
    """TEK satırın işi (kart id'sini döner) — çağıran SAVEPOINT içinde çağırır.

    SIRA ÖNEMLİ: föy upsert'i alan doğrulamasından ÖNCE gelir; bozuk bir alan
    savepoint'i geri alırken föyü de geri alır. Yarım föy (kimlik yazılmış,
    veri yazılmamış) bırakmak, düzeltme listesiyle zaten geri gelecek bir satır
    için sessiz bozulma olurdu.
    """
    sistem_no = _metin(satir.degerler.get("sistem_no"))
    if not sistem_no:
        raise SatirHatasi("SistemNo boş (föyün kimliği)")

    case = _kart_coz(db, satir, foy_haritasi, dosya_haritasi, sistem_no)
    yeni_foy = foy_map.get_foy(db, sistem_no) is None

    foy_map.upsert_foy(
        db, case,
        sistem_no=sistem_no,
        tku_no=_metin(satir.degerler.get("tku_no")),
        hasar_no=_metin(satir.degerler.get("hasar_no")),
        source=foy_source,
    )

    atlanan_alanlar: List[Tuple[str, str]] = []
    bosaltilanlar: List[str] = []
    degisenler = _kart_alanlarini_yaz(
        db, case, satir, source,
        celiskili_alanlar=(kart_celiskileri or {}).get(case.id, frozenset()),
        atlanan_alanlar=atlanan_alanlar,
        sistem_no=sistem_no, duzeltmeler=duzeltmeler, bosaltilanlar=bosaltilanlar,
    )
    sonuc.bosaltilan += len(bosaltilanlar)
    eklenen_avukatlar = _avukatlari_yaz(db, case, satir, source)
    eklenen_taraflar = _taraflari_yaz(db, case, satir, source)

    if yeni_foy:
        # Föyün kartla EŞLENMESİ de bir değişikliktir; provenance imzası
        # (D8/K1) yalnız alan değişikliğine bağlı kalırsa, hiçbir alanı
        # değişmeyen aktarım kartı "elle açılmış" kovada görünürdü.
        db.add(models.CaseHistory(
            case_id=case.id, field_name="case_foys.sistem_no",
            old_value=None, new_value=sistem_no,
            changed_by=DEGISTIREN, source=source,
        ))
        sonuc.foy_yeni += 1
    else:
        sonuc.foy_guncellenen += 1

    if eklenen_avukatlar:
        sonuc.avukat_eklenen += len(eklenen_avukatlar)
        sonuc.degisen_kartlar.add(cast(int, case.id))
    if eklenen_taraflar:
        sonuc.taraf_eklenen += len(eklenen_taraflar)
        sonuc.degisen_kartlar.add(cast(int, case.id))

    if degisenler or yeni_foy or eklenen_avukatlar or eklenen_taraflar:
        # Türetilmiş eksik-alan kovasını TEK yazma yolundan tazele (D8):
        # aktarım imzası yeni düştüyse kayıt AKTARIM kovasına geçmeli.
        case_manager.refresh_missing_required(db, case)
    if degisenler:
        sonuc.alan_degisikligi += len(degisenler)
        sonuc.degisen_kartlar.add(cast(int, case.id))

    # G104 — alan-düzeyi atlamalar satır raporuna HATA olarak düşer (satır
    # İŞLENDİ, yalnız o alan yazılmadı; sebep sütunu bunu açıkça söyler).
    # SATIR düşerse (SatirHatasi yukarı fırlar) buraya hiç gelinmez ve rapora
    # yalnız satır-düzeyi kayıt düşer — ikileme yok.
    for alan, sebep in atlanan_alanlar:
        sonuc.rapor_satirlari.append(RaporSatiri(
            satir_no=satir.satir_no, sistem_no=sistem_no,
            dosya_no=_metin(satir.degerler.get("dosya_no")) or "",
            tur="HATA", sebep=f"{alan} yazılmadı: {sebep}",
        ))
        logger.warning(
            f"Satır {satir.satir_no} ({sistem_no}) {alan} yazılmadı: {sebep}"
        )

    sonuc.islenen += 1
    return cast(int, case.id)


# ═══════════════════════════════════════════════════════════════════════════
# Kardeş föy çelişki raporu
# ═══════════════════════════════════════════════════════════════════════════

def celiskileri_bul(kayitlar: Sequence[Dict[str, Any]]) -> List[Celiski]:
    """Aynı kart (yoksa aynı TKU) altındaki föylerde künye uyuşmazlıkları.

    Tasarım paketinin VAKALAR kanıtı: id-7189 K.2018/143 ile id-7190
    K.2016/768 aynı olayın iki föyü ama karar künyeleri farklı — kartta künye
    TEK SLOT olduğu için biri diğerini ezerdi. Bu satırlar
    `dogrulama_durumu = BELIRSIZ` ile işaretlenmeye ADAYDIR; işaretleme bu
    turda YOK, rapor üretimi yeterli (görev tanımı).

    Kart çözülemeyen satırlar TKU ile gruplanır: eşleşmemiş satırların
    çelişkisi de karşı tarafa borçlu olduğumuz bir bulgudur.
    """
    gruplar: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for kayit in kayitlar:
        if kayit.get("case_id"):
            anahtar = ("KART", str(kayit["case_id"]))
        elif kayit.get("tku_no"):
            anahtar = ("TKU", str(kayit["tku_no"]))
        else:
            continue
        gruplar.setdefault(anahtar, []).append(kayit)

    celiskiler: List[Celiski] = []
    for (kume, kume_anahtari), uyeler in sorted(gruplar.items()):
        if len(uyeler) < 2:
            continue
        for alan in KUNYE_ALANLARI:
            dolu = [(u["sistem_no"], u.get(alan) or "") for u in uyeler if u.get(alan)]
            if len({d for _, d in dolu}) > 1:
                celiskiler.append(Celiski(
                    kume=kume, kume_anahtari=kume_anahtari, alan=alan,
                    degerler=" | ".join(f"{s}={d}" for s, d in sorted(dolu)),
                ))
    return celiskiler


# ═══════════════════════════════════════════════════════════════════════════
# Koşu
# ═══════════════════════════════════════════════════════════════════════════

def asama_satirlarini_oku(yol: Path, *, sheet: str = ASAMA_SAYFASI) -> List[HamSatir]:
    """`Karar_Asamalari` sayfasını okur; sayfa yoksa BOŞ liste (hata değil).

    Sayfa 18.08 teslimiyle geldi; daha eski paketlerde yok ve aktarımın geri
    kalanı onsuz da çalışmalı.
    """
    import openpyxl

    wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            logger.info(f"{sheet!r} sayfası yok — aşama aktarımı atlandı")
            return []
        ws = wb[sheet]
        akis = ws.iter_rows(values_only=True)
        baslik = next(akis, None)
        if baslik is None:
            return []
        dosyadaki = {}
        for i, ham in enumerate(baslik):
            anahtar = _baslik_anahtari(ham)
            if anahtar and anahtar not in dosyadaki:
                dosyadaki[anahtar] = i
        indeksler = {}
        for alan, adaylar in ASAMA_SUTUNLARI.items():
            for aday in adaylar:
                if _baslik_anahtari(aday) in dosyadaki:
                    indeksler[alan] = dosyadaki[_baslik_anahtari(aday)]
                    break
        eksik = [a for a in ("sistem_no", "asama") if a not in indeksler]
        if eksik:
            raise AktarimHatasi(f"{sheet}: zorunlu sütun(lar) yok: {', '.join(eksik)}")
        satirlar: List[HamSatir] = []
        for sira, ham in enumerate(akis, start=2):
            if ham is None or all(_metin(h) is None for h in ham):
                continue
            satirlar.append(HamSatir(
                satir_no=sira,
                degerler={a: (ham[i] if i < len(ham) else None)
                          for a, i in indeksler.items()},
            ))
        return satirlar
    finally:
        wb.close()


def _asama_imzasi(satir: HamSatir) -> Tuple[str, ...]:
    """Kardeş föylerin aynı aşamayı aynı anlatıp anlatmadığının anahtarı."""
    return (
        _baslik_anahtari(_metin(satir.degerler.get("mahkeme")) or ""),
        _metin(satir.degerler.get("esas_no")) or "",
        _metin(satir.degerler.get("karar_no")) or "",
        _tarih_yumusak(satir.degerler.get("karar_tarihi"), "karar_tarihi"),
        _metin(satir.degerler.get("karar_durumu")) or "",
    )


def _basvuran_taraf(satir: HamSatir, stage: str,
                    foy_satirlari: Dict[str, HamSatir]) -> Optional[str]:
    """Başvuran taraf: önce aşama sayfası, boşsa föyün 68 sütunluk değeri.

    68 sütundaki "İstinaf Mahkemesi Başvuran Taraf" (406 dolu) kart kolonuna
    DEĞİL buraya akar; kolonun tek yazıcısı aşama fotoğrafıdır.
    """
    donustur = _esleme(ISTINAF_BASVURAN_ESLEMESI)
    deger = donustur(satir.degerler.get("basvuran_taraf"), "basvuran_taraf")
    if deger or stage != "ISTINAF":
        return deger
    foy = foy_satirlari.get(_metin(satir.degerler.get("sistem_no")) or "")
    return donustur(foy.degerler.get("istinaf_basvuran"), "basvuran_taraf") if foy else None


def _asama_sira(satir: HamSatir) -> int:
    try:
        return int(str(_metin(satir.degerler.get("asama_no")) or "0").strip())
    except ValueError:
        return 0


def asamalari_yaz(db, asama_satirlari: Sequence[HamSatir], *,
                  foy_haritasi: Dict[str, int], foy_satirlari: Dict[str, HamSatir],
                  source: str, sonuc: AktarimSonucu) -> None:
    """Aşama satırlarını `case_stage_decisions` + esas tarihçesine yazar.

    Kurallar:

    * **Kart başına, aşama başına** çalışır. Bir kartın birden çok föyü aynı
      aşamayı anlatıyorsa imzalar karşılaştırılır: aynıysa TEK kez yazılır,
      farklıysa o aşama YAZILMAZ ve çelişki raporuna düşer (D9'un aşama
      tarafındaki karşılığı — 279 kartta künyeler gerçekten çelişiyor).
    * **Dolu aşamaya dokunulmaz.** Kartın o aşamasında zaten satır varsa
      aktarım geçer: elle girilmiş bir kararı ezmek, tek yazma yolunun amacını
      bozardı. İdempotentlik de buradan gelir — ikinci koşu hiçbir satır
      eklemez.
    * **"Önceki" bir karar değildir**: görevsizlik/yenileme öncesi esas
      numarasıdır, `case_esas_numbers`a ONCEKI olarak düşer (güncel işaret
      DEĞİŞMEZ).
    * Kapalı havuza uymayan `karar_durumu` satırı DÜŞÜRMEZ: durum boş
      bırakılıp değer açıklamaya taşınır ve rapora yazılır (8.354 satırın
      yalnız 8'i böyle — teslim bu sayfayı bizim havuzlarımıza göre
      normalize etmiş).
    """
    kart_asamalari: Dict[Tuple[int, str], Dict[str, List[HamSatir]]] = {}
    onceki_esaslar: Dict[int, List[HamSatir]] = {}
    for satir in asama_satirlari:
        sistem_no = _metin(satir.degerler.get("sistem_no"))
        case_id = foy_haritasi.get(sistem_no or "")
        if case_id is None:
            continue                      # föy bu koşuda kartla eşleşmedi
        etiket = _baslik_anahtari(_metin(satir.degerler.get("asama")) or "")
        if etiket == ASAMA_ONCEKI:
            onceki_esaslar.setdefault(case_id, []).append(satir)
            continue
        stage = ASAMA_ESLEMESI.get(etiket)
        if stage is None:
            continue
        kart_asamalari.setdefault((case_id, stage), {}).setdefault(
            sistem_no or "", []).append(satir)

    for case_id, satirlar in sorted(onceki_esaslar.items()):
        case = db.get(models.Case, case_id)
        if case is None or case.deleted_at is not None:
            continue
        for satir in satirlar:
            if case_manager.add_historical_esas(
                db, case, _metin(satir.degerler.get("esas_no")),
                court=_metin(satir.degerler.get("mahkeme")), source=source,
            ) is not None:
                sonuc.onceki_esas_eklenen += 1

    for (case_id, stage), foyler in sorted(kart_asamalari.items()):
        imzalar = {
            foy: tuple(_asama_imzasi(s) for s in sorted(satirlar, key=_asama_sira))
            for foy, satirlar in foyler.items()
        }
        if len(set(imzalar.values())) > 1:
            sonuc.celiskiler.append(Celiski(
                kume="KART", kume_anahtari=str(case_id), alan=f"asama:{stage}",
                degerler=" | ".join(
                    f"{foy}={'/'.join(i[2] or i[1] or '-' for i in imza)}"
                    for foy, imza in sorted(imzalar.items())
                ),
            ))
            continue

        case = db.get(models.Case, case_id)
        if case is None or case.deleted_at is not None:
            continue
        mevcut = db.query(models.CaseStageDecision.id).filter(
            models.CaseStageDecision.case_id == case_id,
            models.CaseStageDecision.stage == stage,
        ).first()
        if mevcut is not None:
            continue                      # dolu aşamaya dokunulmaz

        kanonik = sorted(next(iter(foyler.values())), key=_asama_sira)
        for sira, satir in enumerate(kanonik, start=1):
            durum = _metin(satir.degerler.get("karar_durumu"))
            aciklama = _metin(satir.degerler.get("aciklama"))
            damga = GUVEN_ESLEMESI.get(
                _baslik_anahtari(_metin(satir.degerler.get("guven")) or ""), "BELIRSIZ")
            # İki deneme: önce kaynağın durumu, reddedilirse durumsuz + şerh.
            # Bayrak AYRI taşınır: "deneme is None" fallback'i BELİRTMEZ — durum
            # zaten boşken ilk deneme de None olur ve o okuma açıklamaya Python'ın
            # `None`'unu basardı ("havuz dışı durum: None"; 2026-08-19 koşusunda
            # 833 satır, hepsi kullanıcıya bu hâliyle görünüyordu, G076).
            for havuz_disi in (False, True):
                deneme = None if havuz_disi else durum
                try:
                    stage_decisions.add_stage_decision(
                        db, case, stage=stage, sira_no=sira,
                        mahkeme=_metin(satir.degerler.get("mahkeme")),
                        esas_no=_metin(satir.degerler.get("esas_no")),
                        karar_no=_metin(satir.degerler.get("karar_no")),
                        karar_tarihi=_tarih(satir.degerler.get("karar_tarihi"), "karar_tarihi"),
                        karar_durumu=deneme,
                        teblig_tarihi=_tarih(satir.degerler.get("teblig_tarihi"), "teblig_tarihi"),
                        basvuran_taraf=_basvuran_taraf(satir, stage, foy_satirlari),
                        aciklama=" · ".join(x for x in (aciklama, f"havuz dışı durum: {durum}") if x)
                        if havuz_disi else aciklama,
                        dogrulama_durumu=damga, source=source,
                    )
                    sonuc.asama_eklenen += 1
                    break
                except stage_decisions.InvalidDecisionStatusError:
                    if havuz_disi:
                        raise
                    sonuc.havuz_disi_durum += 1
                    logger.warning(
                        f"Aşama karar durumu havuz dışı, durum boş yazıldı: "
                        f"kart {case_id} {stage} {durum!r}"
                    )
                except SatirHatasi as exc:
                    logger.warning(f"Aşama satırı düştü (kart {case_id} {stage}): {exc}")
                    break


def _statement_timeout_yukselt(db, ms: int) -> bool:
    """Koşu süresince statement_timeout'u yükseltir (yalnız Postgres).

    `SET` bind parametresi almaz; `set_config` alır (üçüncü argüman is_local
    = false: oturum boyu). sqlite birim koşusunda sessizce atlanır.
    """
    if ms <= 0 or db.get_bind().dialect.name != "postgresql":
        return False
    db.execute(text("SELECT set_config('statement_timeout', :ms, false)"), {"ms": str(ms)})
    logger.info(f"statement_timeout koşu için {ms} ms'e yükseltildi")
    return True


def aktarimi_kos(session_factory, *, girdi: Path, sheet: Optional[str] = None,
                 limit: Optional[int] = None, dry_run: bool = False,
                 source: Optional[str] = None, rapor_dizini: Optional[Path] = None,
                 statement_timeout_ms: int = VARSAYILAN_TIMEOUT_MS) -> AktarimSonucu:
    """Çekirdek akış: oku → normalize → föy upsert → kart alanları → raporlar.

    TEK transaction, TEK commit: belge envanteri kapısı commit'ten ÖNCE ölçer,
    parti başına commit o kapıyı bölerdi. Kapı kırmızıysa (ya da `dry_run`)
    koşu tamamen geri alınır ve NONZERO döner.
    """
    girdi = Path(girdi)
    satirlar, bulunan_basliklar = xlsx_oku(girdi, sheet=sheet, limit=limit)
    asama_satirlari = asama_satirlarini_oku(girdi) if limit is None else []
    # Düzeltme_Logu (G112) limit'ten bağımsız okunur: SistemNo anahtarlı,
    # yalnız işlenen satırların kayıtları kullanılır.
    duzeltmeler = duzeltme_logunu_oku(girdi)
    kaynak_imzasi = source or f"{AKTARIM_SOURCE_PREFIX}_{girdi.name}"
    foy_source = kaynak_imzasi[:100]
    if len(kaynak_imzasi) > 100:
        logger.warning(
            f"Föy kaynağı 100 karaktere kırpıldı: {kaynak_imzasi!r} → {foy_source!r}"
        )

    sonuc = AktarimSonucu(okunan=len(satirlar), dry_run=dry_run,
                          kaynak_imzasi=kaynak_imzasi)
    logger.info(
        f"Aktarım başlıyor: {girdi.name} · {len(satirlar)} satır · "
        f"{'KURU KOŞU' if dry_run else 'YAZMA'} · bulunan sütunlar: "
        f"{', '.join(sorted(bulunan_basliklar))}"
    )

    db = session_factory()
    try:
        _statement_timeout_yukselt(db, statement_timeout_ms)
        sonuc.envanter_once = belge_envanteri.snapshot(db)

        foy_haritasi = foy_map.map_sistem_no_to_case(
            db, [_metin(s.degerler.get("sistem_no")) for s in satirlar]
        )
        dosya_haritasi = _dosya_no_haritasi(db)
        avukat_haritasi_kur(db)

        # ÖN GEÇİŞ (yazmaz): kardeş föylerin kart alanlarında uzlaşıp
        # uzlaşmadığı ÖNCE bilinmeli — satır satır yazarken öğrenilseydi ilk
        # föy zaten kartı bir kez ezmiş olurdu.
        kart_celiskileri, kart_celiski_raporu = kart_alan_celiskileri(
            db, satirlar, foy_haritasi=foy_haritasi, dosya_haritasi=dosya_haritasi,
            duzeltmeler=duzeltmeler,
        )
        if kart_celiski_raporu:
            logger.warning(
                f"{len(kart_celiski_raporu)} kart alanında kardeş föyler uzlaşmadı — "
                f"o alanlar YAZILMADI (rapora düştü)"
            )

        kunye_kayitlari: List[Dict[str, Any]] = []

        for satir in satirlar:
            sistem_no = _metin(satir.degerler.get("sistem_no")) or ""
            try:
                with db.begin_nested():
                    case_id = _satiri_isle(
                        db, satir,
                        foy_haritasi=foy_haritasi, dosya_haritasi=dosya_haritasi,
                        source=kaynak_imzasi, foy_source=foy_source, sonuc=sonuc,
                        kart_celiskileri=kart_celiskileri, duzeltmeler=duzeltmeler,
                    )
            except SatirHatasi as exc:
                # Savepoint geri alındı; bellekteki (flush edilmemiş) hâl bayat.
                db.expire_all()
                tur = "ATLANDI" if isinstance(exc, SatirAtlandi) else "HATA"
                if tur == "ATLANDI":
                    sonuc.atlanan += 1
                sonuc.rapor_satirlari.append(RaporSatiri(
                    satir_no=satir.satir_no, sistem_no=sistem_no,
                    dosya_no=_metin(satir.degerler.get("dosya_no")) or "",
                    tur=tur, sebep=str(exc),
                ))
                logger.warning(f"Satır {satir.satir_no} ({sistem_no}) {tur}: {exc}")
            except SQLAlchemyError as exc:
                db.expire_all()
                sonuc.rapor_satirlari.append(RaporSatiri(
                    satir_no=satir.satir_no, sistem_no=sistem_no,
                    dosya_no=_metin(satir.degerler.get("dosya_no")) or "",
                    tur="HATA", sebep=f"{type(exc).__name__}: {exc}",
                ))
                logger.warning(f"Satır {satir.satir_no} ({sistem_no}) DB hatası: {exc}")
            else:
                # Aynı dosyada ikinci kez geçen SistemNo doğrudan bu karta
                # düşsün (Dosya No köprüsüne ikinci kez gitmeye gerek yok).
                foy_haritasi[sistem_no] = case_id
                kunye_kayitlari.append({
                    "sistem_no": sistem_no,
                    "case_id": case_id,
                    "tku_no": _metin(satir.degerler.get("tku_no")),
                    "karar_no": _metin(satir.degerler.get("karar_no")),
                    "karar_tarihi": _tarih_yumusak(
                        satir.degerler.get("karar_tarihi"), "karar_tarihi"),
                })

        # Künye çelişkileri (yazılmayan alanlar) + kart alanı çelişkileri
        # (yazımı ATLANAN alanlar) TEK raporda buluşur: ikisi de "kardeş föyler
        # uzlaşmadı, kartta kur'a çekmedik" demektir.
        # Aşama katmanı ana döngüden SONRA: föy→kart haritası ancak burada tam
        # (bir föy kartına ilk kez bu koşuda bağlanmış olabilir).
        asamalari_yaz(
            db, asama_satirlari, foy_haritasi=foy_haritasi,
            foy_satirlari={
                _metin(s.degerler.get("sistem_no")) or "": s for s in satirlar
            },
            source=kaynak_imzasi, sonuc=sonuc,
        )

        sonuc.celiskiler = kart_celiski_raporu + celiskileri_bul(kunye_kayitlari) + sonuc.celiskiler

        db.flush()
        sonuc.envanter_sonra = belge_envanteri.snapshot(db)
        sonuc.envanter_farki = belge_envanteri.diff(sonuc.envanter_once, sonuc.envanter_sonra)

        if sonuc.envanter_farki:
            # Belge koruma şartı: hasar KALICI OLMADAN durdurulur.
            logger.error(
                "Aktarım GERİ ALINDI — " + belge_envanteri.bicimle(sonuc.envanter_farki)
            )
            db.rollback()
        elif dry_run:
            db.rollback()
            logger.info("KURU KOŞU: hiçbir tabloya yazılmadı (geri alındı)")
        else:
            db.commit()
            sonuc.yazildi = True
    finally:
        db.close()

    sonuc.raporlar = _raporlari_yaz(sonuc, girdi=girdi, rapor_dizini=rapor_dizini)
    return sonuc


# ═══════════════════════════════════════════════════════════════════════════
# Raporlar
# ═══════════════════════════════════════════════════════════════════════════

def _raporlari_yaz(sonuc: AktarimSonucu, *, girdi: Path,
                   rapor_dizini: Optional[Path]) -> List[Path]:
    """Satır raporu + kardeş föy çelişki raporu (CSV; UTF-8 BOM, ';' ayraç).

    Kuru koşuda DA yazılır — kuru koşunun ürünü zaten rapordur. CSV seçildi:
    xlsx'in biçimlendirmesine ihtiyaç yok, ';' + BOM ile Türkçe Excel'de
    doğrudan açılıyor ve dosya diff'lenebilir kalıyor.
    """
    if not sonuc.rapor_satirlari and not sonuc.celiskiler:
        return []
    dizin = Path(rapor_dizini) if rapor_dizini else girdi.parent / "aktarim-raporlari"
    dizin.mkdir(parents=True, exist_ok=True)
    damga = datetime.now().strftime("%Y%m%d-%H%M%S")

    yazilan: List[Path] = []
    if sonuc.rapor_satirlari:
        yazilan.append(_csv_yaz(
            dizin / f"satir-raporu_{damga}.csv",
            ("satir_no", "sistem_no", "dosya_no", "tur", "sebep"),
            [(r.satir_no, r.sistem_no, r.dosya_no, r.tur, r.sebep)
             for r in sonuc.rapor_satirlari],
        ))
    if sonuc.celiskiler:
        yazilan.append(_csv_yaz(
            dizin / f"kardes-foy-celiskileri_{damga}.csv",
            ("kume", "kume_anahtari", "alan", "degerler"),
            [(c.kume, c.kume_anahtari, c.alan, c.degerler) for c in sonuc.celiskiler],
        ))
    return yazilan


def _csv_yaz(yol: Path, basliklar: Sequence[str], satirlar: Iterable[Sequence[Any]]) -> Path:
    with open(yol, "w", newline="", encoding="utf-8-sig") as dosya:
        yazici = csv.writer(dosya, delimiter=";")
        yazici.writerow(basliklar)
        yazici.writerows(satirlar)
    return yol


def ozet_metni(sonuc: AktarimSonucu) -> str:
    """Koşunun tek ekranlık özeti (stdout + log)."""
    satirlar = [
        "=" * 78,
        f"HUKDOK aktarımı — {sonuc.kaynak_imzasi}",
        "=" * 78,
        f"  okunan satır      : {sonuc.okunan}",
        f"  işlenen satır     : {sonuc.islenen}",
        f"  yeni föy          : {sonuc.foy_yeni}",
        f"  güncellenen föy   : {sonuc.foy_guncellenen}",
        f"  alan değişikliği  : {sonuc.alan_degisikligi} ({sonuc.kart_degisen} kart)",
        f"  boşaltılan alan   : {sonuc.bosaltilan} (Düzeltme_Logu açık talimatı)",
        f"  avukat satırı     : {sonuc.avukat_eklenen}",
        f"  taraf satırı      : {sonuc.taraf_eklenen}",
        f"  aşama satırı      : {sonuc.asama_eklenen} (önceki esas: {sonuc.onceki_esas_eklenen}"
        f"{f', havuz dışı durum: {sonuc.havuz_disi_durum}' if sonuc.havuz_disi_durum else ''})",
        f"  atlanan (kart yok): {sonuc.atlanan}",
        f"  satır hatası      : {len(sonuc.hatalar)}",
        f"  kardeş çelişkisi  : {len(sonuc.celiskiler)}",
        f"  yazıldı mı        : {'HAYIR (kuru koşu)' if sonuc.dry_run else ('EVET' if sonuc.yazildi else 'HAYIR')}",
        "  " + belge_envanteri.bicimle(sonuc.envanter_farki).replace("\n", "\n  "),
    ]
    for yol in sonuc.raporlar:
        satirlar.append(f"  rapor             : {yol}")
    satirlar.append("=" * 78)
    return "\n".join(satirlar)


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="HUKDOK teslim paketini kartlara aktarır (çekirdek yazma yolu, G064)",
    )
    parser.add_argument("--input", required=True, help="teslim paketi (.xlsx)")
    parser.add_argument("--sheet", default=None, help="sayfa adı (varsayılan: ilk sayfa)")
    parser.add_argument("--limit", type=int, default=None, help="ilk N veri satırı")
    parser.add_argument("--dry-run", action="store_true",
                        help="hiçbir tabloya yazma; yalnız rapor üret")
    parser.add_argument("--source", default=None,
                        help=f"provenance imzası (varsayılan: {AKTARIM_SOURCE_PREFIX}_<dosya adı>)")
    parser.add_argument("--rapor-dizini", default=None,
                        help="rapor CSV'lerinin yazılacağı dizin")
    parser.add_argument("--statement-timeout-ms", type=int, default=VARSAYILAN_TIMEOUT_MS,
                        help="koşu süresince statement_timeout (0 = dokunma)")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    # Loglama TEK noktadan yapılandırılır (Faz 2-B): dağınık basicConfig yasak,
    # bekçisi tests/test_faz2_logging.py::test_no_basicconfig_left_in_backend.
    from logging_setup import configure_logging

    configure_logging()

    import database

    try:
        sonuc = aktarimi_kos(
            database.SessionLocal,
            girdi=Path(args.input),
            sheet=args.sheet,
            limit=args.limit,
            dry_run=args.dry_run,
            source=args.source,
            rapor_dizini=Path(args.rapor_dizini) if args.rapor_dizini else None,
            statement_timeout_ms=args.statement_timeout_ms,
        )
    except AktarimHatasi as exc:
        logger.error(f"Aktarım başlamadı: {exc}")
        return CIKIS_GIRDI

    print(ozet_metni(sonuc))
    return sonuc.cikis_kodu


if __name__ == "__main__":
    sys.exit(main())

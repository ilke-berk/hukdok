"""format_converter testleri.

Görüntü → PDF dönüşümü saf Pillow ile test edilir (harici bağımlılık yok).
Office → PDF testleri konteynerde koşar; soffice yoksa atlanır.
convert_to_pdfa2b dispatch testleri alt dönüştürücüleri mock'lar.
"""
import contextlib
import os
import shutil
import threading
import zipfile
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import fitz  # pymupdf
import pytest
from PIL import Image

from pdf.format_converter import (
    CONVERTIBLE_EXTENSIONS,
    IMAGE_EXTENSIONS,
    OFFICE_EXTENSIONS,
    ensure_pdf,
    image_to_pdf,
    office_to_pdf,
)


def _make_tiff(path, pages=3, size=(120, 80)):
    frames = [Image.new("RGB", size, color=(i * 40, 100, 200)) for i in range(pages)]
    frames[0].save(path, "TIFF", save_all=True, append_images=frames[1:])


def _pdf_page_count(pdf_path):
    with fitz.open(pdf_path) as doc:
        return doc.page_count


# ── image_to_pdf ─────────────────────────────────────────────────────────────

class TestImageToPdf:
    def test_multipage_tiff_roundtrip(self, tmp_path):
        src = tmp_path / "tarama.tif"
        _make_tiff(str(src), pages=3)
        out = image_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) == 3

    def test_single_page_jpg(self, tmp_path):
        src = tmp_path / "foto.jpg"
        Image.new("RGB", (100, 100), "red").save(str(src), "JPEG")
        out = image_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) == 1

    def test_rgba_png_normalized(self, tmp_path):
        # RGBA doğrudan PDF'e gömülemez — RGB'ye çevrilmeli
        src = tmp_path / "ekran.png"
        Image.new("RGBA", (100, 100), (255, 0, 0, 128)).save(str(src), "PNG")
        out = image_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) == 1

    def test_temp_output_when_no_path_given(self, tmp_path):
        src = tmp_path / "foto.png"
        Image.new("RGB", (50, 50), "blue").save(str(src), "PNG")
        out = image_to_pdf(str(src))
        try:
            assert out.endswith(".pdf")
            assert _pdf_page_count(out) == 1
        finally:
            os.remove(out)

    def test_corrupt_image_raises(self, tmp_path):
        src = tmp_path / "bozuk.tif"
        src.write_bytes(b"II*\x00" + b"\xff" * 32)
        with pytest.raises(RuntimeError):
            image_to_pdf(str(src), str(tmp_path / "out.pdf"))

    def test_multipage_1bit_g4_tiff_roundtrip(self, tmp_path):
        # Adliye taramalarının tipik formatı: çok sayfalı 1-bit G4 TIFF.
        # 2026-07-29 OOM kök nedeni bu yoldu — kareler RGB'ye açılıp listede
        # tutuluyordu (24x şişme × tüm sayfalar aynı anda).
        src = tmp_path / "g4.tif"
        frames = [Image.new("1", (1200, 1600), 1) for _ in range(4)]
        frames[0].save(
            str(src), "TIFF", save_all=True, append_images=frames[1:], compression="group4"
        )
        out = image_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) == 4

    def test_no_leftover_page_temps(self, tmp_path):
        import glob
        import tempfile as tf
        src = tmp_path / "tarama.tif"
        _make_tiff(str(src), pages=3)
        before = set(glob.glob(os.path.join(tf.gettempdir(), "imgpg_*.pdf")))
        image_to_pdf(str(src), str(tmp_path / "out.pdf"))
        after = set(glob.glob(os.path.join(tf.gettempdir(), "imgpg_*.pdf")))
        assert after == before


class TestNormalizeFrame:
    def test_1bit_mode_preserved(self):
        # RGB'ye açmak sayfa başına 24x bellek demek — "1" olduğu gibi kalmalı
        from pdf.format_converter import _normalize_frame
        norm = _normalize_frame(Image.new("1", (100, 100), 1))
        assert norm.mode == "1"

    def test_oversized_1bit_downscaled_as_grayscale(self, monkeypatch):
        # Küçültme kararı RGB dönüşümünden ÖNCE, orijinal mod üzerinde verilmeli
        from pdf import format_converter as fc
        monkeypatch.setattr(fc, "MAX_IMAGE_WIDTH", 500)
        monkeypatch.setattr(fc, "MAX_IMAGE_HEIGHT", 500)
        norm = fc._normalize_frame(Image.new("1", (1000, 800), 1))
        assert max(norm.size) <= 500
        assert norm.mode == "L"

    def test_oversized_rgba_downscaled_to_rgb(self, monkeypatch):
        from pdf import format_converter as fc
        monkeypatch.setattr(fc, "MAX_IMAGE_WIDTH", 400)
        monkeypatch.setattr(fc, "MAX_IMAGE_HEIGHT", 400)
        norm = fc._normalize_frame(Image.new("RGBA", (800, 600), (10, 20, 30, 128)))
        assert max(norm.size) <= 400
        assert norm.mode == "RGB"

    def test_returns_copy_not_original(self):
        # ImageSequence iterator'ı buffer'ı yeniden kullanır — kopya şart
        from pdf.format_converter import _normalize_frame
        img = Image.new("L", (50, 50), 128)
        assert _normalize_frame(img) is not img


# ── ensure_pdf dispatch ──────────────────────────────────────────────────────

class TestEnsurePdf:
    def test_image_dispatch(self, tmp_path, monkeypatch):
        import pdf.format_converter as fc
        called = {}
        monkeypatch.setattr(fc, "image_to_pdf", lambda s, o=None: called.update(img=s) or "x.pdf")
        assert fc.ensure_pdf("belge.tiff") == "x.pdf"
        assert called["img"] == "belge.tiff"

    def test_office_dispatch(self, tmp_path, monkeypatch):
        import pdf.format_converter as fc
        called = {}
        monkeypatch.setattr(fc, "office_to_pdf", lambda s, o=None: called.update(office=s) or "x.pdf")
        assert fc.ensure_pdf("belge.xlsx") == "x.pdf"
        assert called["office"] == "belge.xlsx"

    def test_unsupported_extension_raises(self):
        with pytest.raises(ValueError):
            ensure_pdf("belge.txt")

    def test_extension_sets_consistent(self):
        assert CONVERTIBLE_EXTENSIONS == IMAGE_EXTENSIONS | OFFICE_EXTENSIONS
        assert ".pdf" not in CONVERTIBLE_EXTENSIONS
        assert ".udf" not in CONVERTIBLE_EXTENSIONS


# ── office_to_pdf (LibreOffice gerektirir — konteynerde koşar) ───────────────

_soffice_missing = shutil.which("soffice") is None


@pytest.mark.skipif(_soffice_missing, reason="LibreOffice (soffice) kurulu değil")
class TestOfficeToPdf:
    def test_xlsx_to_pdf(self, tmp_path):
        from openpyxl import Workbook
        src = tmp_path / "tablo.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Dava No"
        ws["B1"] = "2024/123 Esas"
        wb.save(str(src))

        out = office_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) >= 1

    def test_docx_to_pdf(self, tmp_path):
        # Minimal geçerli OOXML docx (python-docx bağımlılığı olmadan)
        src = tmp_path / "belge.docx"
        with zipfile.ZipFile(src, "w") as z:
            z.writestr(
                "[Content_Types].xml",
                '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
                "</Types>",
            )
            z.writestr(
                "_rels/.rels",
                '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
                "</Relationships>",
            )
            z.writestr(
                "word/document.xml",
                '<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                "<w:body><w:p><w:r><w:t>Gerekçeli karar örneği — Türkçe ğüşiöç</w:t></w:r></w:p></w:body></w:document>",
            )

        out = office_to_pdf(str(src), str(tmp_path / "out.pdf"))
        assert _pdf_page_count(out) >= 1

    def test_corrupt_docx_raises(self, tmp_path):
        src = tmp_path / "bozuk.docx"
        src.write_bytes(b"PK\x03\x04 bozuk arsiv")
        with pytest.raises(RuntimeError):
            office_to_pdf(str(src), str(tmp_path / "out.pdf"))


# ── G024: ofis dosyasındaki harici bağlı görsel (SSRF) ──────────────────────


class _ProbeHandler(BaseHTTPRequestHandler):
    """Giden isteği KAYDEDER; gövde döndürmez (404 yeter — kanıt istektir)."""

    protocol_version = "HTTP/1.0"

    def _record(self):
        self.server.hits.append(f"{self.command} {self.path}")
        self.send_response(404)
        self.send_header("Content-Length", "0")
        self.end_headers()

    do_GET = _record
    do_HEAD = _record
    do_POST = _record
    do_PUT = _record
    do_OPTIONS = _record
    do_PROPFIND = _record

    def send_error(self, code, message=None, explain=None):
        # LO WebDAV keşfi tanımadığımız metotlar da deneyebilir — 501 de kanıttır
        self.server.hits.append(f"ERR{code} {getattr(self, 'path', '?')}")
        BaseHTTPRequestHandler.send_error(self, code, message)

    def log_message(self, *args):
        pass


class _ProbeServer(ThreadingHTTPServer):
    daemon_threads = True

    def __init__(self):
        super().__init__(("127.0.0.1", 0), _ProbeHandler)
        self.hits = []


@contextlib.contextmanager
def _probe():
    """Gerçek dinleyici: taban URL + isteğe bakan liste."""
    srv = _ProbeServer()
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    try:
        yield f"http://127.0.0.1:{srv.server_address[1]}", srv.hits
    finally:
        srv.shutdown()
        srv.server_close()


_EXT_RELS = (
    '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rIdX" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"'
    ' Target="{url}" TargetMode="External"/></Relationships>'
)


def _docx_with_linked_image(path, url):
    """`<a:blip r:link>` + `TargetMode="External"` taşıyan minimal .docx."""
    with zipfile.ZipFile(path, "w") as z:
        z.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Default Extension="png" ContentType="image/png"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            "</Types>",
        )
        z.writestr(
            "_rels/.rels",
            '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            "</Relationships>",
        )
        z.writestr(
            "word/document.xml",
            '<?xml version="1.0"?><w:document '
            'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture"><w:body>'
            "<w:p><w:r><w:t>Tebligat ekli belge</w:t></w:r></w:p>"
            '<w:p><w:r><w:drawing><wp:inline distT="0" distB="0" distL="0" distR="0">'
            '<wp:extent cx="1905000" cy="1905000"/><wp:docPr id="1" name="P1"/><a:graphic>'
            '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture"><pic:pic>'
            '<pic:nvPicPr><pic:cNvPr id="1" name="P1"/><pic:cNvPicPr/></pic:nvPicPr>'
            '<pic:blipFill><a:blip r:link="rIdX"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
            '<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="1905000" cy="1905000"/></a:xfrm>'
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
            "</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r></w:p>"
            "</w:body></w:document>",
        )
        z.writestr("word/_rels/document.xml.rels", _EXT_RELS.format(url=url))
    return path


def _xlsx_with_linked_image(path, url):
    """Aynı sınıf, Calc tarafı: çizim parçasında `r:link` + harici ilişki."""
    with zipfile.ZipFile(path, "w") as z:
        z.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
            "</Types>",
        )
        z.writestr(
            "_rels/.rels",
            '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        z.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Sayfa1" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        z.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>",
        )
        z.writestr(
            "xl/worksheets/sheet1.xml",
            '<?xml version="1.0"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheetData><row r="1"><c r="A1" t="inlineStr"><is><t>Dava No</t></is></c></row></sheetData>'
            '<drawing r:id="rIdDr"/></worksheet>',
        )
        z.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rIdDr" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
            "</Relationships>",
        )
        z.writestr(
            "xl/drawings/drawing1.xml",
            '<?xml version="1.0"?><xdr:wsDr '
            'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><xdr:oneCellAnchor>'
            "<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>4</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            '<xdr:ext cx="1905000" cy="1905000"/><xdr:pic>'
            '<xdr:nvPicPr><xdr:cNvPr id="1" name="P1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
            '<xdr:blipFill><a:blip r:link="rIdX"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="1905000" cy="1905000"/></a:xfrm>'
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            "</xdr:pic><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>",
        )
        z.writestr("xl/drawings/_rels/drawing1.xml.rels", _EXT_RELS.format(url=url))
    return path


class TestSeedLoProfile:
    """Tohumun kendisi (soffice gerektirmez)."""

    def test_writes_registrymodifications(self, tmp_path):
        from pdf.format_converter import _seed_lo_profile
        _seed_lo_profile(str(tmp_path))
        written = (tmp_path / "user" / "registrymodifications.xcu").read_text(encoding="utf-8")
        assert 'oor:name="ooInetProxyType"' in written
        assert 'oor:name="BlockUntrustedRefererLinks"' in written
        assert "127.0.0.1" in written

    def test_fail_closed_when_profile_unwritable(self, tmp_path, monkeypatch):
        # Tohum yazılamıyorsa dönüşüm BAŞLAMAZ — sertleştirilmemiş profille
        # devam etmek SSRF yüzeyini sessizce geri açardı.
        import pdf.format_converter as fc

        def _boom(*a, **kw):
            raise OSError("disk dolu")

        monkeypatch.setattr(fc.os, "makedirs", _boom)
        with pytest.raises(RuntimeError):
            fc._seed_lo_profile(str(tmp_path))


@pytest.mark.skipif(_soffice_missing, reason="LibreOffice (soffice) kurulu değil")
class TestOfficeRemoteResourceSsrf:
    """Gerçek soffice + gerçek dinleyici. Kanıt = giden istek sayısı."""

    def test_docx_linked_image_makes_no_request(self, tmp_path):
        with _probe() as (base, hits):
            src = _docx_with_linked_image(tmp_path / "d1.docx", f"{base}/D1_BLIP.png")
            out = office_to_pdf(str(src), str(tmp_path / "d1.pdf"))
            assert _pdf_page_count(out) >= 1
        assert hits == []

    def test_xlsx_linked_image_makes_no_request(self, tmp_path):
        with _probe() as (base, hits):
            src = _xlsx_with_linked_image(tmp_path / "x1.xlsx", f"{base}/X1_BLIP.png")
            out = office_to_pdf(str(src), str(tmp_path / "x1.pdf"))
            assert _pdf_page_count(out) >= 1
        assert hits == []

    def test_negative_control_without_hardening_does_request(self, tmp_path, monkeypatch):
        # Testin İŞ YAPTIĞININ kanıtı: tohum devre dışıyken aynı yük istek üretir.
        # Kırmızıya dönmezse yukarıdaki iki test hiçbir şey ölçmüyor demektir.
        import pdf.format_converter as fc
        monkeypatch.setattr(fc, "_seed_lo_profile", lambda profile_dir: None)
        with _probe() as (base, hits):
            src = _docx_with_linked_image(tmp_path / "d1.docx", f"{base}/D1_BLIP.png")
            office_to_pdf(str(src), str(tmp_path / "d1.pdf"))
        assert hits, "negatif kontrol istek üretmedi — yük artık delmiyor olabilir"

    def test_embedded_image_document_still_converts(self, tmp_path):
        # Meşru dosya bozulmamalı: gömülü görsel + metin + tablo PDF'te durmalı
        from openpyxl import Workbook
        src = tmp_path / "tablo.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Dava No"
        ws["B1"] = "2024/123 Esas"
        wb.save(str(src))
        out = office_to_pdf(str(src), str(tmp_path / "out.pdf"))
        with fitz.open(out) as doc:
            text = doc[0].get_text()
        assert "Dava No" in text and "2024/123 Esas" in text


# ── convert_to_pdfa2b dispatch ───────────────────────────────────────────────

class TestConvertToPdfa2bDispatch:
    def test_image_goes_through_image_branch(self, tmp_path, monkeypatch):
        import pdf.pdf_converter as pc
        src = tmp_path / "tarama.tif"
        _make_tiff(str(src), pages=2)

        # GhostScript adımını kopyalamayla taklit et — ara PDF çıktıya taşınır
        def _fake_gs(s, o, deadline=None):
            shutil.copy(s, o)
            return o

        monkeypatch.setattr(pc, "_pdf_to_pdfa2b", _fake_gs)
        result = pc.convert_to_pdfa2b(str(src))
        try:
            assert result != str(src)
            assert _pdf_page_count(result) == 2
        finally:
            os.remove(result)

    def test_office_failure_raises_not_fallback(self, tmp_path, monkeypatch):
        import pdf.pdf_converter as pc
        src = tmp_path / "bozuk.xlsx"
        src.write_bytes(b"PK\x03\x04 bozuk")

        def _boom(s, o=None, deadline=None):
            raise RuntimeError("LibreOffice dönüşüm hatası")

        monkeypatch.setattr(pc, "office_to_pdf", _boom)
        with pytest.raises(RuntimeError):
            pc.convert_to_pdfa2b(str(src))

    def test_image_failure_raises_not_fallback(self, tmp_path, monkeypatch):
        import pdf.pdf_converter as pc
        src = tmp_path / "bozuk.tif"
        src.write_bytes(b"II*\x00 bozuk")

        def _boom(s, o=None, deadline=None):
            raise RuntimeError("Görüntü dönüşüm hatası")

        monkeypatch.setattr(pc, "image_to_pdf", _boom)
        with pytest.raises(RuntimeError):
            pc.convert_to_pdfa2b(str(src))

    def test_unsupported_extension_raises_valueerror(self, tmp_path):
        import pdf.pdf_converter as pc
        src = tmp_path / "belge.txt"
        src.write_bytes(b"metin")
        with pytest.raises(ValueError):
            pc.convert_to_pdfa2b(str(src))

    def test_pdf_gs_failure_falls_back_to_original(self, tmp_path, monkeypatch):
        # GS timeout/çökme durumunda PDF kaynak orijinal haliyle döner, hata yok
        import pdf.pdf_converter as pc
        src = tmp_path / "buyuk_tarama.pdf"
        src.write_bytes(b"%PDF-1.4 sahte")

        def _boom(s, o, deadline=None):
            raise Exception("PDF/A-2b dönüşümü timeout (240s)")

        monkeypatch.setattr(pc, "_pdf_to_pdfa2b", _boom)
        assert pc.convert_to_pdfa2b(str(src)) == str(src)

    def test_image_gs_failure_falls_back_to_intermediate_pdf(self, tmp_path, monkeypatch):
        # Kaynak → PDF başarılı ama PDF/A adımı çökerse ara PDF ile devam edilir
        import pdf.pdf_converter as pc
        src = tmp_path / "tarama.tif"
        _make_tiff(str(src), pages=2)

        def _boom(s, o):
            raise Exception("GhostScript hatası: sahte")

        monkeypatch.setattr(pc, "_pdf_to_pdfa2b", _boom)
        result = pc.convert_to_pdfa2b(str(src))
        try:
            assert result != str(src)
            assert _pdf_page_count(result) == 2
        finally:
            os.remove(result)


class TestGsTimeout:
    """Faz 5-A: GS tavanının evi config/settings.py — env boot'ta bir kez okunur.

    Eski testler env'i çağrı anında değiştiriyordu; yeni sözleşmede env
    plumbing'i taze Settings() örneğiyle, çalışma zamanı patch'i settings
    attribute'uyla doğrulanır (bozuk env toleransı settings katmanında).
    """

    def test_default_240(self):
        import pdf.pdf_converter as pc
        assert pc._gs_timeout() == 240

    def test_reads_live_settings_attribute(self, monkeypatch):
        import pdf.pdf_converter as pc
        from config.settings import settings
        monkeypatch.setattr(settings, "gs_timeout_seconds", 90)
        assert pc._gs_timeout() == 90

    def test_env_override_via_fresh_settings(self, monkeypatch):
        from config.settings import Settings
        monkeypatch.setenv("GS_TIMEOUT_SECONDS", "90")
        assert Settings().gs_timeout_seconds == 90

    def test_invalid_env_falls_back_to_default(self, monkeypatch):
        # Eski _gs_timeout ValueError-fallback sözleşmesi settings'e taşındı:
        # bozuk değer uygulamayı düşürmez, varsayılana düşer.
        from config.settings import Settings
        monkeypatch.setenv("GS_TIMEOUT_SECONDS", "hizli")
        assert Settings().gs_timeout_seconds == 240

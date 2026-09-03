"""G107 — Teslim defteri (`aktarim_teslimleri`) + `services/teslim_kutusu.py`
durum makinesi: kaydet / doğrula / kuru koş / kapı / uygula / açılış toparlama.

Plan: docs/plan/veri-teslim-otomasyonu-plani-2026-09-03.md §2.1–2.2.

**TEST VERİSİ KURALI (A.2 dersi):** gerçek teslim paketi REPOYA GİRMEZ; bütün
paketler openpyxl ile SENTETİK üretilir (test_g064 düzeni) — `Sheet` +
isteğe bağlı `DEGISIKLIK_OZETI` sayfası, uydurma satırlar.

Katmanlar:

1. **Birim** — model/migrasyon sözleşmesi, özet sayfası ayrıştırıcı, eşikler.
2. **sqlite (StaticPool)** — durum makinesi uçtan uca (G064 fixture reçetesi:
   FK + ÇALIŞAN SAVEPOINT; defter index'leri migrasyon op'undan uygulanır).
3. **dbtest (gerçek Postgres, scratch DB)** — 3-ortam kuralı: DB yoksa SKIP;
   tablo `create_all`'dan, kısmi UNIQUE + partial index migrasyondan geliyor,
   ikinci `init_db` idempotent.
"""
import io
import logging
import os
from pathlib import Path

import pytest
from sqlalchemy import create_engine, event, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import NullPool, StaticPool

import models
from database import _MIGRATIONS, Base
from scripts import hukdok_aktarim
from services import teslim_kutusu as tk
from test_migration_path import _live_indexes, _run_init_db, _scratch_database

BASLIKLAR = ["SistemNo", "TKU", "Hasar No", "Dosya No", "Arşiv Tarihi", "Tıbbi Olay"]


def _paket(satirlar, *, basliklar=None, veri_sayfasi=tk.VERI_SAYFASI,
           ozet=None, ozet_sayfasi=False, ozet_satirlari=None) -> bytes:
    """Sentetik teslim paketi (bayt). `ozet` verilirse DEGISIKLIK_OZETI sayfasına
    "Önceki teslim | <ozet>" satırı yazılır; `ozet_satirlari` ham satır listesidir."""
    from openpyxl import Workbook

    kullanilan = list(basliklar if basliklar is not None else BASLIKLAR)
    wb = Workbook()
    ws = wb.active
    ws.title = veri_sayfasi
    ws.append(kullanilan)
    for satir in satirlar:
        ws.append([satir.get(b) for b in kullanilan])
    if ozet is not None or ozet_sayfasi or ozet_satirlari is not None:
        oz = wb.create_sheet(tk.OZET_SAYFASI)
        if ozet_satirlari is not None:
            for satir in ozet_satirlari:
                oz.append(list(satir))
        else:
            oz.append(["Alan", "Değer"])
            oz.append(["Bu teslim", "HUKDOK_TESLIM_SIMDI.xlsx · 3 satır × 6 sütun"])
            if ozet is not None:
                oz.append(["Önceki teslim", ozet])
    tampon = io.BytesIO()
    wb.save(tampon)
    wb.close()
    return tampon.getvalue()


def _satir(sistem_no, dosya_no, **extra):
    temel = {"SistemNo": sistem_no, "Dosya No": dosya_no, "TKU": "TKU-100"}
    temel.update(extra)
    return temel


def _iki_satir(damga="a"):
    """Yalnız eşleşen iki satır (kapı eşiklerinin İÇİNDE kalan teslim)."""
    return [
        _satir("SSTMN-1", "D-1", **{"Arşiv Tarihi": "15.03.2021", "Hasar No": f"H-{damga}"}),
        _satir("SSTMN-2", "D-2", **{"Tıbbi Olay": f"Enfeksiyon {damga}"}),
    ]


def _uc_satir(damga="a"):
    """İki eşleşen + bir eşleşmeyen satır (okunan 3, işlenen 2, atlanan 1 —
    eşleşmeyen oranı 1/3, kapı eşiğinin DIŞINDA)."""
    return _iki_satir(damga) + [_satir("SSTMN-9", "D-YOK")]


# ═══════════════════════════════════════════════════════════════════════════
# 1. Birim — sözleşme kilitleri, ayrıştırıcı, eşikler
# ═══════════════════════════════════════════════════════════════════════════

def _index_ops(table):
    return [sql for op in _MIGRATIONS if op[0] == "index" and op[1] == table for sql in op[2]]


def test_model_sozlesme_kolonlari():
    """Görev sözleşmesindeki kolon listesi modelde birebir (G108 şeması buna yazıldı)."""
    kolonlar = {c.name for c in models.AktarimTeslimi.__table__.columns}
    beklenen = {
        "id", "dosya_adi", "sha256", "kaynak", "sharepoint_item_id", "spool_path",
        "durum", "durum_gecmisi", "onceki_teslim_adi", "zincir_tamam",
        "okunan", "islenen", "atlanan", "hata_sayisi", "alan_degisikligi", "kart_degisen",
        "envanter_denk", "kapi_karari", "kapi_gerekcesi", "rapor_dizini", "cevap_yuklendi",
        "uygulayan", "hata_mesaji", "created_at", "updated_at", "done_at",
    }
    assert kolonlar == beklenen
    assert models.AktarimTeslimi.__tablename__ == "aktarim_teslimleri"
    assert "aktarim_teslimleri" in Base.metadata.tables          # create_all yaratır


def test_migrasyon_index_oplari_kosulsuz_index_opunda():
    """G041 kuralı: UNIQUE + partial index ("index", ...) op'unda, IF NOT EXISTS ile;
    ("table", ...) op'u YOK (create_all tabloyu yarattığı için ölü kod olurdu)."""
    sqls = _index_ops("aktarim_teslimleri")
    assert len(sqls) == 2
    assert all("IF NOT EXISTS" in sql for sql in sqls)
    unique = next(s for s in sqls if "uq_aktarim_teslimleri_sha256" in s)
    assert "UNIQUE" in unique and "WHERE durum <> 'yinelenen'" in unique
    bekleyen = next(s for s in sqls if "idx_aktarim_teslimleri_bekleyen" in s)
    for durum in tk.BEKLEYEN_DURUMLAR:
        assert f"'{durum}'" in bekleyen
    assert not [op for op in _MIGRATIONS if op[0] == "table" and op[1] == "aktarim_teslimleri"]


def test_durum_kumeleri_plan_ile_uyumlu():
    assert set(tk.DURUMLAR) == {
        "alindi", "yinelenen", "reddedildi", "dogrulandi", "kuru_kosuldu",
        "inceleme_bekliyor", "uygulaniyor", "uygulandi", "basarisiz",
    }
    assert tk.NIHAI_DURUMLAR == {"yinelenen", "reddedildi", "uygulandi", "basarisiz"}
    assert tk.KAPI_ESIKLERI is tk.kapi_esikleri


@pytest.mark.parametrize("satirlar,beklenen", [
    ([["Önceki teslim", "HUKDOK_TESLIM_A.xlsx · 8.409 satır × 68 sütun"]], "HUKDOK_TESLIM_A.xlsx"),
    ([["Onceki Teslim:", "HUKDOK_TESLIM_B.xlsx"]], "HUKDOK_TESLIM_B.xlsx"),
    ([["Önceki teslim: HUKDOK_TESLIM_C.xlsx · 10 satır"]], "HUKDOK_TESLIM_C.xlsx"),
    ([["Bu teslim", "X.xlsx"], [None, None], ["ÖNCEKİ TESLİM", None, "HUKDOK_TESLIM_D.xlsx"]],
     "HUKDOK_TESLIM_D.xlsx"),
    ([["Önceki teslim", "—"]], None),
    ([["Önceki teslim", None]], None),
    ([["Bu teslim", "X.xlsx"]], None),
])
def test_onceki_teslim_adi_ayristirma(satirlar, beklenen):
    """"Önceki teslim" etiketi aksan/boşluk/iki nokta duyarsız; değer aynı
    hücrede ya da sonraki dolu hücrede; `·` sonrası atılır; yer tutucu None."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for satir in satirlar:
        ws.append(satir)
    assert tk.onceki_teslim_adi_oku(ws) == beklenen


def test_kapi_esikleri_envden_okunur(monkeypatch):
    monkeypatch.delenv("TESLIM_KAPI_HATA_ORANI", raising=False)
    monkeypatch.delenv("TESLIM_KAPI_ESLESMEYEN_ORANI", raising=False)
    monkeypatch.delenv("TESLIM_KAPI_ALAN_DEGISIKLIGI", raising=False)
    assert tk.kapi_esikleri() == {"hata_orani": 0.02, "eslesmeyen_orani": 0.05, "alan_degisikligi": 10000}

    monkeypatch.setenv("TESLIM_KAPI_HATA_ORANI", "0.10")
    monkeypatch.setenv("TESLIM_KAPI_ESLESMEYEN_ORANI", "0.25")
    monkeypatch.setenv("TESLIM_KAPI_ALAN_DEGISIKLIGI", "500")
    assert tk.KAPI_ESIKLERI() == {"hata_orani": 0.10, "eslesmeyen_orani": 0.25, "alan_degisikligi": 500}


def test_kapi_esigi_bozuk_env_varsayilana_duser(monkeypatch, caplog):
    monkeypatch.setenv("TESLIM_KAPI_HATA_ORANI", "yüzde iki")
    with caplog.at_level(logging.WARNING, logger="services.teslim_kutusu"):
        assert tk.kapi_esikleri()["hata_orani"] == 0.02
    assert any("TESLIM_KAPI_HATA_ORANI" in r.getMessage() for r in caplog.records)


@pytest.mark.parametrize("ham,beklenen", [
    ("HUKDOK_TESLIM_X.xlsx", "HUKDOK_TESLIM_X.xlsx"),
    ("../../etc/passwd", "passwd"),
    ("C:\\Users\\x\\teslim.xlsx", "teslim.xlsx"),
    ("a:b*c?.xlsx", "a_b_c_.xlsx"),
    ("", "teslim.xlsx"),
])
def test_spool_adi_guvenli(ham, beklenen):
    assert tk._guvenli_ad(ham) == beklenen


def test_timeout_sifirla_yalniz_postgreste_reset_atar():
    """`aktarimi_kos` statement_timeout'u oturum boyu yükseltir; bağlantı
    havuza dönmeden RESET atılmalı (yalnız PG — sqlite'ta dokunulmaz)."""
    class _Dialect:
        def __init__(self, name):
            self.name = name

    class _Conn:
        def __init__(self, name):
            self.dialect = _Dialect(name)
            self.calisan = []
            self.commit_sayisi = 0

        def rollback(self):
            self.calisan.append("ROLLBACK")

        def execute(self, stmt):
            self.calisan.append(str(stmt))

        def commit(self):
            self.commit_sayisi += 1

    pg = _Conn("postgresql")
    tk._timeout_sifirla(pg)
    assert pg.calisan == ["ROLLBACK", "RESET statement_timeout"] and pg.commit_sayisi == 1

    lite = _Conn("sqlite")
    tk._timeout_sifirla(lite)
    assert lite.calisan == [] and lite.commit_sayisi == 0


# ═══════════════════════════════════════════════════════════════════════════
# 2. Davranış — sqlite (G064 reçetesi: FK + ÇALIŞAN SAVEPOINT + defter index'leri)
# ═══════════════════════════════════════════════════════════════════════════

@pytest.fixture()
def spool(tmp_path, monkeypatch):
    """Spool + rapor dizinleri test dizinine (gerçek /app/data kirlenmesin)."""
    dizin = tmp_path / "teslim_spool"
    monkeypatch.setenv("TESLIM_SPOOL_DIR", str(dizin))
    for ad in ("TESLIM_KAPI_HATA_ORANI", "TESLIM_KAPI_ESLESMEYEN_ORANI", "TESLIM_KAPI_ALAN_DEGISIKLIGI"):
        monkeypatch.delenv(ad, raising=False)
    return dizin


@pytest.fixture()
def db_env(spool):
    """In-memory sqlite + `case_foys` ve `aktarim_teslimleri` migrasyon index'leri
    + FK + çalışan SAVEPOINT (pysqlite BEGIN reçetesi — gerekçe test_g064)."""
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _fk_ac(dbapi_connection, _record):
        dbapi_connection.isolation_level = None      # pysqlite BEGIN yaymasın
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    @event.listens_for(engine, "begin")
    def _begin(conn):
        conn.exec_driver_sql("BEGIN")

    Base.metadata.create_all(engine)
    with engine.begin() as conn:
        for tablo in ("case_foys", "aktarim_teslimleri"):
            for sql in _index_ops(tablo):
                conn.execute(text(sql))
    yield sessionmaker(bind=engine, autocommit=False, autoflush=False)
    engine.dispose()


def _kart(db, tracking, klasor, **extra):
    case = models.Case(tracking_no=tracking, status="DERDEST", klasor_no_2=klasor, **extra)
    db.add(case)
    db.flush()
    return case


@pytest.fixture()
def iki_kart(db_env):
    db = db_env()
    try:
        for i in (1, 2):
            _kart(db, f"HA.G107.{i}", f"D-{i}")
        db.commit()
    finally:
        db.close()
    return db_env


def _defter(db, **alanlar):
    """Doğrudan defter satırı (kapı testleri kuru koşu koşturmadan sayaç verir)."""
    temel = dict(dosya_adi="HUKDOK_TESLIM_T.xlsx", sha256="0" * 64, kaynak="yukleme",
                 durum=tk.DURUM_KURU_KOSULDU, durum_gecmisi=[],
                 okunan=100, islenen=99, atlanan=1, hata_sayisi=0,
                 alan_degisikligi=50, kart_degisen=40, envanter_denk=True, zincir_tamam=True)
    temel.update(alanlar)
    teslim = models.AktarimTeslimi(**temel)
    db.add(teslim)
    db.commit()
    return teslim.id


def _foy_sayisi(fabrika):
    db = fabrika()
    try:
        return db.query(models.CaseFoy).count()
    finally:
        db.close()


# ─── teslim_kaydet ───────────────────────────────────────────────────────────

def test_kaydet_yeni_icerik_alindi_spoola_yazar(iki_kart, spool):
    db = iki_kart()
    try:
        icerik = _paket(_uc_satir())
        tid = tk.teslim_kaydet(icerik=icerik, dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db)
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.durum == "alindi"
        assert teslim.sha256 and len(teslim.sha256) == 64
        assert Path(teslim.spool_path) == spool / f"{tid}_HUKDOK_TESLIM_A.xlsx"
        assert Path(teslim.spool_path).read_bytes() == icerik
        assert teslim.durum_gecmisi[0]["durum"] == "alindi" and teslim.durum_gecmisi[0]["at"]
        assert teslim.cevap_yuklendi is False and teslim.done_at is None
    finally:
        db.close()


def test_kaydet_ayni_icerik_yinelenen_spoolsuz_ilk_id_notta(iki_kart, spool):
    """Kabul: aynı içerik ikinci kez → yeni satır `yinelenen`, spool'a yazılmaz,
    notunda ilk id var; farklı içerik aynı ad → `alindi`."""
    db = iki_kart()
    try:
        icerik = _paket(_uc_satir())
        ilk = tk.teslim_kaydet(icerik=icerik, dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db)
        ikinci = tk.teslim_kaydet(icerik=icerik, dosya_adi="HUKDOK_TESLIM_A (1).xlsx",
                                  kaynak="sharepoint", sharepoint_item_id="item-2", db=db)
        assert ikinci != ilk
        yinelenen = db.get(models.AktarimTeslimi, ikinci)
        assert yinelenen.durum == "yinelenen"
        assert yinelenen.spool_path is None
        assert f"#{ilk}" in yinelenen.durum_gecmisi[-1]["not"]
        assert yinelenen.done_at is not None                      # nihai
        assert yinelenen.sharepoint_item_id == "item-2"
        assert len(list(spool.glob("*.xlsx"))) == 1              # spool'da tek dosya
        assert db.get(models.AktarimTeslimi, ilk).durum == "alindi"   # mevcut satıra dokunulmadı

        ucuncu = tk.teslim_kaydet(icerik=_paket(_uc_satir("b")), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                                  kaynak="yukleme", db=db)
        assert db.get(models.AktarimTeslimi, ucuncu).durum == "alindi"
    finally:
        db.close()


def test_kaydet_kaynak_dogrulanir(iki_kart):
    db = iki_kart()
    try:
        with pytest.raises(ValueError, match="kaynak"):
            tk.teslim_kaydet(icerik=b"x", dosya_adi="a.xlsx", kaynak="whatsapp", db=db)
    finally:
        db.close()


def test_kismi_unique_index_yinelenen_disini_tekler(iki_kart):
    """Migrasyon index'i sqlite'ta da uygulanıyor: aynı sha256 ile ikinci
    yinelenen-dışı satır IntegrityError; yinelenen satır serbest."""
    db = iki_kart()
    try:
        _defter(db, sha256="a" * 64, durum="alindi")
        _defter(db, sha256="a" * 64, durum="yinelenen")          # serbest
        with pytest.raises(IntegrityError):
            _defter(db, sha256="a" * 64, durum="dogrulandi")
        db.rollback()
    finally:
        db.close()


# ─── teslim_dogrula ──────────────────────────────────────────────────────────

def test_dogrula_sheet_yoksa_reddedildi(iki_kart, caplog):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir(), veri_sayfasi="Föyler"),
                               dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db)
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_dogrula(tid, db=db) == "reddedildi"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert "'Sheet' sayfası yok" in teslim.hata_mesaji and "Föyler" in teslim.hata_mesaji
        assert teslim.done_at is not None
        assert not [r for r in caplog.records if r.levelno >= logging.ERROR]   # reddedildi = WARNING
        assert any("reddedildi" in r.getMessage() for r in caplog.records)
        with pytest.raises(ValueError):                           # nihai durumdan yeniden doğrulama yok
            tk.teslim_dogrula(tid, db=db)
    finally:
        db.close()


def test_dogrula_zorunlu_baslik_yoksa_reddedildi(iki_kart):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(
            icerik=_paket([{"SistemNo": "S-1"}], basliklar=["SistemNo", "TKU"]),
            dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslim_dogrula(tid, db=db) == "reddedildi"
        assert "dosya_no" in db.get(models.AktarimTeslimi, tid).hata_mesaji

        tid2 = tk.teslim_kaydet(
            icerik=_paket([{"Dosya No": "D-1"}], basliklar=["Dosya No", "TKU"]),
            dosya_adi="HUKDOK_TESLIM_B.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslim_dogrula(tid2, db=db) == "reddedildi"
        assert "sistem_no" in db.get(models.AktarimTeslimi, tid2).hata_mesaji
    finally:
        db.close()


def test_dogrula_bozuk_dosya_reddedildi(iki_kart):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=b"bu bir xlsx degil", dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid, db=db) == "reddedildi"
        assert "açılamadı" in db.get(models.AktarimTeslimi, tid).hata_mesaji
    finally:
        db.close()


def test_dogrula_ozet_sayfasi_yoksa_zincir_null(iki_kart):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir()), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid, db=db) == "dogrulandi"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.zincir_tamam is None and teslim.onceki_teslim_adi is None
        assert teslim.durum_gecmisi[-1]["durum"] == "dogrulandi"
    finally:
        db.close()


def test_dogrula_onceki_teslim_defterde_uygulandi_ise_zincir_tamam(iki_kart):
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="uygulandi")
        tid = tk.teslim_kaydet(
            icerik=_paket(_uc_satir(), ozet="HUKDOK_TESLIM_ONCEKI.xlsx · 3 satır × 6 sütun"),
            dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslim_dogrula(tid, db=db) == "dogrulandi"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.onceki_teslim_adi == "HUKDOK_TESLIM_ONCEKI.xlsx"
        assert teslim.zincir_tamam is True
    finally:
        db.close()


def test_dogrula_onceki_teslim_uygulanmamis_ise_zincir_eksik(iki_kart):
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="inceleme_bekliyor")
        tid = tk.teslim_kaydet(
            icerik=_paket(_uc_satir(), ozet="HUKDOK_TESLIM_ONCEKI.xlsx · 3 satır"),
            dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslim_dogrula(tid, db=db) == "dogrulandi"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.onceki_teslim_adi == "HUKDOK_TESLIM_ONCEKI.xlsx"
        assert teslim.zincir_tamam is False

        # Özet sayfası var ama "Önceki teslim" boş/yer tutucu: zincir doğrulanamadı → False
        tid2 = tk.teslim_kaydet(
            icerik=_paket(_uc_satir("z"), ozet="—"),
            dosya_adi="HUKDOK_TESLIM_B.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslim_dogrula(tid2, db=db) == "dogrulandi"
        teslim2 = db.get(models.AktarimTeslimi, tid2)
        assert teslim2.onceki_teslim_adi is None and teslim2.zincir_tamam is False
    finally:
        db.close()


# ─── teslim_kuru_kos ─────────────────────────────────────────────────────────

def test_kuru_kos_sayaclar_deftere_dbye_yazilmaz(iki_kart, spool):
    """Kabul: sayaçlar defterde, `case_foys` boş kalır, `rapor_dizini` dolu."""
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir()), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        with pytest.raises(ValueError, match="kuru koşu"):        # alindi'dan doğrudan kuru koşu yok
            tk.teslim_kuru_kos(tid, db=db)
        assert tk.teslim_dogrula(tid, db=db) == "dogrulandi"
        assert tk.teslim_kuru_kos(tid, db=db) == "kuru_kosuldu"

        teslim = db.get(models.AktarimTeslimi, tid)
        assert (teslim.okunan, teslim.islenen, teslim.atlanan, teslim.hata_sayisi) == (3, 2, 1, 0)
        assert teslim.alan_degisikligi >= 2 and teslim.kart_degisen == 2
        assert teslim.envanter_denk is True
        assert teslim.rapor_dizini == str(spool / f"{tid}_raporlar")
        rapor = Path(teslim.rapor_dizini)
        assert rapor.is_dir()
        assert (rapor / "kuru-kosu-ozeti.txt").read_text(encoding="utf-8").startswith("=")
        assert any(p.name.startswith("satir-raporu_") for p in rapor.iterdir())   # atlanan satır rapora düştü
        assert teslim.durum_gecmisi[-1]["durum"] == "kuru_kosuldu"
    finally:
        db.close()
    assert _foy_sayisi(iki_kart) == 0                            # KURU koşu: hiçbir föy yazılmadı


def test_kuru_kos_istisnasi_basarisiz_tek_error(iki_kart, monkeypatch, caplog):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir()), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        tk.teslim_dogrula(tid, db=db)

        def _patla(*_a, **_k):
            raise RuntimeError("disk dolu")

        monkeypatch.setattr(hukdok_aktarim, "aktarimi_kos", _patla)
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_kuru_kos(tid, db=db) == "basarisiz"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert "RuntimeError: disk dolu" in teslim.hata_mesaji
        errors = [r for r in caplog.records if r.levelno >= logging.ERROR]
        assert len(errors) == 1 and f"#{tid}" in errors[0].getMessage()
    finally:
        db.close()


# ─── kapi_degerlendir ────────────────────────────────────────────────────────

def test_kapi_bos_defterde_inceleme_ilk_teslim(iki_kart):
    db = iki_kart()
    try:
        tid = _defter(db)
        assert tk.kapi_degerlendir(tid, db=db) == "inceleme"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.durum == "inceleme_bekliyor"
        assert teslim.kapi_karari == "inceleme" and "ilk_teslim" in teslim.kapi_gerekcesi
        assert teslim.durum_gecmisi[-1]["durum"] == "inceleme_bekliyor"
    finally:
        db.close()


def test_kapi_onceki_uygulandi_ve_esik_ici_otomatik(iki_kart):
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="uygulandi")
        tid = _defter(db)                                          # zincir_tamam=True, sayaçlar eşik içi
        assert tk.kapi_degerlendir(tid, db=db) == "otomatik"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.durum == "kuru_kosuldu"                      # durum DEĞİŞMEZ; uygulama ayrı adım
        assert teslim.kapi_karari == "otomatik" and teslim.kapi_gerekcesi is None
    finally:
        db.close()


@pytest.mark.parametrize("alanlar,kural", [
    ({"envanter_denk": False}, "envanter_denk_degil"),
    ({"envanter_denk": None}, "envanter_denk_degil"),
    ({"zincir_tamam": False}, "zincir_eksik"),
    ({"okunan": 100, "hata_sayisi": 3}, "hata_orani"),             # 0.03 > 0.02
    ({"okunan": 100, "atlanan": 6}, "eslesmeyen_orani"),           # 0.06 > 0.05
    ({"alan_degisikligi": 10001}, "alan_degisikligi"),
    ({"okunan": 0, "islenen": 0, "atlanan": 0}, "bos_teslim"),
])
def test_kapi_her_esik_ihlali_tek_tek_inceleme(iki_kart, alanlar, kural):
    """Kabul: her kural tek başına `inceleme` üretir ve gerekçe metni kuralın adını taşır."""
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="uygulandi")
        tid = _defter(db, **alanlar)
        assert tk.kapi_degerlendir(tid, db=db) == "inceleme"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.durum == "inceleme_bekliyor"
        assert kural in teslim.kapi_gerekcesi
        assert kural in tk.KAPI_KURALLARI
    finally:
        db.close()


def test_kapi_ihlaller_birlesir_hepsi_degerlendirilir(iki_kart):
    """İlk ihlalde durmaz: gerekçe `;` ile birleşir (admin hepsini görsün)."""
    db = iki_kart()
    try:
        tid = _defter(db, envanter_denk=False, zincir_tamam=False, alan_degisikligi=20000)
        assert tk.kapi_degerlendir(tid, db=db) == "inceleme"
        gerekce = db.get(models.AktarimTeslimi, tid).kapi_gerekcesi
        for kural in ("envanter_denk_degil", "ilk_teslim", "zincir_eksik", "alan_degisikligi"):
            assert kural in gerekce
        assert gerekce.count(";") == 3
    finally:
        db.close()


def test_kapi_esikleri_envden_uygulanir(iki_kart, monkeypatch):
    """Kabul: eşik env'den ÇAĞRI ANINDA okunur — aynı sayaçlar gevşek eşikle geçer."""
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="uygulandi")
        tid = _defter(db, okunan=100, hata_sayisi=3, atlanan=6, alan_degisikligi=10001)
        assert tk.kapi_degerlendir(tid, db=db) == "inceleme"

        monkeypatch.setenv("TESLIM_KAPI_HATA_ORANI", "0.05")
        monkeypatch.setenv("TESLIM_KAPI_ESLESMEYEN_ORANI", "0.10")
        monkeypatch.setenv("TESLIM_KAPI_ALAN_DEGISIKLIGI", "20000")
        assert tk.kapi_degerlendir(tid, db=db) == "otomatik"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.kapi_karari == "otomatik" and teslim.kapi_gerekcesi is None
        assert teslim.durum == "inceleme_bekliyor"                 # geriye dönüş YOK (tek yönlü)
    finally:
        db.close()


def test_kapi_yanlis_durumdan_value_error(iki_kart):
    db = iki_kart()
    try:
        tid = _defter(db, durum="alindi")
        with pytest.raises(ValueError, match="kapı"):
            tk.kapi_degerlendir(tid, db=db)
    finally:
        db.close()


# ─── teslim_uygula + teslimi_isle ────────────────────────────────────────────

def test_uctan_uca_ilk_teslim_inceleme_elle_uygula_ikinci_gece_otomatik(iki_kart, spool):
    """Kabul: sqlite uçtan uca — ilk teslim kapıda `inceleme_bekliyor` (defter boş),
    admin uygular → `uygulandi` + föy satırları; aynı içerik → `yinelenen`;
    zincirli ikinci teslim gece turunda kendiliğinden `uygulandi` (gece-job)."""
    db = iki_kart()
    try:
        ilk_icerik = _paket(_uc_satir())
        t1 = tk.teslim_kaydet(icerik=ilk_icerik, dosya_adi="HUKDOK_TESLIM_1.xlsx", kaynak="sharepoint",
                              sharepoint_item_id="item-1", db=db)
        assert tk.teslimi_isle(t1, otomatik_uygula=True, db=db) == "inceleme_bekliyor"
        teslim1 = db.get(models.AktarimTeslimi, t1)
        assert "ilk_teslim" in teslim1.kapi_gerekcesi
        assert teslim1.uygulayan is None
        assert [g["durum"] for g in teslim1.durum_gecmisi] == [
            "alindi", "dogrulandi", "kuru_kosuldu", "inceleme_bekliyor",
        ]
        assert db.query(models.CaseFoy).count() == 0

        assert tk.teslim_uygula(t1, uygulayan="admin@buro.test", db=db) == "uygulandi"
        teslim1 = db.get(models.AktarimTeslimi, t1)
        assert teslim1.uygulayan == "admin@buro.test" and teslim1.done_at is not None
        assert (teslim1.okunan, teslim1.islenen, teslim1.atlanan) == (3, 2, 1)
        assert teslim1.envanter_denk is True and teslim1.hata_mesaji is None
        assert [g["durum"] for g in teslim1.durum_gecmisi][-2:] == ["uygulaniyor", "uygulandi"]
        assert (Path(teslim1.rapor_dizini) / "uygulama-ozeti.txt").exists()
        assert db.query(models.CaseFoy).count() == 2
        db.expire_all()
        assert db.query(models.Case).filter_by(klasor_no_2="D-1").one().arsiv_tarihi is not None

        # Aynı içerik ikinci kez: yinelenen; teslimi_isle dokunmaz
        t1b = tk.teslim_kaydet(icerik=ilk_icerik, dosya_adi="HUKDOK_TESLIM_1.xlsx", kaynak="sharepoint",
                               sharepoint_item_id="item-1", db=db)
        assert db.get(models.AktarimTeslimi, t1b).durum == "yinelenen"
        assert tk.teslimi_isle(t1b, otomatik_uygula=True, db=db) == "yinelenen"

        # Zincirli, eşik içi ikinci teslim: önceki adı defterde uygulandi → gece otomatik
        t2 = tk.teslim_kaydet(
            icerik=_paket(_iki_satir("b"), ozet="HUKDOK_TESLIM_1.xlsx · 3 satır × 6 sütun"),
            dosya_adi="HUKDOK_TESLIM_2.xlsx", kaynak="sharepoint", sharepoint_item_id="item-2", db=db,
        )
        assert tk.teslimi_isle(t2, otomatik_uygula=True, db=db) == "uygulandi"
        teslim2 = db.get(models.AktarimTeslimi, t2)
        assert (teslim2.okunan, teslim2.islenen, teslim2.atlanan, teslim2.hata_sayisi) == (2, 2, 0, 0)
        assert teslim2.zincir_tamam is True and teslim2.kapi_karari == "otomatik"
        assert teslim2.uygulayan == tk.GECE_UYGULAYAN == "gece-job"
        db.expire_all()
        assert db.query(models.CaseFoy).count() == 2               # aynı SistemNo'lar: föy ikilenmedi
        assert db.query(models.Case).filter_by(klasor_no_2="D-2").one().tibbi_olay == "Enfeksiyon b"
    finally:
        db.close()


def test_isle_otomatik_uygula_kapali_kuru_kosuldu_kalir(iki_kart):
    """Boot telafisi / admin yükleme yolu: kapı otomatik dese bile uygulama YOK."""
    db = iki_kart()
    try:
        _defter(db, dosya_adi="HUKDOK_TESLIM_ONCEKI.xlsx", sha256="b" * 64, durum="uygulandi")
        tid = tk.teslim_kaydet(
            icerik=_paket(_iki_satir(), ozet="HUKDOK_TESLIM_ONCEKI.xlsx"),
            dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db,
        )
        assert tk.teslimi_isle(tid, otomatik_uygula=False, db=db) == "kuru_kosuldu"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert teslim.kapi_karari == "otomatik" and teslim.uygulayan is None
    finally:
        db.close()
    assert _foy_sayisi(iki_kart) == 0


def test_isle_reddedilen_teslimde_durur(iki_kart):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir(), veri_sayfasi="Veri"),
                               dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db)
        assert tk.teslimi_isle(tid, otomatik_uygula=True, db=db) == "reddedildi"
        assert db.get(models.AktarimTeslimi, tid).okunan is None   # kuru koşuya hiç gelmedi
    finally:
        db.close()


def test_uygula_yanlis_durumdan_value_error(iki_kart):
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir()), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        with pytest.raises(ValueError, match="uygulama"):
            tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db)
        tk.teslim_dogrula(tid, db=db)
        with pytest.raises(ValueError, match="uygulama"):
            tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db)
        assert db.get(models.AktarimTeslimi, tid).durum == "dogrulandi"
        with pytest.raises(ValueError, match="Teslim yok"):
            tk.teslim_uygula(9999, uygulayan="admin@buro.test", db=db)
    finally:
        db.close()
    assert _foy_sayisi(iki_kart) == 0


def test_uygula_istisna_basarisiz_hata_mesaji_tek_error(iki_kart, monkeypatch, caplog):
    """Kabul: `aktarimi_kos` istisna fırlatınca `basarisiz` + `hata_mesaji`;
    log sözleşmesi: teslim başına TEK ERROR."""
    db = iki_kart()
    try:
        tid = tk.teslim_kaydet(icerik=_paket(_uc_satir()), dosya_adi="HUKDOK_TESLIM_A.xlsx",
                               kaynak="yukleme", db=db)
        tk.teslimi_isle(tid, otomatik_uygula=False, db=db)
        assert db.get(models.AktarimTeslimi, tid).durum == "inceleme_bekliyor"

        def _patla(*_a, **_k):
            raise RuntimeError("bağlantı koptu")

        monkeypatch.setattr(hukdok_aktarim, "aktarimi_kos", _patla)
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db) == "basarisiz"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert "RuntimeError: bağlantı koptu" in teslim.hata_mesaji
        assert teslim.done_at is not None and teslim.uygulayan == "admin@buro.test"
        assert [g["durum"] for g in teslim.durum_gecmisi][-2:] == ["uygulaniyor", "basarisiz"]
        errors = [r for r in caplog.records if r.levelno >= logging.ERROR]
        assert len(errors) == 1 and f"#{tid}" in errors[0].getMessage()
        with pytest.raises(ValueError):                             # nihai: yeniden uygulama yok
            tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db)
    finally:
        db.close()


def test_uygula_envanter_kapisi_basarisiz_ikinci_error_yok(db_env, spool, monkeypatch, caplog):
    """Belge koruma şartı: envanter denk değilse script koşuyu geri alır ve
    ERROR'u kendisi basar; servis `basarisiz` yazar, İKİNCİ ERROR eklemez."""
    db = db_env()
    try:
        case = _kart(db, "HA.G107.B1", "D-1")
        taraf = models.CaseParty(case_id=case.id, name="Ali V.", role="Davacı", party_type="CLIENT")
        db.add(taraf)
        db.flush()
        db.add(models.CaseDocument(
            case_id=case.id, case_party_id=taraf.id,
            original_filename="tensip.pdf", stored_filename="TENSIP.pdf",
            sharepoint_url="https://sp/tensip.pdf", link_mode="LINKED",
        ))
        db.commit()
        taraf_id = taraf.id

        tid = tk.teslim_kaydet(icerik=_paket([_satir("SSTMN-1", "D-1")]),
                               dosya_adi="HUKDOK_TESLIM_A.xlsx", kaynak="yukleme", db=db)
        assert tk.teslimi_isle(tid, otomatik_uygula=False, db=db) == "inceleme_bekliyor"

        gercek = hukdok_aktarim._kart_alanlarini_yaz

        def _bagi_kopar(db, case, satir, source, **kwargs):
            db.query(models.CaseDocument).filter(
                models.CaseDocument.case_id == case.id
            ).update({"case_party_id": None}, synchronize_session=False)
            return gercek(db, case, satir, source, **kwargs)

        monkeypatch.setattr(hukdok_aktarim, "_kart_alanlarini_yaz", _bagi_kopar)
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db) == "basarisiz"
        teslim = db.get(models.AktarimTeslimi, tid)
        assert "envanter" in teslim.hata_mesaji and teslim.envanter_denk is False
        errors = [r for r in caplog.records if r.levelno >= logging.ERROR]
        assert len(errors) == 1 and errors[0].name != "services.teslim_kutusu"
        db.expire_all()
        assert db.query(models.CaseDocument).one().case_party_id == taraf_id   # bağ yerinde
        assert db.query(models.CaseFoy).count() == 0                            # koşu geri alındı
    finally:
        db.close()


# ─── acilis_toparla ──────────────────────────────────────────────────────────

def test_acilis_toparla_uygulaniyor_inceleme_bekliyora_duser(iki_kart, caplog):
    db = iki_kart()
    try:
        kesik = _defter(db, sha256="c" * 64, durum="uygulaniyor", kapi_gerekcesi=None)
        saglam = _defter(db, sha256="d" * 64, durum="kuru_kosuldu")
        bitmis = _defter(db, sha256="e" * 64, durum="uygulandi")
        with caplog.at_level(logging.WARNING):
            assert tk.acilis_toparla(db=db) == 1
        teslim = db.get(models.AktarimTeslimi, kesik)
        assert teslim.durum == "inceleme_bekliyor"
        assert teslim.kapi_karari == "inceleme" and "uygulama_kesildi" in teslim.kapi_gerekcesi
        assert "uygulaniyor" in teslim.durum_gecmisi[-1]["not"]
        assert db.get(models.AktarimTeslimi, saglam).durum == "kuru_kosuldu"
        assert db.get(models.AktarimTeslimi, bitmis).durum == "uygulandi"
        assert any("uygulaniyor" in r.getMessage() for r in caplog.records)
        assert not [r for r in caplog.records if r.levelno >= logging.ERROR]
        assert tk.acilis_toparla(db=db) == 0                       # idempotent
    finally:
        db.close()


def test_oturumsuz_cagri_sessionlocal_acar_kapatir(monkeypatch, iki_kart):
    """`db` verilmezse fonksiyon kendi oturumunu açar ve KAPATIR (deadline_scanner deseni)."""
    kapatilan = []
    fabrika = iki_kart

    def _yeni():
        s = fabrika()
        orijinal = s.close

        def _kapat():
            kapatilan.append(True)
            orijinal()
        s.close = _kapat
        return s

    monkeypatch.setattr(tk, "SessionLocal", _yeni)
    assert tk.acilis_toparla() == 0
    assert kapatilan == [True]


# ═══════════════════════════════════════════════════════════════════════════
# 3. dbtest — gerçek Postgres (scratch DB; 3-ortam kuralı: DB yoksa SKIP)
# ═══════════════════════════════════════════════════════════════════════════

@pytest.fixture(scope="module")
def fresh_db():
    """Boş scratch DB üzerinde bir kez init_db() koşmuş engine (gerçek DB'ye yazılmaz).

    test_migration_path'teki `admin_engine` + `fresh_db` ikilisinin ikizi; fixture
    import edip yeniden bağlamak F811 üretiyor, bakım bağlantısı bu yüzden burada
    kurulur. DB yoksa/ulaşılamıyorsa SKIP (3-ortam kuralı)."""
    url = os.getenv("MIGRATION_TEST_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
    if not url.startswith("postgresql"):
        pytest.skip("MIGRATION_TEST_DATABASE_URL/DATABASE_URL postgresql:// değil")
    admin = create_engine(
        url, isolation_level="AUTOCOMMIT", poolclass=NullPool, connect_args={"connect_timeout": 3},
    )
    try:
        with admin.connect() as conn:
            conn.execute(text("SELECT 1"))
    except Exception as exc:
        admin.dispose()
        pytest.skip(f"Gerçek Postgres'e ulaşılamadı ({type(exc).__name__}) — G107 dbtest atlandı")
    try:
        with _scratch_database(admin, "g107") as engine:
            _run_init_db(engine)
            yield engine
    finally:
        admin.dispose()


@pytest.mark.dbtest
def test_pg_tablo_create_all_indexler_migrasyondan_ikinci_kosu_idempotent(fresh_db):
    """Kabul: tablo `create_all` ile geliyor, UNIQUE + partial index ("index", …)
    op'undan; ikinci init_db şemayı değiştirmiyor; kısmi UNIQUE gerçekten
    yinelenen-dışını tekliyor."""
    with fresh_db.connect() as conn:
        assert conn.execute(text("SELECT to_regclass('public.aktarim_teslimleri')")).scalar() is not None
        kolonlar = {
            r[0] for r in conn.execute(text(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name = 'aktarim_teslimleri'"
            ))
        }
    assert {"sha256", "durum", "durum_gecmisi", "cevap_yuklendi", "done_at"} <= kolonlar

    once = _live_indexes(fresh_db)
    assert "uq_aktarim_teslimleri_sha256" in once and "idx_aktarim_teslimleri_bekleyen" in once
    assert "UNIQUE" in once["uq_aktarim_teslimleri_sha256"]
    assert "yinelenen" in once["uq_aktarim_teslimleri_sha256"]      # kısmi: WHERE durum <> 'yinelenen'
    assert "WHERE" in once["idx_aktarim_teslimleri_bekleyen"]
    assert "inceleme_bekliyor" in once["idx_aktarim_teslimleri_bekleyen"]

    _run_init_db(fresh_db)                                           # ikinci koşu
    assert _live_indexes(fresh_db) == once

    ekle = (
        "INSERT INTO aktarim_teslimleri (dosya_adi, sha256, kaynak, durum, cevap_yuklendi) "
        "VALUES (:ad, 'f' || repeat('0', 63), 'yukleme', :durum, false)"
    )
    with fresh_db.connect() as conn:
        with conn.begin():
            conn.execute(text(ekle), {"ad": "a.xlsx", "durum": "alindi"})
            conn.execute(text(ekle), {"ad": "a.xlsx", "durum": "yinelenen"})      # kısmi index dışı
        with conn.begin():
            assert conn.execute(text("SELECT COUNT(*) FROM aktarim_teslimleri")).scalar() == 2
        with pytest.raises(IntegrityError):
            with conn.begin():
                conn.execute(text(ekle), {"ad": "b.xlsx", "durum": "dogrulandi"})   # aynı sha, yinelenen dışı

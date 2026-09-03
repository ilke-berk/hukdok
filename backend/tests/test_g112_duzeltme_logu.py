"""G112 — `Düzeltme_Logu` provenance + `(boş)` açık boşaltma yolu (üçlü şart) +
`DEGER_HAVUZLARI` fark raporu.

Plan: docs/plan/veri-teslim-otomasyonu-plani-2026-09-03.md §4. Sözleşme
(gorevler/gorev/G112.md "Pazarlıksız kurallar"):

* Sütun adı → kart alanı çevirisi `KART_ALANLARI`'nın ters haritası; bilinmeyen
  sütun adı yok sayılır (DEBUG, rapora düşmez). Aynı (SistemNo, alan) için en
  yeni `Tarih` kazanır.
* Provenance: değişen alanın `case_history.source` imzasına gerekçe eklenir; imza
  `HUKDOK_TESLIM_` ile başlamaya DEVAM eder (D8 kovası bozulmaz). Log satırı
  olmayan değişiklikte imza eski biçimde.
* Açık boşaltma ÜÇLÜ şartla: log `(boş)` VE `Sheet` hücresi gerçekten boş VE bizde
  dolu. Künye (stage_decisions tek yazıcı) ve içerik-modu alanları (`court`/
  `sub_type`) boşaltılmaz, rapora düşer. İkinci koşu 0 değişiklik.
* DEGER_HAVUZLARI: iki yönlü fark CSV'si (cevap paketine girer) + fark varsa admin
  bildirimi (`teslim:<id>:havuz:<alıcı>`); listeye YAZMA yok; fark yokken dosya
  ve bildirim yok.

**TEST VERİSİ KURALI (A.2 dersi):** gerçek teslim paketi REPOYA GİRMEZ; paketler
openpyxl ile SENTETİK üretilir (test_g064 düzeni + ek sayfalar).
"""
import io
import logging
from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import models
import sharepoint.sharepoint_uploader_graph as spu
from database import _MIGRATIONS, Base
from managers import case_manager
from required_fields import AKTARIM_SOURCE_PREFIX, is_aktarim_source
from scripts import hukdok_aktarim
from scripts.hukdok_aktarim import (
    BOSALTMA_ISARETI,
    CIKIS_SATIR_HATASI,
    CIKIS_TAMAM,
    DUZELTME_SAYFASI,
    aktarimi_kos,
    duzeltme_logunu_oku,
)
from services import app_settings
from services import teslim_cevap as tc
from services import teslim_kutusu as tk

ADMIN = "yonetici@hanyaloglu-acar.av.tr"
BASLIKLAR = ["SistemNo", "TKU", "Hasar No", "Dosya No", "Arşiv Tarihi", "Tıbbi Olay",
             "Yerel Mahkeme", "Esas", "Karar No"]
LOG_BASLIKLARI = ["Excel Satırı", "SistemNo", "DosyaNo", "Sütun", "Eski Değer", "Yeni Değer",
                  "Gerekçe", "Tarih"]


# ═══════════════════════════════════════════════════════════════════════════
# Sentetik paket üreticileri
# ═══════════════════════════════════════════════════════════════════════════

def _kitap(satirlar, *, basliklar=None, sayfa="Sheet", log=None, log_basliklar=None,
           havuz=None, havuz_basliklar=("Havuz / Sütun", "Değer")):
    """Sentetik paket: veri sayfası + isteğe bağlı `Düzeltme_Logu` + `DEGER_HAVUZLARI`.

    `log`: sözlük listesi (LOG_BASLIKLARI anahtarlı). `havuz`: [(havuz adı, değer)]
    uzun biçim satırları (havuz_basliklar başlığıyla)."""
    from openpyxl import Workbook

    kullanilan = list(basliklar if basliklar is not None else BASLIKLAR)
    wb = Workbook()
    ws = wb.active
    ws.title = sayfa
    ws.append(kullanilan)
    for satir in satirlar:
        ws.append([satir.get(b) for b in kullanilan])
    if log is not None:
        lb = list(log_basliklar if log_basliklar is not None else LOG_BASLIKLARI)
        wl = wb.create_sheet(DUZELTME_SAYFASI)
        wl.append(lb)
        for satir in log:
            wl.append([satir.get(b) for b in lb])
    if havuz is not None:
        wh = wb.create_sheet(tc.HAVUZ_SAYFASI)
        wh.append(list(havuz_basliklar))
        for satir in havuz:
            wh.append(list(satir))
    return wb


def _paket_yaz(yol, satirlar, **kw):
    wb = _kitap(satirlar, **kw)
    wb.save(yol)
    wb.close()
    return Path(yol)


def _paket_bayt(satirlar, **kw) -> bytes:
    wb = _kitap(satirlar, **kw)
    tampon = io.BytesIO()
    wb.save(tampon)
    wb.close()
    return tampon.getvalue()


def _satir(sistem_no, dosya_no, **extra):
    temel = {"SistemNo": sistem_no, "Dosya No": dosya_no, "TKU": "TKU-112"}
    temel.update(extra)
    return temel


def _log(sistem_no, sutun, yeni, *, gerekce="veri ekibi düzeltmesi", tarih=None, eski=None):
    return {"Excel Satırı": 5, "SistemNo": sistem_no, "DosyaNo": "D-1", "Sütun": sutun,
            "Eski Değer": eski, "Yeni Değer": yeni, "Gerekçe": gerekce, "Tarih": tarih}


# ═══════════════════════════════════════════════════════════════════════════
# 1. Birim — ters harita, okuyucu (DB yok)
# ═══════════════════════════════════════════════════════════════════════════

def test_ters_harita_kart_alanlarindan_kurulur_yeni_sozluk_yok():
    """Sütun adı → alan çevirisi KART_ALANLARI + SUTUN_ADAYLARI'ndan türetilir;
    künye kaynakları raporlanabilsin diye haritada, ama BOSALTMA_DISI'nde."""
    h = hukdok_aktarim.DUZELTME_ALAN_HARITASI
    k = hukdok_aktarim._baslik_anahtari
    assert h[k("Tıbbi Olay")] == "tibbi_olay"
    assert h[k("TIBBİ OLAY")] == "tibbi_olay"                 # aksan/büyük-küçük toleransı
    assert h[k("Arşiv Tarihi")] == "arsiv_tarihi"
    assert h[k("Yerel Mahkeme")] == "court"
    assert h[k("Esas")] == "esas_no"
    assert h[k("Karar No")] == "karar_no" and h[k("Karar Tarihi")] == "karar_tarihi"
    assert h[k("İstinaf Mahkemesi Başvuran Taraf")] == "istinaf_basvuran_taraf"
    # KART_ALANLARI'nda olmayan sütunlar haritada YOK (bilinmeyen → yok sayılır)
    assert k("İstinaf Karar Durumu") not in h
    assert k("Dava Değeri TL") not in h and k("Müvekkil") not in h
    # Haritadaki her kart alanı gerçekten KART_ALANLARI'nda ya da künye
    for alan in set(h.values()):
        assert alan in hukdok_aktarim.KART_ALANLARI or alan in hukdok_aktarim.BOSALTMA_DISI_ALANLAR
    assert hukdok_aktarim.BOSALTMA_DISI_ALANLAR >= {"karar_no", "karar_tarihi", "istinaf_basvuran_taraf",
                                                     "court", "sub_type"}


def test_duzeltme_logu_sayfa_yoksa_bos(tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")])
    assert duzeltme_logunu_oku(paket) == {}


def test_duzeltme_logu_bilinmeyen_sutun_yok_sayilir_en_yeni_tarih_kazanir(tmp_path, caplog):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")], log=[
        _log("S-1", "İstinaf Karar Durumu", "Kaldırma"),              # bilinmeyen sütun
        _log("S-1", "Tıbbi Olay", "Kanama", gerekce="eski", tarih=datetime(2026, 7, 20)),
        _log("S-1", "Tıbbi Olay", "Enfeksiyon", gerekce="yeni", tarih="02.08.2026 14:05"),
        _log("S-1", "Tıbbi Olay", "Tarihsiz", gerekce="tarihsiz"),      # en eski sayılır
        _log("S-2", "Arşiv Tarihi", "(boş)", gerekce="arşive girmedi", tarih=date(2026, 8, 1)),
        _log("", "Tıbbi Olay", "x"),                                    # SistemNo boş → atlanır
    ])
    with caplog.at_level(logging.WARNING):
        kayitlar = duzeltme_logunu_oku(paket)
    assert set(kayitlar) == {("S-1", "tibbi_olay"), ("S-2", "arsiv_tarihi")}
    k = kayitlar[("S-1", "tibbi_olay")]
    assert (k.yeni, k.gerekce, k.bosalt) == ("Enfeksiyon", "yeni", False)
    assert k.tarih == datetime(2026, 8, 2, 14, 5)
    b = kayitlar[("S-2", "arsiv_tarihi")]
    assert b.bosalt and b.gerekce == "arşive girmedi" and b.satir_no == 6
    # bilinmeyen sütun rapora/WARNING'e düşmez (20k satırlık sayfa, gürültü)
    assert not [r for r in caplog.records if r.levelno >= logging.WARNING]


@pytest.mark.parametrize("ham,beklenen", [("(boş)", True), ("(BOŞ)", True), ("( boş )", True),
                                          ("boş", False), ("", False), (None, False), ("Kanama", False)])
def test_bosalt_isareti_yalniz_parantezli_bos(tmp_path, ham, beklenen):
    """`(boş)` AÇIK işarettir; çıplak "boş" ya da boş hücre boşaltma talimatı DEĞİLDİR."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       log=[_log("S-1", "Tıbbi Olay", ham)])
    kayitlar = duzeltme_logunu_oku(paket)
    assert kayitlar[("S-1", "tibbi_olay")].bosalt is beklenen
    assert BOSALTMA_ISARETI == "(boş)"


def test_duzeltme_logu_zorunlu_baslik_yoksa_tek_warning_bos(tmp_path, caplog):
    """Sütun adı başlığı yoksa sayfa uygulanamaz: koşu DURMAZ, tek WARNING, boş harita
    (boşaltma talimatı okunamadığı için uygulanmaz — güvenli taraf)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       log=[{"SistemNo": "S-1", "Yeni Değer": "(boş)"}],
                       log_basliklar=["Excel Satırı", "SistemNo", "Eski Değer", "Yeni Değer", "Gerekçe"])
    with caplog.at_level(logging.WARNING, logger="HukdokAktarim"):
        assert duzeltme_logunu_oku(paket) == {}
    uyarilar = [r for r in caplog.records if r.levelno == logging.WARNING]
    assert len(uyarilar) == 1 and "sutun" in uyarilar[0].getMessage()


# ═══════════════════════════════════════════════════════════════════════════
# 2. sqlite — provenance, boşaltma, idempotency
# ═══════════════════════════════════════════════════════════════════════════

def _index_ops(table):
    return [sql for op in _MIGRATIONS if op[0] == "index" and op[1] == table for sql in op[2]]


def _engine():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _fk_ac(dbapi_connection, _record):
        dbapi_connection.isolation_level = None      # pysqlite BEGIN yaymasın (G064 reçetesi)
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    @event.listens_for(engine, "begin")
    def _begin(conn):
        conn.exec_driver_sql("BEGIN")

    Base.metadata.create_all(engine)
    return engine


@pytest.fixture()
def db_env():
    engine = _engine()
    with engine.begin() as conn:
        for sql in _index_ops("case_foys"):
            conn.execute(text(sql))
    yield sessionmaker(bind=engine, autocommit=False, autoflush=False)
    engine.dispose()


def _kart(db, tracking, klasor, **extra):
    case = models.Case(tracking_no=tracking, status="DERDEST", klasor_no_2=klasor, **extra)
    db.add(case)
    db.flush()
    return case


@pytest.fixture()
def dolu_kart(db_env):
    """D-1 kartı tıbbi olay + arşiv tarihi + karar no DOLU (boşaltma zemini); D-2 boş."""
    db = db_env()
    try:
        _kart(db, "HA.G112.1", "D-1", tibbi_olay="Enfeksiyon", arsiv_tarihi=date(2021, 3, 15),
              karar_no="2018/143", court="İzmir 4. İdare Mahkemesi")
        _kart(db, "HA.G112.2", "D-2")
        db.commit()
    finally:
        db.close()
    return db_env


def _kart_oku(fabrika, klasor):
    db = fabrika()
    try:
        c = db.query(models.Case).filter_by(klasor_no_2=klasor).one()
        return SimpleNamespace(id=c.id, tibbi_olay=c.tibbi_olay, arsiv_tarihi=c.arsiv_tarihi,
                               karar_no=c.karar_no, court=c.court, esas_no=c.esas_no,
                               bucket=c.missing_required_bucket)
    finally:
        db.close()


def _tarihce(fabrika, alan):
    db = fabrika()
    try:
        return [(k.old_value, k.new_value, k.source) for k in
                db.query(models.CaseHistory).filter_by(field_name=alan).order_by(models.CaseHistory.id)]
    finally:
        db.close()


def _sayim(fabrika):
    db = fabrika()
    try:
        return {m.__name__: db.query(m).count() for m in (models.Case, models.CaseFoy, models.CaseHistory)}
    finally:
        db.close()


def test_provenance_gerekce_source_imzasina_eklenir_imza_korunur(dolu_kart, tmp_path):
    """Kabul: değişen alanın tarihçesinde gerekçe var; `source` HUKDOK_TESLIM_ ile başlıyor
    (D8 kovası); log satırı olmayan değişiklikte imza eski biçimde."""
    paket = _paket_yaz(tmp_path / "duzeltme.xlsx", [
        _satir("S-1", "D-1", **{"Tıbbi Olay": "Kanama", "Arşiv Tarihi": "20.04.2022"}),
    ], log=[_log("S-1", "Tıbbi Olay", "Kanama", gerekce="epikrizde kanama yazıyor", eski="Enfeksiyon")])

    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")

    assert sonuc.cikis_kodu == CIKIS_TAMAM and sonuc.alan_degisikligi == 2 and sonuc.bosaltilan == 0
    (tibbi,) = _tarihce(dolu_kart, "tibbi_olay")
    assert tibbi[:2] == ("Enfeksiyon", "Kanama")
    assert tibbi[2].startswith(sonuc.kaynak_imzasi) and tibbi[2].startswith(AKTARIM_SOURCE_PREFIX)
    assert tibbi[2] == f"{sonuc.kaynak_imzasi} · gerekçe: epikrizde kanama yazıyor"
    assert is_aktarim_source(tibbi[2])                          # kova testi
    (arsiv,) = _tarihce(dolu_kart, "arsiv_tarihi")
    assert arsiv[2] == sonuc.kaynak_imzasi                       # log satırı yok → eski biçim
    assert _kart_oku(dolu_kart, "D-1").bucket == "AKTARIM"


def test_provenance_uzun_gerekce_source_sinirina_kirpilir(dolu_kart, tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1", **{"Tıbbi Olay": "Kanama"})],
                       log=[_log("S-1", "Tıbbi Olay", "Kanama", gerekce="x" * 500)])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    (tibbi,) = _tarihce(dolu_kart, "tibbi_olay")
    assert len(tibbi[2]) == hukdok_aktarim._SOURCE_SINIRI and tibbi[2].startswith(sonuc.kaynak_imzasi)


def test_bosaltma_uclu_sart_null_tarihce_ikinci_kosu_sifir(dolu_kart, tmp_path):
    """Kabul: log (boş) + Sheet hücresi boş + bizde dolu → NULL + "boşaltıldı: <gerekçe>";
    ikinci koşu 0 değişiklik (boşaltılan alan zaten NULL)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [
        _satir("S-1", "D-1", **{"Tıbbi Olay": None, "Arşiv Tarihi": "15.03.2021"}),
    ], log=[_log("S-1", "Tıbbi Olay", "(boş)", gerekce="hasta beyanı geri çekildi", eski="Enfeksiyon")])

    once = _kart_oku(dolu_kart, "D-1")
    assert once.tibbi_olay == "Enfeksiyon"
    ilk = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")

    assert ilk.cikis_kodu == CIKIS_TAMAM
    assert ilk.bosaltilan == 1 and ilk.alan_degisikligi == 1 and ilk.kart_degisen == 1
    sonra = _kart_oku(dolu_kart, "D-1")
    assert sonra.tibbi_olay is None and sonra.arsiv_tarihi == date(2021, 3, 15)
    (kayit,) = _tarihce(dolu_kart, "tibbi_olay")
    assert kayit == ("Enfeksiyon", None, f"{ilk.kaynak_imzasi} · boşaltıldı: hasta beyanı geri çekildi")
    assert is_aktarim_source(kayit[2])
    assert "boşaltılan alan   : 1" in hukdok_aktarim.ozet_metni(ilk)

    sayim = _sayim(dolu_kart)
    ikinci = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert ikinci.cikis_kodu == CIKIS_TAMAM
    assert ikinci.alan_degisikligi == 0 and ikinci.bosaltilan == 0 and ikinci.kart_degisen == 0
    assert _sayim(dolu_kart) == sayim
    assert _kart_oku(dolu_kart, "D-1").tibbi_olay is None


def test_sheet_hucresi_doluysa_bosaltma_yok_deger_yazilir(dolu_kart, tmp_path):
    """Üçlü şartın ikinci ayağı: log (boş) dese de Sheet değer taşıyorsa boşaltma YOK;
    değer yazılır ve (boş) satırının gerekçesi o yazıma iliştirilmez (çelişkili talimat)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1", **{"Tıbbi Olay": "Kanama"})],
                       log=[_log("S-1", "Tıbbi Olay", "(boş)", gerekce="çelişkili")])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert sonuc.bosaltilan == 0 and sonuc.alan_degisikligi == 1
    assert _kart_oku(dolu_kart, "D-1").tibbi_olay == "Kanama"
    (kayit,) = _tarihce(dolu_kart, "tibbi_olay")
    assert kayit == ("Enfeksiyon", "Kanama", sonuc.kaynak_imzasi)


def test_log_yoksa_bos_hucre_bosaltmaz(dolu_kart, tmp_path):
    """Üçlü şartın birinci ayağı: Sheet hücresi boş ama Düzeltme_Logu yok → None = "bu
    teslimde yok", mevcut değer KORUNUR (G064 sözleşmesi bozulmadı)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1", **{"Tıbbi Olay": None})])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert sonuc.bosaltilan == 0 and sonuc.alan_degisikligi == 0
    assert _kart_oku(dolu_kart, "D-1").tibbi_olay == "Enfeksiyon"


def test_sutun_paketten_eksikse_bosaltma_yok(dolu_kart, tmp_path):
    """Partili teslim güvencesi: sütun büsbütün yoksa Sheet "boş" TEYİT EDEMEZ →
    log (boş) tek başına boşaltmaz."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       basliklar=["SistemNo", "TKU", "Dosya No"],
                       log=[_log("S-1", "Tıbbi Olay", "(boş)")])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert sonuc.bosaltilan == 0 and sonuc.alan_degisikligi == 0
    assert _kart_oku(dolu_kart, "D-1").tibbi_olay == "Enfeksiyon"


def test_bizde_zaten_bossa_bosaltma_tarihce_yazmaz(dolu_kart, tmp_path):
    """Üçüncü ayak: D-2 tıbbi olayı zaten NULL → talimat sessizce geçer (tarihçe şişmez)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-2", "D-2", **{"Tıbbi Olay": None})],
                       log=[_log("S-2", "Tıbbi Olay", "(boş)")])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert sonuc.bosaltilan == 0 and sonuc.alan_degisikligi == 0
    assert _tarihce(dolu_kart, "tibbi_olay") == []


def test_kunye_ve_icerik_alani_bosaltilmaz_rapora_duser(dolu_kart, tmp_path):
    """Kabul: künye (`karar_no`, tek yazıcı stage_decisions) ve içerik-modu (`court`)
    talimatı UYGULANMAZ; satır raporuna HATA olarak düşer, koşu NONZERO."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [
        _satir("S-1", "D-1", **{"Karar No": None, "Yerel Mahkeme": None, "Tıbbi Olay": None}),
    ], log=[
        _log("S-1", "Karar No", "(boş)", gerekce="künye yanlış"),
        _log("S-1", "Yerel Mahkeme", "(boş)", gerekce="mahkeme yanlış"),
        _log("S-1", "Tıbbi Olay", "(boş)", gerekce="beyan geri çekildi"),   # bu uygulanır
    ])
    sonuc = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")

    assert sonuc.cikis_kodu == CIKIS_SATIR_HATASI and sonuc.islenen == 1
    kart = _kart_oku(dolu_kart, "D-1")
    assert kart.karar_no == "2018/143" and kart.court == "İzmir 4. İdare Mahkemesi"
    assert kart.tibbi_olay is None and sonuc.bosaltilan == 1
    sebepler = sorted(r.sebep for r in sonuc.hatalar)
    assert len(sebepler) == 2
    assert sebepler[0].startswith("court yazılmadı: boşaltılmadı — içerik-karşılaştırmalı")
    assert sebepler[1].startswith("karar_no yazılmadı: boşaltılmadı — karar künyesi tek yazıcı stage_decisions")
    assert _tarihce(dolu_kart, "karar_no") == [] and _tarihce(dolu_kart, "court") == []
    rapor = [y for y in sonuc.raporlar if "satir-raporu" in y.name]
    assert "boşaltılmadı" in rapor[0].read_text(encoding="utf-8-sig")


def test_kardes_foy_bosalt_derken_digeri_deger_tasirsa_alan_celiskili(dolu_kart, tmp_path):
    """Boşaltma da bir değerdir: kardeş föyler uzlaşmıyorsa alan yazılmaz, rapora düşer;
    ikinci koşu salınmaz (idempotency)."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [
        _satir("S-1", "D-1", **{"Tıbbi Olay": None}),
        _satir("S-1b", "D-1", **{"Tıbbi Olay": "Kanama"}),
    ], log=[_log("S-1", "Tıbbi Olay", "(boş)")])
    ilk = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert ilk.bosaltilan == 0 and ilk.alan_degisikligi == 0
    assert _kart_oku(dolu_kart, "D-1").tibbi_olay == "Enfeksiyon"
    celiski = [c for c in ilk.celiskiler if c.alan == "tibbi_olay"]
    assert len(celiski) == 1 and "S-1=(boş)" in celiski[0].degerler and "S-1b=Kanama" in celiski[0].degerler
    sayim = _sayim(dolu_kart)
    ikinci = aktarimi_kos(dolu_kart, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert ikinci.alan_degisikligi == 0 and _sayim(dolu_kart) == sayim


def test_esas_no_bosaltma_tek_yoldan_tarihce_satiri_kalir(db_env, tmp_path):
    """`esas_no` türetilmiş: boşaltma `sync_current_esas` üzerinden — kolon NULL, tarihçe
    satırı SİLİNMEZ (eski esasla arama sürer), güncel işaret düşer."""
    db = db_env()
    try:
        kart = _kart(db, "HA.G112.E", "D-1")
        case_manager.sync_current_esas(db, kart, "2021/588", source="test")
        db.commit()
        kart_id = kart.id
    finally:
        db.close()
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1", **{"Esas": None})],
                       log=[_log("S-1", "Esas", "(boş)", gerekce="görevsizlik")])
    sonuc = aktarimi_kos(db_env, girdi=paket, rapor_dizini=tmp_path / "rapor")
    assert sonuc.bosaltilan == 1
    db = db_env()
    try:
        assert db.get(models.Case, kart_id).esas_no is None
        satirlar = db.query(models.CaseEsasNumber).filter_by(case_id=kart_id).all()
        assert [(s.esas_no, s.is_current) for s in satirlar] == [("2021/588", False)]
    finally:
        db.close()
    assert _tarihce(db_env, "esas_no")[-1][:2] == ("2021/588", None)


def test_kapsam_kilidi_kunye_kolonlarina_bosaltma_yazmaz():
    """Mekanik kilit (G064 ikizi): boşaltma yolu künye kolonlarına setattr etmez;
    `BOSALTMA_DISI_ALANLAR` künye + içerik-modu kümesini kapsar."""
    kaynak = Path(hukdok_aktarim.__file__).read_text(encoding="utf-8")
    for yasak in ("case.karar_no =", "case.karar_tarihi =", "case.istinaf_basvuran_taraf ="):
        assert yasak not in kaynak
    assert hukdok_aktarim.ICERIK_KARSILASTIRMALI_ALANLAR <= hukdok_aktarim.BOSALTMA_DISI_ALANLAR
    assert set(hukdok_aktarim._DUZELTME_KUNYE_KAYNAKLARI.values()) <= hukdok_aktarim.BOSALTMA_DISI_ALANLAR
    assert "(boş)" in (hukdok_aktarim.__doc__ or "") and "BOŞALTILMAZ" in (hukdok_aktarim.__doc__ or "")


# ═══════════════════════════════════════════════════════════════════════════
# 3. DEGER_HAVUZLARI — okuyucu, fark, cevap paketi + bildirim
# ═══════════════════════════════════════════════════════════════════════════

def test_havuz_okuyucu_uzun_bicim_ve_eslemesiz_havuz_atlanir(tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")], havuz=[
        ("İstinaf Karar Durumu", "Kaldırma"),
        ("İstinaf Karar Durumu", "Kaldırma/Yeniden Hüküm ; Başvuru Ret"),
        ("İstinaf Karar Durumu", "Kaldırma"),                     # mükerrer
        ("Para Birimi", "TL"),                                     # eşlemesi yok → atlanır
        ("Olay Türü", "Tıbbi Olay"),
        ("Olay Türü", None),
    ])
    assert tc.havuz_degerlerini_oku(paket) == {
        "İstinaf Karar Durumu": ["Kaldırma", "Kaldırma/Yeniden Hüküm", "Başvuru Ret"],
        "Olay Türü": ["Tıbbi Olay"],
    }


def test_havuz_okuyucu_genis_bicim_ve_sayfa_yoksa_bos(tmp_path):
    genis = _paket_yaz(tmp_path / "g.xlsx", [_satir("S-1", "D-1")],
                       havuz=[("Kaldırma", "Bozma", "TL"), ("Başvuru Ret", None, "USD")],
                       havuz_basliklar=("İstinaf Karar Durumu", "Yargıtay Onama Durumu", "Para Birimi"))
    assert tc.havuz_degerlerini_oku(genis) == {
        "İstinaf Karar Durumu": ["Kaldırma", "Başvuru Ret"],
        "Yargıtay Onama Durumu": ["Bozma"],
    }
    yok = _paket_yaz(tmp_path / "y.xlsx", [_satir("S-1", "D-1")])
    assert tc.havuz_degerlerini_oku(yok) == {}


def test_havuz_eslemesi_sozlesme_ve_registry():
    """Sabit küçük sözlük: altı havuz, hepsi LIST_REGISTRY'de (yalnız okunur)."""
    from managers.reference_lists import LIST_REGISTRY

    assert set(tc.HAVUZ_LISTE_ESLEMESI.values()) == {
        "alleged_faults", "appeal_decisions", "cassation_decisions",
        "local_decisions", "event_types", "judgment_roles",
    }
    assert set(tc.HAVUZ_LISTE_ESLEMESI.values()) <= set(LIST_REGISTRY)
    assert tc.HAVUZ_LISTE_ESLEMESI[tk._anahtar("İDDİA EDİLEN KUSUR")] == "alleged_faults"


def _liste_doldur(fabrika):
    db = fabrika()
    try:
        db.add_all([
            models.AppealDecision(code="KALDIRMA", name="Kaldırma", sequence=0),
            models.AppealDecision(code="BASVURU-RET", name="Başvuru Ret", sequence=1),
            models.EventType(code="TIBBI", name="Tıbbi Olay", sequence=0),
        ])
        db.commit()
    finally:
        db.close()


def _liste_sayimlari(fabrika):
    from managers.reference_lists import LIST_REGISTRY

    db = fabrika()
    try:
        return {ad: db.query(LIST_REGISTRY[ad].model).count() for ad in tc.HAVUZ_LISTE_ESLEMESI.values()}
    finally:
        db.close()


def test_havuz_farki_iki_yon_yazim_toleransli_listeye_yazmaz(db_env):
    _liste_doldur(db_env)
    once = _liste_sayimlari(db_env)
    db = db_env()
    try:
        farklar = tc.havuz_farki(db, {
            "İstinaf Karar Durumu": ["KALDIRMA", "Kaldırma/Yeniden Hüküm"],   # yazım farkı = aynı
            "Olay Türü": ["Tıbbi Olay"],                                       # tam örtüşme
            "İddia Edilen Kusur": ["Teşhis Hatası"],                           # bizde liste boş
        })
    finally:
        db.close()
    assert farklar == [
        ("İstinaf Karar Durumu", "appeal_decisions", tc.YON_TESLIMDE_VAR, "Kaldırma/Yeniden Hüküm"),
        ("İstinaf Karar Durumu", "appeal_decisions", tc.YON_BIZDE_VAR, "Başvuru Ret"),
        ("İddia Edilen Kusur", "alleged_faults", tc.YON_TESLIMDE_VAR, "Teşhis Hatası"),
    ]
    assert _liste_sayimlari(db_env) == once                       # listeye YAZMA yok


@pytest.fixture()
def teslim_env(tmp_path, monkeypatch):
    """G110 `env` ikizi: sqlite + defter/föy/bildirim index'leri + spool + iki kart + admin."""
    monkeypatch.setenv("TESLIM_SPOOL_DIR", str(tmp_path / "teslim_spool"))
    monkeypatch.setenv("ADMIN_EMAILS", ADMIN)
    monkeypatch.setenv("SHAREPOINT_FOLDER_TESLIM_NAME", "03_VERI_TESLIM")
    for ad in ("TESLIM_KAPI_HATA_ORANI", "TESLIM_KAPI_ESLESMEYEN_ORANI", "TESLIM_KAPI_ALAN_DEGISIKLIGI"):
        monkeypatch.delenv(ad, raising=False)
    engine = _engine()
    with engine.begin() as conn:
        for tablo in ("case_foys", "aktarim_teslimleri", "notifications"):
            for sql in _index_ops(tablo):
                conn.execute(text(sql))
    maker = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    monkeypatch.setattr(tk, "SessionLocal", maker)
    monkeypatch.setattr(app_settings, "SessionLocal", maker)
    monkeypatch.setattr(spu, "list_folder_children", lambda folder_name: [])
    monkeypatch.setattr(spu, "upload_file_to_sharepoint",
                        lambda *a, **kw: (_ for _ in ()).throw(AssertionError("upload beklenmiyor")))
    db = maker()
    try:
        for i in (1, 2):
            _kart(db, f"HA.G112.T{i}", f"D-{i}")
        db.commit()
    finally:
        db.close()
    yield SimpleNamespace(db=maker, spool=tmp_path / "teslim_spool")
    engine.dispose()


def _bildirimler(env):
    db = env.db()
    try:
        return [(n.dedupe_key, n.title, n.body, n.severity) for n in
                db.query(models.Notification).filter_by(type=tk.BILDIRIM_TURU).order_by(models.Notification.id)]
    finally:
        db.close()


def _rapor(env, tid):
    db = env.db()
    try:
        return Path(db.get(models.AktarimTeslimi, tid).rapor_dizini)
    finally:
        db.close()


def test_teslim_havuz_farki_csv_cevap_paketinde_bildirim_dedupe_liste_sabit(teslim_env, caplog):
    """Kabul: fark CSV'si iki yönü listeler (rapor dizininde → cevap paketine girer), admin
    bildirimi `teslim:<id>:havuz:<alıcı>`; kuru koşu + uygulama ikinci bildirim ÜRETMEZ;
    referans tablolarında satır sayısı koşu öncesi/sonrası AYNI."""
    _liste_doldur(teslim_env.db)
    once = _liste_sayimlari(teslim_env.db)
    icerik = _paket_bayt(
        [_satir("S-1", "D-1", **{"Tıbbi Olay": "Enfeksiyon"}), _satir("S-2", "D-2")],
        havuz=[("İstinaf Karar Durumu", "Kaldırma"), ("İstinaf Karar Durumu", "Kaldırma/Yeniden Hüküm"),
               ("Olay Türü", "Tıbbi Olay"), ("Para Birimi", "TL")],
    )
    db = teslim_env.db()
    try:
        tid = tk.teslim_kaydet(icerik=icerik, dosya_adi="HUKDOK_TESLIM_H.xlsx", kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid, db=db) == tk.DURUM_DOGRULANDI
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_kuru_kos(tid, db=db) == tk.DURUM_KURU_KOSULDU
        rapor = _rapor(teslim_env, tid)
        dosya = rapor / "deger-havuzu-farki_HUKDOK_TESLIM_H.csv"
        assert dosya.is_file()
        ham = dosya.read_bytes()
        assert ham.startswith(b"\xef\xbb\xbf" + ";".join(tc.HAVUZ_FARKI_BASLIKLARI).encode())
        satirlar = ham.decode("utf-8-sig").splitlines()[1:]
        assert satirlar == [
            f"İstinaf Karar Durumu;appeal_decisions;{tc.YON_TESLIMDE_VAR};Kaldırma/Yeniden Hüküm",
            f"İstinaf Karar Durumu;appeal_decisions;{tc.YON_BIZDE_VAR};Başvuru Ret",
        ]
        # cevap paketi dosyayı kendi adıyla alır
        assert dosya.name in {ad for _yol, ad in tc._cevap_dosyalari(rapor, "HUKDOK_TESLIM_H")}

        bildirimler = _bildirimler(teslim_env)
        assert len(bildirimler) == 1
        anahtar, baslik, govde, severity = bildirimler[0]
        assert anahtar == f"teslim:{tid}:havuz:{ADMIN}" == tk.bildirim_dedupe_key(tid, tk.BILDIRIM_HAVUZ, ADMIN)
        assert baslik == "Veri teslimi değer havuzu farkı: HUKDOK_TESLIM_H.xlsx" and severity == "warning"
        assert "2 fark" in govde and "İstinaf Karar Durumu" in govde
        assert not [r for r in caplog.records if r.levelno >= logging.WARNING
                    and r.name.startswith("services.teslim_")]

        # kapı (ilk teslim → inceleme) + uygulama: dosya tazelenir, bildirim İKİLENMEZ
        assert tk.kapi_degerlendir(tid, db=db) == tk.KAPI_INCELEME
        assert tk.teslim_uygula(tid, uygulayan="admin@buro.test", db=db) == tk.DURUM_UYGULANDI
    finally:
        db.close()          # cevap denemesi (anahtar kapalı) okuma transaction'ı açık bırakır
    assert dosya.is_file()
    assert [b[0] for b in _bildirimler(teslim_env)].count(f"teslim:{tid}:havuz:{ADMIN}") == 1
    assert _liste_sayimlari(teslim_env.db) == once
    assert _kart_oku(teslim_env.db, "D-1").tibbi_olay == "Enfeksiyon"    # aktarımın kendisi bozulmadı


def test_teslim_havuz_farki_yokken_dosya_ve_bildirim_yok(teslim_env):
    """Kabul: fark yokken (tam örtüşme ya da sayfa yok) CSV üretilmez, bildirim yok;
    bayat fark dosyası varsa temizlenir."""
    _liste_doldur(teslim_env.db)
    icerik = _paket_bayt([_satir("S-1", "D-1")],
                         havuz=[("İstinaf Karar Durumu", "Kaldırma ; Başvuru Ret"), ("Olay Türü", "Tıbbi Olay")])
    db = teslim_env.db()
    try:
        tid = tk.teslim_kaydet(icerik=icerik, dosya_adi="HUKDOK_TESLIM_Y.xlsx", kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid, db=db) == tk.DURUM_DOGRULANDI
        # bayat kopya: önceki bir koşudan kalmış gibi
        rapor = tk._rapor_dizini(db.get(models.AktarimTeslimi, tid))
        bayat = rapor / "deger-havuzu-farki_HUKDOK_TESLIM_Y.csv"
        bayat.write_text("eski", encoding="utf-8")
        assert tk.teslim_kuru_kos(tid, db=db) == tk.DURUM_KURU_KOSULDU
        assert not bayat.exists()
        assert not list(rapor.glob("deger-havuzu-farki_*.csv"))
        assert _bildirimler(teslim_env) == []

        # sayfa hiç yok
        tid2 = tk.teslim_kaydet(icerik=_paket_bayt([_satir("S-2", "D-2")]), dosya_adi="HUKDOK_TESLIM_Z.xlsx",
                                kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid2, db=db) == tk.DURUM_DOGRULANDI
        assert tk.teslim_kuru_kos(tid2, db=db) == tk.DURUM_KURU_KOSULDU
        assert not list(_rapor(teslim_env, tid2).glob("deger-havuzu-farki_*.csv"))
        assert _bildirimler(teslim_env) == []
    finally:
        db.close()


def test_havuz_farki_istisnasi_kuru_kosuyu_dusurmez(teslim_env, monkeypatch, caplog):
    """Yan ürün: fark hesabı patlarsa WARNING, teslim yine `kuru_kosuldu`."""
    def _patla(*a, **kw):
        raise RuntimeError("havuz okunamadı")

    monkeypatch.setattr(tc, "teslim_havuz_farki", _patla)
    db = teslim_env.db()
    try:
        tid = tk.teslim_kaydet(icerik=_paket_bayt([_satir("S-1", "D-1")]), dosya_adi="HUKDOK_TESLIM_P.xlsx",
                               kaynak="yukleme", db=db)
        assert tk.teslim_dogrula(tid, db=db) == tk.DURUM_DOGRULANDI
        with caplog.at_level(logging.WARNING):
            assert tk.teslim_kuru_kos(tid, db=db) == tk.DURUM_KURU_KOSULDU
        assert db.get(models.AktarimTeslimi, tid).durum == tk.DURUM_KURU_KOSULDU
    finally:
        db.close()
    uyarilar = [r for r in caplog.records if r.levelno == logging.WARNING and "havuzu farkı" in r.getMessage()]
    assert len(uyarilar) == 1 and not [r for r in caplog.records if r.levelno >= logging.ERROR]
    assert _bildirimler(teslim_env) == []

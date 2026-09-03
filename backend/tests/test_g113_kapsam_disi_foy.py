"""G113 — `Silinen_Föyler` / `Kapsam_Dışı` sayfaları → `case_foys` kapsam işareti.

Veri ekibi kapsamdan çıkardığı föyleri (mükerrer/hatalı açılış → `SILINDI`;
malpraktis dışı → `KAPSAM_DISI`) gerekçe + tarihle ayrı sayfalarda gönderiyor. Föy
SİLİNMEZ (kart ve belgeler dokunulmaz — belge koruma şartı), yalnız işaretlenir.

Sözleşme (gorevler/gorev/G113.md):

* `case_foys.kapsam_durumu` VARCHAR(20) NULL (`SILINDI` | `KAPSAM_DISI`; NULL =
  kapsamda), `kapsam_gerekcesi` VARCHAR NULL, `kapsam_tarihi` DATE NULL — migration
  `("columns", "case_foys", {...})`, index yok.
* Aktarım: SistemNo bizde varsa üç alan yazılır, tarihçe kart düzeyinde ("föy
  kapsam dışı: <gerekçe>"); ikinci koşu 0 değişiklik; bizde olmayan → ATLANDI;
  sayfa yoksa hata değil.
* Geri dönüş: föy ana sayfada (`Sheet`) yeniden görünürse işaret NULL + tarihçe.
* D9: kapsam dışı föy kardeş-föy uzlaşısına (kart alanı + künye) katılmaz.
* `case_relations_auto`: kapsam dışı föy TKU ilişkisi üretmez.
* `get_case.foyler[]`: her föyde üç alan; kapsamda föyde NULL.

Katmanlar: şema (model + migrasyon op'u) · sqlite (aktarım, ilişki, kart) · dbtest
(gerçek Postgres'te üç kolon; 3-ortam kuralı: şema göçmemişse SKIP).

**TEST VERİSİ KURALI (A.2 dersi):** gerçek teslim paketi REPOYA GİRMEZ; paketler
openpyxl ile SENTETİK üretilir (test_g064/test_g112 düzeni + kapsam sayfaları).
"""
import logging
import os
from datetime import date
from pathlib import Path

import pytest
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import NullPool, StaticPool

import models
from database import _MIGRATIONS, Base
from managers import case_manager, foy_map
from scripts import hukdok_aktarim
from scripts.hukdok_aktarim import (
    CIKIS_TAMAM,
    KAPSAM_DISI,
    KAPSAM_SILINDI,
    KAPSAM_TARIHCE_ALANI,
    AktarimHatasi,
    aktarimi_kos,
    kapsam_kayitlarini_oku,
)
from services import case_relations_auto as ilis

BASLIKLAR = ["SistemNo", "TKU", "Hasar No", "Dosya No", "Tıbbi Olay", "Karar No"]
SILINEN_BASLIKLAR = ["SistemNo", "Dosya No", "Silinme Gerekçesi", "Tarih"]
KAPSAM_DISI_BASLIKLAR = ["SistemNo", "Dosya No", "Kapsam Dışı Gerekçesi", "Tarih"]
TENANT = "tenant-hanyaloglu"


# ═══════════════════════════════════════════════════════════════════════════
# Sentetik paket üreticileri
# ═══════════════════════════════════════════════════════════════════════════

def _kitap(satirlar, *, basliklar=None, sayfa="Sheet", silinen=None, silinen_basliklar=None,
           kapsam_disi=None, kapsam_disi_basliklar=None, asamalar=None):
    """Ana sayfa + isteğe bağlı `Silinen_Föyler` / `Kapsam_Dışı` / `Karar_Asamalari`."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sayfa
    kullanilan = list(basliklar if basliklar is not None else BASLIKLAR)
    ws.append(kullanilan)
    for satir in satirlar:
        ws.append([satir.get(b) for b in kullanilan])
    for ad, veri, varsayilan, verilen in (
        ("Silinen_Föyler", silinen, SILINEN_BASLIKLAR, silinen_basliklar),
        ("Kapsam_Dışı", kapsam_disi, KAPSAM_DISI_BASLIKLAR, kapsam_disi_basliklar),
    ):
        if veri is None:
            continue
        sb = list(verilen if verilen is not None else varsayilan)
        w = wb.create_sheet(ad)
        w.append(sb)
        for satir in veri:
            w.append([satir.get(b) for b in sb])
    if asamalar is not None:
        ab = ["SistemNo", "AsamaNo", "Aşama", "Mahkeme", "Esas No", "Karar No",
              "Karar Tarihi", "Karar Durumu", "Güven"]
        wa = wb.create_sheet(hukdok_aktarim.ASAMA_SAYFASI)
        wa.append(ab)
        for satir in asamalar:
            wa.append([satir.get(b) for b in ab])
    return wb


def _paket_yaz(yol, satirlar, **kw):
    wb = _kitap(satirlar, **kw)
    wb.save(yol)
    wb.close()
    return Path(yol)


def _satir(sistem_no, dosya_no, **extra):
    temel = {"SistemNo": sistem_no, "Dosya No": dosya_no, "TKU": "TKU-113"}
    temel.update(extra)
    return temel


def _silinen(sistem_no, gerekce="mükerrer açılış", tarih="12.08.2026", dosya_no="D-1"):
    return {"SistemNo": sistem_no, "Dosya No": dosya_no, "Silinme Gerekçesi": gerekce, "Tarih": tarih}


def _kapsam_disi(sistem_no, gerekce="malpraktis dışı", tarih=date(2026, 8, 20), dosya_no="D-1"):
    return {"SistemNo": sistem_no, "Dosya No": dosya_no, "Kapsam Dışı Gerekçesi": gerekce, "Tarih": tarih}


# ═══════════════════════════════════════════════════════════════════════════
# 1. Şema — model + migrasyon op'u
# ═══════════════════════════════════════════════════════════════════════════

def _case_foys_kolon_oplari():
    ops = {}
    for op in _MIGRATIONS:
        if op[0] == "columns" and op[1] == "case_foys":
            for name, spec in op[2].items():
                ops[name] = spec if isinstance(spec, str) else spec[0]
    return ops


def test_uc_kolon_columns_opunda_ve_modelle_uyumlu():
    """Migration `("columns", "case_foys", {...})`: üç kolon, DDL modelle aynı
    (VARCHAR(20) / VARCHAR / DATE), hepsi NULL kabul eder. `("columns", ...)` op'u
    KOŞULLUDUR (mevcut kolonu atlar) — ikinci koşu bu yüzden yapısal olarak
    idempotent; gerçek Postgres kanıtı dbtest bölümünde."""
    ops = _case_foys_kolon_oplari()
    assert set(ops) >= {"kapsam_durumu", "kapsam_gerekcesi", "kapsam_tarihi"}
    assert ops["kapsam_durumu"].upper().startswith("VARCHAR(20)")
    assert ops["kapsam_gerekcesi"].upper().startswith("VARCHAR")
    assert ops["kapsam_tarihi"].upper().startswith("DATE")
    for ddl in ops.values():
        assert "NOT NULL" not in ddl.upper() and "DEFAULT" not in ddl.upper()
    kolonlar = models.CaseFoy.__table__.columns
    assert kolonlar["kapsam_durumu"].type.length == 20
    assert all(kolonlar[k].nullable for k in ("kapsam_durumu", "kapsam_gerekcesi", "kapsam_tarihi"))
    # index YOK (G041/G042): kapsam kolonları hiçbir ("index", ...) op'unda geçmez
    for op in _MIGRATIONS:
        if op[0] == "index" and op[1] == "case_foys":
            assert not any("kapsam" in sql for sql in op[2])


def test_kapsam_sayfalari_ve_durum_degerleri_sozlesmeye_uygun():
    assert hukdok_aktarim.KAPSAM_SAYFALARI == {"Silinen_Föyler": "SILINDI", "Kapsam_Dışı": "KAPSAM_DISI"}
    assert KAPSAM_SILINDI == "SILINDI" and KAPSAM_DISI == "KAPSAM_DISI"
    assert all(len(d) <= 20 for d in hukdok_aktarim.KAPSAM_SAYFALARI.values())


# ═══════════════════════════════════════════════════════════════════════════
# 2. Okuyucu (DB yok)
# ═══════════════════════════════════════════════════════════════════════════

def test_kapsam_sayfasi_yoksa_bos_hata_degil(tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")])
    assert kapsam_kayitlarini_oku(paket) == {}


def test_iki_sayfa_okunur_gerekce_ve_tarih_cozulur(tmp_path, caplog):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       silinen=[_silinen("S-2", "mükerrer", "12.08.2026"),
                                _silinen("S-3", "hatalı açılış", "bozuk-tarih"),
                                _silinen("S-2", "ikinci kez", "01.01.2026")],       # mükerrer → ilk kazanır
                       kapsam_disi=[_kapsam_disi("S-4", "malpraktis dışı", date(2026, 8, 20))])
    with caplog.at_level(logging.WARNING, logger="HukdokAktarim"):
        kayitlar = kapsam_kayitlarini_oku(paket)
    assert set(kayitlar) == {"S-2", "S-3", "S-4"}
    assert (kayitlar["S-2"].durum, kayitlar["S-2"].gerekce, kayitlar["S-2"].tarih) == (
        "SILINDI", "mükerrer", date(2026, 8, 12))
    assert kayitlar["S-2"].satir_no == 2 and kayitlar["S-2"].sayfa == "Silinen_Föyler"
    assert (kayitlar["S-3"].durum, kayitlar["S-3"].tarih) == ("SILINDI", None)   # tarih çözülmedi → None
    assert (kayitlar["S-4"].durum, kayitlar["S-4"].gerekce, kayitlar["S-4"].tarih) == (
        "KAPSAM_DISI", "malpraktis dışı", date(2026, 8, 20))
    uyarilar = [r.getMessage() for r in caplog.records if r.levelno == logging.WARNING]
    assert any("bozuk-tarih" in u for u in uyarilar)
    assert any("ilk kayıt kazandı" in u for u in uyarilar)


def test_sayfa_adi_ve_gerekce_basligi_toleransli(tmp_path):
    """Sayfa adı aksan/alt çizgi duyarsız; "… Gerekçesi" soneki adaylarda yoksa da tanınır."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(BASLIKLAR)
    ws.append(["S-1", "TKU-113", None, "D-1", None, None])
    w = wb.create_sheet("SILINEN FOYLER")
    w.append(["Sistem No", "Föy Silinme Gerekçesi", "Tarih"])
    w.append(["S-2", "çift kayıt", "05.05.2026"])
    yol = tmp_path / "t.xlsx"
    wb.save(yol)
    wb.close()
    kayitlar = kapsam_kayitlarini_oku(yol)
    assert kayitlar["S-2"].gerekce == "çift kayıt" and kayitlar["S-2"].durum == "SILINDI"


def test_sistem_no_sutunu_yoksa_sayfa_uygulanamaz(tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       silinen=[{"Silinme Gerekçesi": "x"}],
                       silinen_basliklar=["Dosya No", "Silinme Gerekçesi", "Tarih"])
    with pytest.raises(AktarimHatasi, match="SistemNo"):
        kapsam_kayitlarini_oku(paket)


def test_gerekce_sutunu_yoksa_tek_warning_isaret_gerekcesiz(tmp_path, caplog):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       silinen=[{"SistemNo": "S-2", "Tarih": "12.08.2026"}],
                       silinen_basliklar=["SistemNo", "Tarih"])
    with caplog.at_level(logging.WARNING, logger="HukdokAktarim"):
        kayitlar = kapsam_kayitlarini_oku(paket)
    assert kayitlar["S-2"].gerekce is None and kayitlar["S-2"].tarih == date(2026, 8, 12)
    assert len([r for r in caplog.records if r.levelno == logging.WARNING]) == 1


# ═══════════════════════════════════════════════════════════════════════════
# 3. sqlite — aktarım
# ═══════════════════════════════════════════════════════════════════════════

def _index_ops(table):
    return [sql for op in _MIGRATIONS if op[0] == "index" and op[1] == table for sql in op[2]]


def _engine():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _fk_ac(dbapi_connection, _record):
        dbapi_connection.isolation_level = None      # pysqlite BEGIN yaymasın (G064 reçetesi)
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    @event.listens_for(engine, "begin")
    def _begin(conn):
        conn.exec_driver_sql("BEGIN")

    Base.metadata.create_all(engine)
    return engine


@pytest.fixture()
def db_env():
    engine = _engine()
    with engine.begin() as conn:
        for sql in _index_ops("case_foys"):
            conn.execute(text(sql))
    yield sessionmaker(bind=engine, autocommit=False, autoflush=False)
    engine.dispose()


def _kart(db, tracking, klasor, **extra):
    case = models.Case(tracking_no=tracking, status="DERDEST", klasor_no_2=klasor, **extra)
    db.add(case)
    db.flush()
    return case


@pytest.fixture()
def belgeli_kart(db_env):
    """D-1 kartı: taraf + o tarafa bağlı işlenmiş belge (belge koruma zemini); D-2 boş."""
    db = db_env()
    try:
        k1 = _kart(db, "HA.G113.1", "D-1")
        taraf = models.CaseParty(case_id=k1.id, name="Ayşe Y.", role="Davacı", party_type="CLIENT")
        db.add(taraf)
        db.flush()
        db.add(models.CaseDocument(
            case_id=k1.id, original_filename="karar.pdf", stored_filename="k.pdf",
            belge_turu_kodu="KARAR_________", case_party_id=taraf.id,
        ))
        _kart(db, "HA.G113.2", "D-2")
        db.commit()
    finally:
        db.close()
    return db_env


def _foy_oku(fabrika, sistem_no):
    db = fabrika()
    try:
        f = foy_map.get_foy(db, sistem_no)
        return None if f is None else (f.case_id, f.kapsam_durumu, f.kapsam_gerekcesi, f.kapsam_tarihi)
    finally:
        db.close()


def _sayim(fabrika):
    db = fabrika()
    try:
        return {m.__name__: db.query(m).count() for m in
                (models.Case, models.CaseFoy, models.CaseHistory, models.CaseParty, models.CaseDocument)}
    finally:
        db.close()


def _kapsam_tarihcesi(fabrika):
    db = fabrika()
    try:
        return [(h.case_id, h.old_value, h.new_value, h.source) for h in
                db.query(models.CaseHistory).filter_by(field_name=KAPSAM_TARIHCE_ALANI)
                .order_by(models.CaseHistory.id)]
    finally:
        db.close()


def test_silinen_foy_isaretlenir_kart_ve_belge_dokunulmaz_ikinci_kosu_sifir(belgeli_kart, tmp_path):
    """Kabul: Silinen_Föyler'deki SistemNo işaretlenir (gerekçe + tarih); kart ve belge
    sayıları DEĞİŞMEZ (envanter denk); tarihçe kart düzeyinde; ikinci koşu 0 değişiklik."""
    # Önceki teslim: S-1 ve S-2 D-1 kartının föyleri
    ilk_paket = _paket_yaz(tmp_path / "t0.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")])
    aktarimi_kos(belgeli_kart, girdi=ilk_paket, rapor_dizini=tmp_path / "r")
    once = _sayim(belgeli_kart)
    assert once["CaseFoy"] == 2 and once["CaseDocument"] == 1

    # Bu teslim: S-1 ana sayfada, S-2 Silinen_Föyler'de (ana sayfada YOK)
    paket = _paket_yaz(tmp_path / "t1.xlsx", [_satir("S-1", "D-1")],
                       silinen=[_silinen("S-2", "mükerrer açılış", "12.08.2026")])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")

    assert sonuc.cikis_kodu == CIKIS_TAMAM and sonuc.yazildi and sonuc.envanter_farki == {}
    assert (sonuc.kapsam_isaretlenen, sonuc.kapsam_geri_alinan, sonuc.kapsam_atlanan) == (1, 0, 0)
    case_id = _foy_oku(belgeli_kart, "S-1")[0]
    assert _foy_oku(belgeli_kart, "S-2") == (case_id, "SILINDI", "mükerrer açılış", date(2026, 8, 12))
    assert _foy_oku(belgeli_kart, "S-1") == (case_id, None, None, None)   # kapsamda föy NULL
    sonra = _sayim(belgeli_kart)
    assert (sonra["Case"], sonra["CaseFoy"], sonra["CaseParty"], sonra["CaseDocument"]) == (
        once["Case"], once["CaseFoy"], once["CaseParty"], once["CaseDocument"])
    assert sonra["CaseHistory"] == once["CaseHistory"] + 1
    [(h_case, eski, yeni, kaynak)] = _kapsam_tarihcesi(belgeli_kart)
    assert h_case == case_id and eski == "S-2: kapsamda"
    assert "föy kapsam dışı: mükerrer açılış" in yeni and yeni.startswith("S-2: SILINDI")
    assert kaynak == sonuc.kaynak_imzasi
    # Belge-taraf bağı kımıldamadı
    db = belgeli_kart()
    try:
        belge = db.query(models.CaseDocument).one()
        assert belge.case_id == case_id and belge.case_party_id is not None
    finally:
        db.close()

    ikinci = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")
    assert ikinci.cikis_kodu == CIKIS_TAMAM
    assert (ikinci.kapsam_isaretlenen, ikinci.kapsam_geri_alinan, ikinci.kapsam_atlanan) == (0, 0, 0)
    assert _sayim(belgeli_kart) == sonra                            # tarihçe büyümedi
    assert "kapsam işareti    : 0 işaretlendi" in hukdok_aktarim.ozet_metni(ikinci)


def test_kapsam_disi_sayfasi_kapsam_disi_durumu_yazar(belgeli_kart, tmp_path):
    ilk = _paket_yaz(tmp_path / "t0.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-2")])
    aktarimi_kos(belgeli_kart, girdi=ilk, rapor_dizini=tmp_path / "r")
    paket = _paket_yaz(tmp_path / "t1.xlsx", [_satir("S-1", "D-1")],
                       kapsam_disi=[_kapsam_disi("S-2", "malpraktis dışı", date(2026, 8, 20))])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")
    assert sonuc.cikis_kodu == CIKIS_TAMAM and sonuc.kapsam_isaretlenen == 1
    assert _foy_oku(belgeli_kart, "S-2")[1:] == ("KAPSAM_DISI", "malpraktis dışı", date(2026, 8, 20))


def test_bizde_olmayan_sistem_no_atlanir_kosu_kirmizi_degil(belgeli_kart, tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1")],
                       silinen=[_silinen("S-YOK", "mükerrer", "12.08.2026")])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")
    assert sonuc.cikis_kodu == CIKIS_TAMAM and sonuc.kapsam_atlanan == 1 and sonuc.kapsam_isaretlenen == 0
    [rapor] = [r for r in sonuc.rapor_satirlari if r.sistem_no == "S-YOK"]
    assert rapor.tur == "ATLANDI" and "Silinen_Föyler" in rapor.sebep and "föy bizde yok" in rapor.sebep
    assert _foy_oku(belgeli_kart, "S-YOK") is None                  # föy YARATILMADI


def test_geri_donus_ana_sayfaya_donen_foy_kapsama_alinir(belgeli_kart, tmp_path):
    """Kabul: föy sonraki pakette `Sheet`'e dönünce işaret NULL + tarihçe; sonraki koşu 0."""
    _ = aktarimi_kos(belgeli_kart, girdi=_paket_yaz(
        tmp_path / "t0.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")]), rapor_dizini=tmp_path / "r")
    _ = aktarimi_kos(belgeli_kart, girdi=_paket_yaz(
        tmp_path / "t1.xlsx", [_satir("S-1", "D-1")], silinen=[_silinen("S-2")]), rapor_dizini=tmp_path / "r")
    assert _foy_oku(belgeli_kart, "S-2")[1] == "SILINDI"

    donus = _paket_yaz(tmp_path / "t2.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")])
    sonuc = aktarimi_kos(belgeli_kart, girdi=donus, rapor_dizini=tmp_path / "r")
    assert sonuc.cikis_kodu == CIKIS_TAMAM
    assert (sonuc.kapsam_isaretlenen, sonuc.kapsam_geri_alinan) == (0, 1)
    assert _foy_oku(belgeli_kart, "S-2")[1:] == (None, None, None)
    tarihce = _kapsam_tarihcesi(belgeli_kart)
    assert len(tarihce) == 2
    assert tarihce[-1][1] == "S-2: SILINDI" and "kapsama geri alındı" in tarihce[-1][2]

    tekrar = aktarimi_kos(belgeli_kart, girdi=donus, rapor_dizini=tmp_path / "r")
    assert (tekrar.kapsam_isaretlenen, tekrar.kapsam_geri_alinan) == (0, 0)
    assert len(_kapsam_tarihcesi(belgeli_kart)) == 2


def test_dusen_satir_adina_geri_donus_verilmez(belgeli_kart, tmp_path):
    """Savepoint'i geri alınan satır (kart yok) ana sayfada olsa da işaret kalır."""
    aktarimi_kos(belgeli_kart, girdi=_paket_yaz(
        tmp_path / "t0.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")]), rapor_dizini=tmp_path / "r")
    aktarimi_kos(belgeli_kart, girdi=_paket_yaz(
        tmp_path / "t1.xlsx", [_satir("S-1", "D-1")], silinen=[_silinen("S-2")]), rapor_dizini=tmp_path / "r")
    # S-2 ana sayfada ama satırı BOZUK (tarih çözümlenemiyor → SatirHatasi → savepoint geri)
    satirlar = [_satir("S-1", "D-1"), _satir("S-2", "D-1", **{"Arşiv Tarihi": "31.02.2026-bozuk"})]
    bozuk = _paket_yaz(tmp_path / "t2.xlsx", satirlar, basliklar=BASLIKLAR + ["Arşiv Tarihi"])
    sonuc = aktarimi_kos(belgeli_kart, girdi=bozuk, rapor_dizini=tmp_path / "r")
    assert [r.sistem_no for r in sonuc.hatalar] == ["S-2"]
    assert sonuc.kapsam_geri_alinan == 0
    assert _foy_oku(belgeli_kart, "S-2")[1] == "SILINDI"


def test_kapsam_disi_foy_celiski_raporuna_girmez_ve_kart_alanini_yazmaz(belgeli_kart, tmp_path):
    """Kabul (D9): iki föylü kart, biri kapsam dışı, künye ve kart alanı farklı → çelişki 0;
    kapsamdaki föyün değeri karta yazılır, kapsam dışı föy yalnız kimlik bırakır."""
    paket = _paket_yaz(tmp_path / "t.xlsx", [
        _satir("S-1", "D-1", **{"Tıbbi Olay": "Enfeksiyon", "Karar No": "2018/143"}),
        _satir("S-2", "D-1", **{"Tıbbi Olay": "Kanama", "Karar No": "2016/768"}),
    ], silinen=[_silinen("S-2", "mükerrer açılış")])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")
    assert sonuc.cikis_kodu == CIKIS_TAMAM
    assert sonuc.celiskiler == []                                   # ne kart alanı ne künye çelişkisi
    assert sonuc.alan_degisikligi == 1 and sonuc.kapsam_isaretlenen == 1
    db = belgeli_kart()
    try:
        kart = db.query(models.Case).filter_by(klasor_no_2="D-1").one()
        assert kart.tibbi_olay == "Enfeksiyon"                      # kapsamdaki föy kazandı, kur'a yok
        foyler = {f.sistem_no: f for f in foy_map.get_case_foys(db, kart.id)}
        assert set(foyler) == {"S-1", "S-2"}                       # kimlik yazıldı, föy SİLİNMEDİ
        assert foyler["S-2"].kapsam_durumu == "SILINDI" and foyler["S-1"].kapsam_durumu is None
    finally:
        db.close()

    # Kontrol: kapsam sayfası OLMASAYDI aynı paket çelişki üretir ve alanı yazmazdı
    kontrol = _paket_yaz(tmp_path / "k.xlsx", [
        _satir("K-1", "D-2", **{"Tıbbi Olay": "Enfeksiyon", "Karar No": "2018/143"}),
        _satir("K-2", "D-2", **{"Tıbbi Olay": "Kanama", "Karar No": "2016/768"}),
    ])
    kontrol_sonuc = aktarimi_kos(belgeli_kart, girdi=kontrol, rapor_dizini=tmp_path / "r")
    assert {c.alan for c in kontrol_sonuc.celiskiler} == {"tibbi_olay", "karar_no"}


def test_kapsam_disi_foyun_asama_satiri_yazilmaz(belgeli_kart, tmp_path):
    paket = _paket_yaz(tmp_path / "t.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")],
                       silinen=[_silinen("S-2")],
                       asamalar=[{"SistemNo": "S-2", "AsamaNo": 1, "Aşama": "Yerel",
                                  "Mahkeme": "Ankara 3. Tüketici Mahkemesi", "Esas No": "2021/55",
                                  "Karar No": "2022/9", "Karar Tarihi": "10.03.2022",
                                  "Karar Durumu": "Kabul", "Güven": "KESİN"}])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r")
    assert sonuc.cikis_kodu == CIKIS_TAMAM and sonuc.asama_eklenen == 0
    db = belgeli_kart()
    try:
        assert db.query(models.CaseStageDecision).count() == 0
    finally:
        db.close()


def test_kuru_kosu_isaret_yazmaz(belgeli_kart, tmp_path):
    aktarimi_kos(belgeli_kart, girdi=_paket_yaz(
        tmp_path / "t0.xlsx", [_satir("S-1", "D-1"), _satir("S-2", "D-1")]), rapor_dizini=tmp_path / "r")
    paket = _paket_yaz(tmp_path / "t1.xlsx", [_satir("S-1", "D-1")], silinen=[_silinen("S-2")])
    sonuc = aktarimi_kos(belgeli_kart, girdi=paket, rapor_dizini=tmp_path / "r", dry_run=True)
    assert sonuc.kapsam_isaretlenen == 1 and not sonuc.yazildi
    assert _foy_oku(belgeli_kart, "S-2")[1] is None


# ═══════════════════════════════════════════════════════════════════════════
# 4. sqlite — ilişki katmanı + kart çıktısı
# ═══════════════════════════════════════════════════════════════════════════

@pytest.fixture
def oturum():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    models.Base.metadata.create_all(engine)
    Fabrika = sessionmaker(bind=engine)
    db = Fabrika()
    yield db
    db.close()
    engine.dispose()


def _iliski_karti(db, tracking_no, **alanlar):
    case = models.Case(tracking_no=tracking_no, status="DERDEST", **alanlar)
    db.add(case)
    db.commit()
    return case


def _foy(db, sistem_no, case_id, tku_no=None, kapsam_durumu=None):
    db.add(models.CaseFoy(sistem_no=sistem_no, case_id=case_id, tku_no=tku_no,
                          kapsam_durumu=kapsam_durumu))
    db.commit()


def _tku_1230(db, kapsam_disi_foy=None):
    """test_iliskili_dava_otomatik deseni: üç kart, TKU-1230; istenen föy işaretli."""
    kartlar = []
    for tracking_no, sistem_no in (
        ("D1.B_GURER....0001.IDARE.00000", "id-7173"),
        ("D1.E_CELIKOGL.0001.IDARE.00000", "id-7174"),
        ("D1.J_HAZNECI..0001.IDARE.00000", "id-7175"),
    ):
        kart = _iliski_karti(db, tracking_no, file_type="İdare",
                             court="İstanbul 8. İdare Mahkemesi", esas_no="2020/2029")
        _foy(db, sistem_no, kart.id, "TKU-1230",
             kapsam_durumu="SILINDI" if sistem_no == kapsam_disi_foy else None)
        kartlar.append(kart)
    return kartlar


def test_kapsam_disi_foy_tku_iliskisi_uretmez_hedef_tarafinda(oturum):
    """Kabul: b'nin föyü kapsam dışı → a'nın ilişkilerinde b YOK, c var (mevcut ilişki
    testleri deseni). Esas ikizi dedektörü etkilenmez: kartlar esas+mahkeme ile
    yine bulunur — bu yüzden esas farklı verilir ki TKU'nun tek bağ olduğu görülsün."""
    a, b, c = _tku_1230(oturum, kapsam_disi_foy="id-7174")
    b.esas_no = "2019/1"
    c.esas_no = "2019/2"
    oturum.commit()
    assert {kart.id for kart, *_ in ilis.iliskileri_bul(oturum, a, TENANT)} == {c.id}


def test_kapsam_disi_foy_tku_iliskisi_uretmez_kaynak_tarafinda(oturum):
    """Kaynak kartın TEK TKU kaynağı kapsam dışı föyse TKU kümesi boştur."""
    a, b, c = _tku_1230(oturum, kapsam_disi_foy="id-7173")
    for kart, esas in ((a, "2019/0"), (b, "2019/1"), (c, "2019/2")):
        kart.esas_no = esas
    oturum.commit()
    assert ilis._tku_kumesi(oturum, a) == set()
    assert ilis.iliskileri_bul(oturum, a, TENANT) == []
    # Kontrol: işaret olmasaydı ilişki üretilirdi
    oturum.query(models.CaseFoy).filter_by(sistem_no="id-7173").one().kapsam_durumu = None
    oturum.commit()
    assert {kart.id for kart, *_ in ilis.iliskileri_bul(oturum, a, TENANT)} == {b.id, c.id}


def test_legacy_kart_tku_kolonu_kapsam_suzgecinden_etkilenmez(oturum):
    """`cases.tku_no` kapsam bilgisi taşımaz, olduğu gibi okunmaya devam eder."""
    a = _iliski_karti(oturum, "D1.A....0001.HUKUK.00000", file_type="Hukuk",
                      court="Ankara 3. Tüketici Mahkemesi", esas_no="2021/55", tku_no="TKU-9")
    b = _iliski_karti(oturum, "D1.B....0001.CEZA.00000", file_type="Ceza",
                      court="Ankara 1. Asliye Ceza Mahkemesi", esas_no="2021/900", tku_no="TKU-9")
    assert [kart.id for kart, *_ in ilis.iliskileri_bul(oturum, a, TENANT)] == [b.id]


@pytest.fixture
def kart_fabrikasi(monkeypatch):
    """In-memory sqlite; `case_manager.SessionLocal` yönlendirilir (test_g065 deseni)."""
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    models.Base.metadata.create_all(engine)
    Fabrika = sessionmaker(bind=engine)
    monkeypatch.setattr(case_manager, "SessionLocal", Fabrika)
    yield Fabrika
    engine.dispose()


def test_get_case_foyler_listesinde_uc_alan_kapsamda_foyde_null(kart_fabrikasi):
    """Kabul: `get_case` föy listesinde üç alan var; kapsamda föyde NULL; sıra sistem_no."""
    db = kart_fabrikasi()
    try:
        kart = models.Case(tracking_no="HA.G113.K", status="DERDEST",
                           maddi_tazminat=0, manevi_tazminat=0)
        db.add(kart)
        db.flush()
        db.add(models.CaseFoy(sistem_no="S-2", case_id=kart.id, tku_no="TKU-1",
                              kapsam_durumu="KAPSAM_DISI", kapsam_gerekcesi="malpraktis dışı",
                              kapsam_tarihi=date(2026, 8, 20)))
        db.add(models.CaseFoy(sistem_no="S-1", case_id=kart.id, tku_no="TKU-1", hasar_no="H-1"))
        db.commit()
        case_id = kart.id
    finally:
        db.close()

    sonuc = case_manager.get_case(case_id)
    assert sonuc is not None
    assert [f["sistem_no"] for f in sonuc["foyler"]] == ["S-1", "S-2"]
    s1, s2 = sonuc["foyler"]
    for f in (s1, s2):
        assert {"kapsam_durumu", "kapsam_gerekcesi", "kapsam_tarihi", "tku_no", "hasar_no",
                "case_party_id", "source", "id"} <= set(f)
    assert (s1["kapsam_durumu"], s1["kapsam_gerekcesi"], s1["kapsam_tarihi"]) == (None, None, None)
    assert (s2["kapsam_durumu"], s2["kapsam_gerekcesi"], s2["kapsam_tarihi"]) == (
        "KAPSAM_DISI", "malpraktis dışı", "2026-08-20")
    assert s1["hasar_no"] == "H-1"


def test_get_case_foysuz_kartta_bos_liste(kart_fabrikasi):
    db = kart_fabrikasi()
    try:
        kart = models.Case(tracking_no="HA.G113.B", status="DERDEST", maddi_tazminat=0, manevi_tazminat=0)
        db.add(kart)
        db.commit()
        case_id = kart.id
    finally:
        db.close()
    assert case_manager.get_case(case_id)["foyler"] == []


def test_case_read_semasinda_foyler_alani_var():
    from schemas import CaseRead

    okuma = CaseRead(id=1, tracking_no="HA.X.0001.2026", status="DERDEST", created_at="2026-09-03T00:00:00")
    assert okuma.foyler == []


# ═══════════════════════════════════════════════════════════════════════════
# 5. dbtest — gerçek Postgres (3-ortam kuralı: to_regclass + SKIP)
# ═══════════════════════════════════════════════════════════════════════════

@pytest.fixture(scope="module")
def pg():
    url = os.getenv("DATABASE_URL") or ""
    if not url.startswith("postgresql"):
        pytest.skip("DATABASE_URL postgresql:// değil")
    engine = create_engine(url, poolclass=NullPool, connect_args={"connect_timeout": 3})
    try:
        conn = engine.connect()
        conn.execute(text("SELECT 1"))
    except Exception as exc:
        engine.dispose()
        pytest.skip(f"Gerçek Postgres'e ulaşılamadı ({type(exc).__name__}) — G113 dbtest atlandı")
    try:
        var = conn.execute(text("SELECT to_regclass('public.case_foys')")).scalar()
    except Exception as exc:
        conn.close()
        engine.dispose()
        pytest.skip(f"Şema sorgulanamadı ({type(exc).__name__})")
    if var is None:
        conn.close()
        engine.dispose()
        pytest.skip("Şema göçmemiş — case_foys yok (migrasyon koşmamış)")
    conn.rollback()
    try:
        yield conn
    finally:
        conn.close()
        engine.dispose()


@pytest.mark.dbtest
def test_uc_kolon_gercek_semada_null_kabul_eder(pg):
    """Migration gerçek Postgres'te koştu: üç kolon var, tipleri ve NULL'luğu sözleşmeye uygun.
    (İkinci koşu idempotentliği: `("columns", ...)` op'u mevcut kolonu atlar — kolon ikinci
    kez eklenmeye çalışılsaydı ALTER hata verir, konteyner açılmazdı.)"""
    satirlar = pg.execute(text(
        "SELECT column_name, data_type, character_maximum_length, is_nullable, column_default "
        "FROM information_schema.columns WHERE table_name = 'case_foys' "
        "AND column_name IN ('kapsam_durumu', 'kapsam_gerekcesi', 'kapsam_tarihi')"
    )).all()
    if len(satirlar) < 3:
        pytest.skip("case_foys kapsam kolonları göçmemiş (migrasyon koşmamış)")
    kolonlar = {ad: (tip, uzunluk, nullable, varsayilan) for ad, tip, uzunluk, nullable, varsayilan in satirlar}
    assert kolonlar["kapsam_durumu"][:3] == ("character varying", 20, "YES")
    assert kolonlar["kapsam_gerekcesi"][0] == "character varying" and kolonlar["kapsam_gerekcesi"][2] == "YES"
    assert kolonlar["kapsam_tarihi"][:1] == ("date",) and kolonlar["kapsam_tarihi"][2] == "YES"
    assert all(v[3] is None for v in kolonlar.values())            # DEFAULT yok: NULL = kapsamda

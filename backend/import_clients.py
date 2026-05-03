"""
Müvekkil Import Scripti
=======================
Excel dosyasından (cari_mikro_guncellendi.xlsx) müvekkilleri okur,
mevcut clients tablosunu temizler (TRUNCATE) ve taze import yapar.

Kullanım:
    python import_clients.py --file "C:/path/to/cari_mikro_guncellendi.xlsx"

Kolon eşleştirmesi:
    Col 1  (Vergi No sayısal)   → cari_kod
    Col 2  (Vergi No isim)      → name
    Col 3  (Tür)                → client_type  ("Şahıs"→"Individual", "Kurum"→"Corporate")
    Col 4  (e-Posta)            → email
    Col 5  (Cep Telefonu)       → mobile_phone
    Col 6  (Telefon)            → phone
    Col 7  (Adres)              → address
    Col 8  (İl)                 → il
    Col 9  (TC / Vergi No)      → tc_no
    Col 10 (Sektörü)            → sektor
    Col 11 (Grup)               → specialty
    Col 12 (Özel Kod)           → category
    Col 13 (YEVMİYE NO)        → yevmiye_no
    Col 14 (NOTERLİK)           → noterlik
    Col 15 (VERİLİŞ TARİHİ)    → vekaletname_tarihi
    Col 16 (VEKİL AVUKATLAR)   → vekil_avukatlar  (normalize → "AD;AD" formatı)
    Col 17 (GEÇERLİLİK TAR.)   → gecerlilik_tarihi
    Col 18 (VEKALET NO)        → vekalet_no
    Col 19 (BÜRO VEKALET NO)   → buro_vekalet_no
    Col 20 (VEKALET AÇIKL.)    → kapsam dışı, import edilmez
"""

import argparse
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

# ── Argümanları EN ÖNCE parse et (database import'undan önce) ──────────────
# database.py engine'i modül yüklenirken oluşturur; DATABASE_URL o anda
# hazır olmalı. Bu yüzden --db-url argümanı DB import'undan önce işlenmeli.
_parser = argparse.ArgumentParser(description="Müvekkil Excel Import", add_help=False)
_parser.add_argument("--file", required=True)
_parser.add_argument("--db-url", default=None)
_early_args, _ = _parser.parse_known_args()

# .env yükle — önce backend/, sonra proje root'unu dene
from dotenv import load_dotenv
_script_dir = Path(__file__).parent
for _env_candidate in [_script_dir / ".env", _script_dir.parent / ".env"]:
    if _env_candidate.exists():
        load_dotenv(_env_candidate)
        print(f"⚙️  .env yüklendi: {_env_candidate}")
        break

# --db-url verilmişse DATABASE_URL'yi override et (docker host'u bypass)
if _early_args.db_url:
    os.environ["DATABASE_URL"] = _early_args.db_url
    _display = _early_args.db_url.split("@")[1] if "@" in _early_args.db_url else _early_args.db_url
    print(f"⚙️  DATABASE_URL override: {_display}")

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl yüklü değil. Çalıştır: pip install openpyxl")
    sys.exit(1)

# DB importları DATABASE_URL set edildikten SONRA yapılıyor
from sqlalchemy import text
from database import SessionLocal, engine
import models  # noqa: F401 — Base.metadata'nın tabloları görmesi için


# ---------------------------------------------------------------------------
# Yardımcı fonksiyonlar
# ---------------------------------------------------------------------------

TUR_MAP = {
    "şahıs": "Individual",
    "sahis": "Individual",
    "kurum": "Corporate",
}


def map_client_type(raw: str) -> str:
    if not raw:
        return None
    return TUR_MAP.get(raw.strip().lower(), raw.strip())


def parse_date(raw) -> date | None:
    """DD.MM.YYYY formatındaki string'i date nesnesine çevirir."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_vekil(raw) -> str | None:
    """
    İki formatı kısa "AD SOYAD;AD SOYAD" formatına normalleştirir.
    - Uzun: "Ayse Acar Yucel ( T.C. Kimlik No: 123...)\n..."
    - Kısa: "AYŞE GÜL HANYALOĞLU;TUGÇE ÜNGÖR YANIK"
    """
    if not raw:
        return None
    raw = str(raw).strip()
    if not raw:
        return None

    # Uzun format tespiti: TC kelimesi veya çok satırlı
    if "\n" in raw or "T.C." in raw.upper() or "KİMLİK NO" in raw.upper() or "KIMLIK NO" in raw.upper():
        names = []
        for line in raw.split("\n"):
            line = line.strip()
            if not line:
                continue
            # Parantez ve içini at: "Ad Soyad ( T.C. ... )" → "Ad Soyad"
            name = re.sub(r"\s*\(.*?\)\s*", "", line).strip()
            if name:
                names.append(name.upper())
        return ";".join(names) if names else None

    # Zaten kısa format
    return raw.strip()


def str_val(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Ana import fonksiyonu
# ---------------------------------------------------------------------------

def run_migration():
    """Yeni kolonları DB'ye ekler (henüz yoksa)."""
    from sqlalchemy import inspect
    new_columns = {
        "il":                  "VARCHAR(100)",
        "sektor":              "VARCHAR(200)",
        "yevmiye_no":          "VARCHAR(50)",
        "noterlik":            "VARCHAR(200)",
        "vekaletname_tarihi":  "DATE",
        "vekil_avukatlar":     "TEXT",
        "gecerlilik_tarihi":   "DATE",
        "vekalet_no":          "VARCHAR(50)",
        "buro_vekalet_no":     "VARCHAR(50)",
    }
    with engine.connect() as conn:
        inspector = inspect(engine)
        existing = [col["name"] for col in inspector.get_columns("clients")]
        added = []
        for col_name, col_type in new_columns.items():
            if col_name not in existing:
                conn.execute(text(f"ALTER TABLE clients ADD COLUMN {col_name} {col_type}"))
                added.append(col_name)
        if added:
            conn.commit()
            print(f"   ✅ Migration: {len(added)} kolon eklendi → {', '.join(added)}")
        else:
            print("   ✅ Migration: tüm kolonlar zaten mevcut.")


def import_clients(filepath: str):
    if not os.path.exists(filepath):
        print(f"❌ Dosya bulunamadı: {filepath}")
        sys.exit(1)

    print("🔧 Migration kontrol ediliyor...")
    run_migration()

    print(f"📂 Dosya okunuyor: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1  # header hariç
    print(f"   Toplam kayıt: {total_rows}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    db = SessionLocal()
    errors = []

    try:
        # 1. TRUNCATE — case_parties'deki client_id referanslarını önce NULL yap
        print("🗑️  Mevcut müvekkiller temizleniyor...")
        db.execute(text("UPDATE case_parties SET client_id = NULL WHERE client_id IS NOT NULL"))
        db.execute(text("DELETE FROM clients"))
        db.commit()
        print("   ✅ Temizlendi.")

        # 2. INSERT
        print("📥 Import başlıyor...")
        inserted = 0

        for i, row in enumerate(rows, start=2):  # satır numarası (header=1)
            try:
                # Sütunları güvenli al
                def col(idx):
                    return row[idx] if idx < len(row) else None

                cari_kod        = str_val(col(0))
                name            = str_val(col(1))
                client_type_raw = str_val(col(2))
                email           = str_val(col(3))
                mobile_phone    = str_val(col(4))
                phone           = str_val(col(5))
                address         = str_val(col(6))
                il              = str_val(col(7))
                tc_no           = str_val(col(8))
                sektor          = str_val(col(9))
                specialty       = str_val(col(10))
                category        = str_val(col(11))
                yevmiye_no      = str_val(col(12))
                noterlik        = str_val(col(13))
                vekaletname_tarihi = parse_date(col(14))
                vekil_avukatlar    = normalize_vekil(col(15))
                gecerlilik_tarihi  = parse_date(col(16))
                vekalet_no         = str_val(col(17))
                buro_vekalet_no    = str_val(col(18))
                # Col 19 (VEKALET AÇIKLAMALAR) kapsam dışı

                if not name:
                    errors.append(f"Satır {i}: name boş, atlandı.")
                    continue

                client = models.Client(
                    name=name,
                    cari_kod=cari_kod,
                    client_type=map_client_type(client_type_raw),
                    email=email,
                    mobile_phone=mobile_phone,
                    phone=phone,
                    address=address,
                    il=il,
                    tc_no=tc_no,
                    sektor=sektor,
                    specialty=specialty,
                    category=category,
                    yevmiye_no=yevmiye_no,
                    noterlik=noterlik,
                    vekaletname_tarihi=vekaletname_tarihi,
                    vekil_avukatlar=vekil_avukatlar,
                    gecerlilik_tarihi=gecerlilik_tarihi,
                    vekalet_no=vekalet_no,
                    buro_vekalet_no=buro_vekalet_no,
                    active=True,
                    contact_type="Client",
                )
                db.add(client)
                inserted += 1

                # Her 200 kayıtta bir commit (bellek yönetimi)
                if inserted % 200 == 0:
                    db.commit()
                    db.expunge_all()
                    print(f"   ... {inserted} kayıt eklendi")

            except Exception as e:
                errors.append(f"Satır {i}: {e}")
                continue

        db.commit()
        print(f"\n✅ Import tamamlandı!")
        print(f"   Eklenen kayıt : {inserted}")
        print(f"   Hatalı satır  : {len(errors)}")
        if errors:
            print("\n⚠️  Hatalar:")
            for err in errors[:20]:
                print(f"   {err}")
            if len(errors) > 20:
                print(f"   ... ve {len(errors) - 20} hata daha")

    except Exception as e:
        db.rollback()
        print(f"❌ Kritik hata, rollback yapıldı: {e}")
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Tam argüman parse (help dahil)
    parser = argparse.ArgumentParser(description="Müvekkil Excel Import")
    parser.add_argument(
        "--file",
        required=True,
        help='Excel dosyasının tam yolu (ör: "C:/Users/.../cari_mikro_guncellendi.xlsx")',
    )
    parser.add_argument(
        "--db-url",
        default=None,
        help='PostgreSQL bağlantı URL\'si. '
             'Local için: "postgresql://hukudok_user:SIFRE@localhost:5432/hukudok"',
    )
    args = parser.parse_args()
    import_clients(args.file)

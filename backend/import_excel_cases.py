"""
BIRLESIK_SONUC_v5_temiz.xlsx → cases DB import scripti
Kullanım:
  python import_excel_cases.py              # Tam import
  python import_excel_cases.py --dry-run    # DB'ye yazmadan logla
  python import_excel_cases.py --limit 100  # İlk 100 satır
"""

import sys
import os
import argparse
import re
from datetime import datetime, date

sys.stdout.reconfigure(encoding="utf-8")
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv("../.env")

import openpyxl
import models
from database import SessionLocal

# ─── CONFIG ──────────────────────────────────────────────────────────────────

EXCEL_PATH = r"C:\Users\ilkeb\OneDrive\Masaüstü\BIRLESIK_SONUC_v5_temiz.xlsx"
SHEET_NAME = "Son Liste"

# ─── TRACKING NUMBER ─────────────────────────────────────────────────────────

CATEGORY_MAP = {
    "DOKTOR": "D1",
    "ÖZEL HASTANE": "H2",
    "SİGORTA": "S0",
    "HASTA": "H1",
}

INSURANCE_CODES = {
    "AK": "1", "ANADOLU": "2", "AXA": "3",
    "CORPUS": "4", "QUICK": "4", "EUREKO": "5",
    "NIPPON": "6", "SOMPO": "7",
}

PROCESS_MAP = {
    "Hukuk": "HUKUK",
    "İdari Yargı": "IDARI",
    "İdare": "IDARE",
    "Ceza": "CEZAA",
    "İcra": "ICRAA",
    "Arabuluculuk": "ARABU",
    "Savcılık": "SAVCI",
    "Tahkim": "TAHKM",
    "Vergi": "VERGI",
    "Danışmanlık": "DANIS",
}

def _tr_upper(s: str) -> str:
    return s.replace("ı", "I").replace("i", "İ").replace("ğ", "Ğ").replace("ü", "Ü") \
            .replace("ş", "Ş").replace("ö", "Ö").replace("ç", "Ç").upper()

def _slugify_name(name: str) -> str:
    if not name:
        return "XXXXXXXXXX"
    clean = re.sub(r"[^A-Z\s]", "", _normalize_ascii(name).upper())
    parts = clean.split()
    if not parts:
        return "XXXXXXXXXX"
    if len(parts) == 1:
        return parts[0].ljust(10, ".")[:10]
    surname = parts[-1]
    first_initial = parts[0][0]  # sadece ilk ismin baş harfi
    return (f"{first_initial}_{surname}").ljust(10, ".")[:10]

def _normalize_ascii(s: str) -> str:
    return (s.replace("ı", "i").replace("ğ", "g").replace("ü", "u")
             .replace("ş", "s").replace("ö", "o").replace("ç", "c")
             .replace("İ", "I").replace("Ğ", "G").replace("Ü", "U")
             .replace("Ş", "S").replace("Ö", "O").replace("Ç", "C"))

def generate_tracking_number(client_name: str, category: str, sequence: int, process_type: str) -> str:
    norm_cat = _tr_upper(category or "")
    norm_name = _tr_upper(client_name or "")

    block1 = "X1"
    for key, val in CATEGORY_MAP.items():
        if _tr_upper(key) in norm_cat:  # normalize key too (Ö→O, İ→I etc.)
            block1 = val
            break

    is_sigorta = "SİGORTA" in norm_cat or "SİGORTA" in norm_name or "SIGORTA" in norm_name
    if is_sigorta:
        if block1 == "X1":
            block1 = "S0"
        for key, code in INSURANCE_CODES.items():
            if key in norm_name:
                block1 = f"S{code}"
                break

    block2 = _slugify_name(client_name or "")
    block3 = str(sequence).zfill(4)
    block4 = PROCESS_MAP.get(process_type or "", "HUKUK")
    block5 = "00000"

    return f"{block1}.{block2}.{block3}.{block4}.{block5}"

# ─── STATUS MAPPING ──────────────────────────────────────────────────────────

STATUS_MAP = {
    "Aktif":  "DERDEST",
    "Arşiv":  "MAHZEN",
}

# ─── KARAR_TURU MAPPING ──────────────────────────────────────────────────────

KARAR_TURU_MAP = {
    "kabul":                    "KABUL",
    "kabul/kısmen":             "KISMI_KABUL",
    "red/esastan":              "RED",
    "red/görev":                "RED",
    "red/husumet":              "RED",
    "red/zamanaşımı":           "RED",
    "red/dilekçenin reddi":     "RED",
    "red/msk kararı gereği":    "RED",
    "red/arabuluculuk ön şart": "RED",
    "red/feragat":              "FERAGAT",
    "feragat":                  "FERAGAT",
    "beraat":                   "RED",
    "anlaşmama":                None,
    "derdest":                  None,
    "kapalı":                   None,
}

def map_karar_turu(val: str) -> str | None:
    if not val:
        return None
    return KARAR_TURU_MAP.get(val.strip().lower(), None)

# ─── DATE PARSING ────────────────────────────────────────────────────────────

DATE_FORMATS = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y%m%d"]

def parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# ─── PARTY PARSING ───────────────────────────────────────────────────────────

def split_names(val) -> list[str]:
    if not val:
        return []
    return [n.strip() for n in str(val).split(";") if n.strip()]

# ─── MAIN IMPORT ─────────────────────────────────────────────────────────────

def run(dry_run: bool = False, limit: int = None):
    print(f"{'[DRY-RUN] ' if dry_run else ''}Excel okunuyor: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    db = SessionLocal()

    added = 0
    errors = 0
    # Per-client sequence tracker
    client_seq: dict[str, int] = {}

    rows = ws.iter_rows(min_row=2, values_only=True)

    try:
        for row_idx, row in enumerate(rows, start=2):
            if limit and row_idx - 2 >= limit:
                break

            def get(col_name):
                idx = col.get(col_name)
                return row[idx] if idx is not None else None

            try:
                klasor_no_2    = get("Klasör No.2")
                muvekkil_str   = get("Müvekkil")
                karsi_str      = get("Karşı Taraf")
                diger_str      = get("Diğer Davalı")
                tarafimiz      = get("Tarafımız")
                ana_tur        = get("Ana Tür")
                dava_konusu    = get("Dava Konusu")
                alt_kirilim    = get("Alt Kırılım")
                ek_alt         = get("Ek Alt Kırılım")
                buro_ozel      = get("Büro Özel Türü")
                mahkeme        = get("Mahkemesi")
                esas_no        = get("Esas Numarası")
                durum          = get("Durum")
                son_durum      = get("Son Durum")
                ymkd           = get("Yerel Mahkeme Karar Durumu")
                dava_tarihi    = get("Dava Tarihi")
                is_kabul       = get("İş Kabul Tarihi")
                atama          = get("Atama Tarihi")
                dosya_ilgilisi = get("Dosya İlgilisi")
                hasar_dosya    = get("Hasar Dosya Numarası")
                hukuk_no       = get("Hukuk Numarası")

                # Müvekkil isimlerini ayır
                muvekkil_names = split_names(muvekkil_str)
                primary_client = muvekkil_names[0] if muvekkil_names else ""

                # Sequence (per müvekkil)
                client_key = _normalize_ascii(primary_client).upper().strip()
                client_seq[client_key] = client_seq.get(client_key, 0) + 1
                seq = client_seq[client_key]

                # Tracking no üret
                tracking_no = generate_tracking_number(
                    client_name=primary_client,
                    category="",   # kategori bilgisi Excel'de yok, X1 default
                    sequence=seq,
                    process_type=str(ana_tur) if ana_tur else "",
                )

                # Duplicate tracking_no koruması
                if not dry_run:
                    existing = db.query(models.Case).filter_by(tracking_no=tracking_no).first()
                    if existing:
                        suffix = 2
                        while db.query(models.Case).filter_by(tracking_no=f"{tracking_no}-{suffix}").first():
                            suffix += 1
                        tracking_no = f"{tracking_no}-{suffix}"

                status = STATUS_MAP.get(str(durum).strip() if durum else "", "MAHZEN")

                case = models.Case(
                    tracking_no            = tracking_no,
                    klasor_no_2            = str(klasor_no_2).strip() if klasor_no_2 else None,
                    esas_no                = str(esas_no).strip() if esas_no else None,
                    status                 = status,
                    file_type              = str(ana_tur).strip() if ana_tur else None,
                    subject                = str(dava_konusu).strip() if dava_konusu else None,
                    sub_type               = str(alt_kirilim).strip() if alt_kirilim else None,
                    sub_type_extra         = str(ek_alt).strip() if ek_alt else None,
                    bureau_type            = str(buro_ozel).strip() if buro_ozel else None,
                    court                  = str(mahkeme).strip() if mahkeme else None,
                    opening_date           = parse_date(dava_tarihi),
                    acceptance_date        = parse_date(is_kabul),
                    atama_tarihi           = parse_date(atama),
                    responsible_lawyer_name= str(dosya_ilgilisi).strip() if dosya_ilgilisi else None,
                    hasar_dosya_no         = str(hasar_dosya).strip() if hasar_dosya else None,
                    hukuk_no               = str(hukuk_no).strip() if hukuk_no else None,
                    dosya_son_durumu       = str(son_durum).strip() if son_durum else None,
                    karar_turu             = map_karar_turu(str(ymkd) if ymkd else ""),
                    active                 = True,
                )

                if not dry_run:
                    db.add(case)
                    db.flush()  # case.id'yi al

                # Taraflar
                parties = []

                # Müvekkil(ler) — CLIENT
                for i, name in enumerate(muvekkil_names):
                    role = str(tarafimiz).strip() if (i == 0 and tarafimiz) else "Müvekkil"
                    parties.append(models.CaseParty(
                        case_id    = case.id if not dry_run else 0,
                        name       = name,
                        role       = role,
                        party_type = "CLIENT",
                    ))

                # Karşı taraf(lar) — COUNTER
                for name in split_names(karsi_str):
                    parties.append(models.CaseParty(
                        case_id    = case.id if not dry_run else 0,
                        name       = name,
                        role       = "Karşı Taraf",
                        party_type = "COUNTER",
                    ))

                # Diğer davalı(lar) — THIRD
                for name in split_names(diger_str):
                    parties.append(models.CaseParty(
                        case_id    = case.id if not dry_run else 0,
                        name       = name,
                        role       = "Diğer Davalı",
                        party_type = "THIRD",
                    ))

                if not dry_run:
                    db.add_all(parties)

                added += 1

                if added % 500 == 0:
                    if not dry_run:
                        db.commit()
                    print(f"  {added} kayıt işlendi...")

            except Exception as e:
                errors += 1
                print(f"  [HATA] Satır {row_idx}: {e}")
                if not dry_run:
                    db.rollback()

        if not dry_run:
            db.commit()

    except Exception as e:
        print(f"[KRİTİK HATA] {e}")
        if not dry_run:
            db.rollback()
    finally:
        db.close()

    print()
    print("=" * 50)
    print(f"{'[DRY-RUN] ' if dry_run else ''}TAMAMLANDI")
    print(f"  ✓ Eklendi : {added}")
    print(f"  ✗ Hatalı  : {errors}")
    print("=" * 50)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="DB'ye yazmadan çalıştır")
    parser.add_argument("--limit", type=int, default=None, help="Kaç satır işlensin")
    args = parser.parse_args()
    run(dry_run=args.dry_run, limit=args.limit)

"""Veri teslim gelen kutusu — defter + durum makinesi (G107).

Plan: `docs/plan/veri-teslim-otomasyonu-plani-2026-09-03.md` §2.1–2.2. Veri
ekibinin bıraktığı teslim paketi (xlsx) `aktarim_teslimleri` defterine ve
spool'a girer; buradaki fonksiyonlar durum makinesini yürütür:

    alindi → dogrulandi → kuru_kosuldu → [kapı] → uygulaniyor → uygulandi
                 │              │                      │
             reddedildi   inceleme_bekliyor        basarisiz
    yinelenen (aynı sha256 daha önce alınmış; nihai, işlenmez)

Bu modül **hiçbir uç ve zamanlayıcı açmaz**: admin uçları G108'in, cevap
paketi G110'un işidir; SharePoint gözcüsü (`sharepoint_tara`), gece turu
(`gece_turu`) ve boot telafisi (`boot_catch_up`) G109 ile buraya geldi ama
zamanlayıcı KAYDI `api.py` lifespan'indedir (lider worker, 04:00 TR). Gerçek
yazma yolu `scripts/hukdok_aktarim.aktarimi_kos` (G064) — burada YALNIZ
import edilir, değiştirilmez.

Tasarım kararları
-----------------
* **Oturum modeli** `deadline_scanner` deseni: her fonksiyon senkron, `db`
  verilmezse `SessionLocal` ile kendi oturumunu açıp kapatır; `db` yalnız
  test/çağıran enjeksiyonu içindir ve o oturumda DA commit edilir (durum
  makinesi commit'e dayanır — `uygulaniyor` çökme izi commit'siz iz değildir).
* **Aktarım ayrı bağlantıda koşar** (`_aktarimi_calistir`): `aktarimi_kos`
  oturum FABRİKASI ister ve `statement_timeout`'u oturum boyu (`set_config …
  false`) yükseltir. Havuzdan gelen bir bağlantı o ayarla havuza dönerse Faz
  3-E'nin 30 sn koruması o bağlantı için sessizce kalkardı; bu yüzden aktarım
  için bağlantı açıkça alınır, iş bitince `RESET statement_timeout` ile
  varsayılana döndürülür ve kapatılır. Defter oturumu aktarım süresince
  KAPALI transaction'dadır (önce commit) — sqlite StaticPool'da (testler)
  ikisi aynı DBAPI bağlantısını paylaşır, iç içe BEGIN patlardı.
* **Yinelenen içerik yeni satırdır**: aynı sha256 ikinci kez gelince mevcut
  satıra dokunulmaz, `yinelenen` durumlu yeni satır açılır ve notunda ilk id
  yazar (veri ekibi "dosyayı iki kez bıraktım" dediğinde defter bunu gösterir).
  Spool'a yazılmaz. Tekillik kısmi UNIQUE index'le yalnız yinelenen-dışı
  satırlarda (database.py madde 39).
* **Log sözleşmesi**: deneme/yapı düzeyi başarısızlık WARNING (`reddedildi`
  dahil — yapı hatası veri ekibinin düzelteceği bir şeydir, nöbetçi alarmı
  değil); nihai `basarisiz` teslim başına TEK ERROR. Envanter kapısı kırmızı
  çıktığında o ERROR'u `aktarimi_kos` zaten basar ("Aktarım GERİ ALINDI") —
  burada ikinci ERROR YAZILMAZ, yalnız defter işlenir.
* **Kapı** (plan §2.2) kuralların HEPSİNİ değerlendirir ve gerekçeyi `;` ile
  birleştirir: admin "neden inceleme" sorusuna tek bakışta cevap alsın, ilk
  ihlalde durup diğerlerini gizlemesin. Eşikler ENV'den ÇAĞRI ANINDA okunur
  (`kapi_esikleri()`; sözleşmedeki `KAPI_ESIKLERI` adı aynı fonksiyonun
  takma adıdır) — recreate'siz `.env` değişikliği yine gelmez, ama test ve
  admin paneli monkeypatch/os.environ ile anlık değer görür.
* **Bildirim** (G108, `bildir`): eşik dışı (`inceleme_bekliyor`) ve nihai
  (`reddedildi` / `basarisiz` / `uygulandi`) durumlarda ADMIN_EMAILS kümesine
  uygulama içi bildirim düşer (`services.notifications.create_notification`,
  `type="veri_teslim"`). `dedupe_key` teslim + durum + ALICI üçlüsüdür (G082
  dersi: anahtar global tekil, alıcı sonda) — aynı geçiş ikinci kez satır
  ikilemez. Bildirim yan üründür: her türlü hatası WARNING ile yutulur, durum
  makinesi bozulmaz. `yinelenen` bildirim üretmez (bilgi defterde, alarm değil).
* **SharePoint gözcüsü** (G109, `sharepoint_tara`): `<SHAREPOINT_FOLDER_TESLIM_NAME>/gelen`
  klasörü `list_folder_children` ile listelenir; ad kalıbı `HUKDOK_TESLIM_*.xlsx`
  (harf duyarsız) dışındakiler `atlanan`. Ucuz eleme: `sharepoint_item_id`
  kolonuna driveItem id'si ile eTag BİRLİKTE (`<id>@<eTag>`) yazılır — aynı
  anahtar defterdeyse dosya indirilmez (`yinelenen` sayılır); eTag değiştiyse
  (dosya yerinde güncellendi) indirilir ve sha256 eşleşirse `teslim_kaydet`
  zaten `yinelenen` satırı açar. Ayrı eTag kolonu için model/migrasyon
  değişikliği görev kapsamı dışıydı; kolon ID+sürüm anahtarıdır, panel bu
  alanı göstermez. Tek dosyanın hatası WARNING, tur devam eder; LİSTELEME
  hatası yükselir — tur düzeyinde tek ERROR'a (gece) ya da WARNING'e (boot)
  çağıran karar verir. Anahtar (`veri_teslim_otomasyonu`) kapalıysa hiç listelenmez.
* **Gece turu** (`gece_turu`, 04:00 TR): `acilis_toparla` → `sharepoint_tara` →
  bekleyen (`alindi`/`dogrulandi`/`kuru_kosuldu`, `created_at` sırası) her
  satıra `teslimi_isle(otomatik_uygula=…)`. **Aynı turda en fazla BİR teslim
  uygulanır**: ilki uygulandıysa sonrakiler `otomatik_uygula=False` ile koşar
  ve kapı "otomatik" dese bile `inceleme_bekliyor`a alınır (gerekçe
  `tek_uygulama`) — ikinci paket aynı gece geldiyse insan baksın. Anahtar
  kapalıysa tur hiçbir durum değiştirmez. `inceleme_bekliyor` satırlarına
  dokunulmaz (insan bekliyor; her gece yeniden kuru koşturmak boşuna).
* **Boot telafisi** (`boot_catch_up`): lider açılışında daemon thread'de bir
  kez; `acilis_toparla` (anahtardan bağımsız — kesilmiş elle uygulama da
  toparlanmalı) + `sharepoint_tara` + yalnız `alindi`/`dogrulandi` satırlara
  `teslimi_isle(otomatik_uygula=False)`. **Uygulama yalnız cron'da** (plan
  §2.3); `kuru_kosuldu` satırlar her restart'ta yeniden kuru koşturulmaz.
  Her istisna tek WARNING ile yutulur (`deadline_scanner.boot_catch_up_scan`).
* **Cevap paketi** (G110, `services/teslim_cevap.py`): rapor dizinine her kuru koşu ve
  uygulamada `ozet.txt` (`ozet_metni` + "kapı kararı" satırı; kapı değerlendirilince
  satır tazelenir) yazılır. `teslim_uygula` başarı yolunda cevap yüklemesini BİR kez
  dener (admin "Uygula" da buradan geçer); `gece_turu` sonunda `uygulandi` +
  `cevap_yuklendi=False` kalan teslimler (bu turda uygulanan HARİÇ — az önce denendi)
  yeniden denenir, retry böylece ertesi geceye kalır. Yükleme hatası teslim durumunu
  DEĞİŞTİRMEZ (WARNING). `teslim_cevap` bu modülü import eder; buradaki çağrılar döngü
  olmasın diye fonksiyon içinde import edilir.
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import unicodedata
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterator, List, Optional, Tuple

from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, sessionmaker

from database import SessionLocal, engine as _varsayilan_engine
from managers.reference_lists import tr_upper
import models
from required_fields import AKTARIM_SOURCE_PREFIX
from scripts import hukdok_aktarim
from services import app_settings, belge_envanteri
from sharepoint import sharepoint_uploader_graph as _spu

logger = logging.getLogger(__name__)

# ─── Durumlar (plan §2.1) ────────────────────────────────────────────────────
DURUM_ALINDI = "alindi"
DURUM_YINELENEN = "yinelenen"
DURUM_REDDEDILDI = "reddedildi"
DURUM_DOGRULANDI = "dogrulandi"
DURUM_KURU_KOSULDU = "kuru_kosuldu"
DURUM_INCELEME = "inceleme_bekliyor"
DURUM_UYGULANIYOR = "uygulaniyor"
DURUM_UYGULANDI = "uygulandi"
DURUM_BASARISIZ = "basarisiz"

DURUMLAR: Tuple[str, ...] = (
    DURUM_ALINDI, DURUM_YINELENEN, DURUM_REDDEDILDI, DURUM_DOGRULANDI,
    DURUM_KURU_KOSULDU, DURUM_INCELEME, DURUM_UYGULANIYOR, DURUM_UYGULANDI,
    DURUM_BASARISIZ,
)
#: Nihai durumlar — `done_at` yazılır, bir daha işlenmez.
NIHAI_DURUMLAR = frozenset({DURUM_YINELENEN, DURUM_REDDEDILDI, DURUM_UYGULANDI, DURUM_BASARISIZ})
#: `teslimi_isle`'nin baştan (doğrulamadan) ele alabildiği durumlar.
ISLENEBILIR_DURUMLAR = frozenset({DURUM_ALINDI, DURUM_DOGRULANDI, DURUM_KURU_KOSULDU, DURUM_INCELEME})
#: Gece turunun / boot telafisinin taradığı "bekleyen" kümesi (partial index ile aynı liste).
BEKLEYEN_DURUMLAR = (DURUM_ALINDI, DURUM_DOGRULANDI, DURUM_KURU_KOSULDU, DURUM_INCELEME)
#: Gece turunun baştan (doğrula → kuru koş → kapı → uygula) ele aldığı durumlar —
#: `inceleme_bekliyor` DIŞARIDA (insan kararı bekliyor, her gece yeniden koşturulmaz).
GECE_ISLENEN_DURUMLAR = (DURUM_ALINDI, DURUM_DOGRULANDI, DURUM_KURU_KOSULDU)
#: Boot telafisinin ele aldığı durumlar — yalnız henüz kuru koşulmamış olanlar
#: (`kuru_kosuldu` gece uygulanmayı bekliyor; her restart'ta yeniden koşturmak boşuna).
BOOT_ISLENEN_DURUMLAR = (DURUM_ALINDI, DURUM_DOGRULANDI)

#: SharePoint teslim klasörü (env `SHAREPOINT_FOLDER_TESLIM_NAME` yoksa) ve gelen alt klasörü.
TESLIM_KLASORU_VARSAYILAN = "03_VERI_TESLIM"
TESLIM_GELEN_ALT_KLASORU = "gelen"
#: Gözcünün aldığı dosya adı kalıbı (harf duyarsız); dışındakiler `atlanan`.
TESLIM_AD_KALIBI = re.compile(r"^HUKDOK_TESLIM_.*\.xlsx$", re.IGNORECASE)
#: Gece turunda ikinci uygulanabilir teslimi incelemeye alan kural etiketi.
KAPI_TEK_UYGULAMA = "tek_uygulama"

_KURU_KOS_DURUMLARI = frozenset({DURUM_DOGRULANDI, DURUM_KURU_KOSULDU, DURUM_INCELEME})
_KAPI_DURUMLARI = frozenset({DURUM_KURU_KOSULDU, DURUM_INCELEME})
_UYGULA_DURUMLARI = frozenset({DURUM_KURU_KOSULDU, DURUM_INCELEME})

KAYNAKLAR: Tuple[str, ...] = ("sharepoint", "yukleme")

KAPI_OTOMATIK = "otomatik"
KAPI_INCELEME = "inceleme"
#: Kapı kuralı adları — `kapi_gerekcesi` bu etiketlerle başlar (G111 paneli okur).
KAPI_KURALLARI: Tuple[str, ...] = (
    "envanter_denk_degil", "ilk_teslim", "zincir_eksik", "bos_teslim",
    "hata_orani", "eslesmeyen_orani", "alan_degisikligi",
)

#: Gece turu otomatik uygularken `uygulayan` kolonuna yazılan değer.
GECE_UYGULAYAN = "gece-job"

#: Bildirim tür etiketi (frontend filtresi / G111 paneli bunu tüketir).
BILDIRIM_TURU = "veri_teslim"
#: Bildirim üreten geçişler → (başlık fiili, severity). Diğer durumlar sessizdir.
BILDIRIM_OLAYLARI: dict = {
    DURUM_INCELEME: ("inceleme bekliyor", "warning"),
    DURUM_BASARISIZ: ("başarısız", "warning"),
    DURUM_REDDEDILDI: ("reddedildi", "warning"),
    DURUM_UYGULANDI: ("uygulandı", "info"),
}

#: Teslim paketinin veri sayfası ve değişiklik özeti sayfası (plan §1, §3).
VERI_SAYFASI = "Sheet"
OZET_SAYFASI = "DEGISIKLIK_OZETI"
#: `xlsx_oku`'nun bulması ZORUNLU alanlar (script yalnız sistem_no'yu zorunlu
#: sayar; dosya_no eşleştirme köprüsüdür, onsuz her satır atlanırdı).
ZORUNLU_BASLIKLAR: Tuple[str, ...] = ("sistem_no", "dosya_no")
#: Özet sayfasında "Önceki teslim" etiketi bu kadar satır içinde aranır.
_OZET_TARAMA_SATIRI = 200
_ONCEKI_ETIKET = "ONCEKITESLIM"
_YER_TUTUCULAR = frozenset({"-", "--", "—", "–", "YOK", "N/A", "NA", "İLK", "ILK"})

HATA_MESAJI_SINIRI = 2000
_GUVENSIZ_AD = re.compile(r'[\\/:*?"<>|\x00-\x1f]')

#: Rapor dizinindeki koşu özeti (`ozet_metni` + kapı kararı satırı) — cevap paketinde
#: `ozet_<teslim>.txt` adıyla yüklenir (G110).
OZET_DOSYASI = "ozet.txt"
_KAPI_SATIRI_ONEKI = "  kapı kararı       : "
_KAPI_HENUZ_YOK = "henüz değerlendirilmedi"


# ═══════════════════════════════════════════════════════════════════════════
# Yardımcılar
# ═══════════════════════════════════════════════════════════════════════════

def get_teslim_spool_dir() -> Path:
    """Spool dizini: TESLIM_SPOOL_DIR env'i ya da <backend>/data/teslim_spool.

    Konteynerde <backend> = /app ve /app/data backend-data volume'üdür →
    spool konteyner recreate'ini atlatır (upload_queue.get_spool_dir ikizi).
    """
    override = os.getenv("TESLIM_SPOOL_DIR", "").strip()
    if override:
        spool = Path(override)
    else:
        spool = Path(__file__).resolve().parent.parent / "data" / "teslim_spool"
    spool.mkdir(parents=True, exist_ok=True)
    return spool


def _env_sayi(ad: str, varsayilan: float, donustur: Callable[[str], Any]) -> Any:
    ham = os.getenv(ad, "").strip()
    if not ham:
        return varsayilan
    try:
        return donustur(ham)
    except ValueError:
        logger.warning("%s=%r sayı değil — varsayılan %s kullanılıyor", ad, ham, varsayilan)
        return varsayilan


def kapi_esikleri() -> dict:
    """Kapı eşikleri (plan §2.2) — env'den ÇAĞRI ANINDA okunur.

    Anahtarlar G108 sözleşmesindeki `esikler` nesnesiyle aynıdır.
    """
    return {
        "hata_orani": _env_sayi("TESLIM_KAPI_HATA_ORANI", 0.02, float),
        "eslesmeyen_orani": _env_sayi("TESLIM_KAPI_ESLESMEYEN_ORANI", 0.05, float),
        "alan_degisikligi": _env_sayi("TESLIM_KAPI_ALAN_DEGISIKLIGI", 10_000, int),
    }


#: Sözleşme adı (G107 görev dosyası): `KAPI_ESIKLERI() -> dict`.
KAPI_ESIKLERI = kapi_esikleri


def _simdi() -> datetime:
    return datetime.now(timezone.utc)


def _kirp(mesaj: str) -> str:
    return mesaj if len(mesaj) <= HATA_MESAJI_SINIRI else mesaj[:HATA_MESAJI_SINIRI - 1] + "…"


def _guvenli_ad(dosya_adi: str) -> str:
    """Spool dosya adı: yol bileşeni ve yasak karakter içermez, 255'i aşmaz."""
    ad = Path(str(dosya_adi or "").replace("\\", "/")).name.strip()
    ad = _GUVENSIZ_AD.sub("_", ad).strip(". ")
    return (ad or "teslim.xlsx")[:255]


def _anahtar(deger: Any) -> str:
    """Etiket karşılaştırma anahtarı (hukdok_aktarim._baslik_anahtari'nın ikizi):
    TR büyük harf + aksan sadeleştirme + yalnız harf/rakam."""
    buyuk = tr_upper(str(deger or ""))
    ayrik = unicodedata.normalize("NFD", buyuk)
    sade = "".join(ch for ch in ayrik if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]", "", sade)


@contextmanager
def _oturum(db: Optional[Session]) -> Iterator[Session]:
    """`db` verildiyse onu kullan (kapatma), yoksa SessionLocal aç/kapat."""
    if db is not None:
        yield db
        return
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _teslim_getir(db: Session, teslim_id: int) -> models.AktarimTeslimi:
    teslim = db.get(models.AktarimTeslimi, teslim_id)
    if teslim is None:
        raise ValueError(f"Teslim yok: #{teslim_id}")
    return teslim


def _durum_kontrol(teslim: models.AktarimTeslimi, izinli: frozenset, islem: str) -> None:
    if teslim.durum not in izinli:
        raise ValueError(
            f"Teslim #{teslim.id} '{teslim.durum}' durumunda — {islem} yalnız "
            f"{', '.join(sorted(izinli))} durumundan yapılır"
        )


def _durum_gecir(teslim: models.AktarimTeslimi, yeni: str, *, not_: Optional[str] = None) -> None:
    """Durumu geçirir ve `durum_gecmisi`ne zaman damgalı kayıt ekler.

    Liste her seferinde YENİDEN atanır (`JSON` kolonu yerinde değişikliği
    izlemez; `append` sessizce kaybolurdu).
    """
    simdi = _simdi()
    gecmis = list(teslim.durum_gecmisi or [])
    gecmis.append({"durum": yeni, "at": simdi.isoformat(), "not": not_})
    teslim.durum_gecmisi = gecmis
    teslim.durum = yeni
    if yeni in NIHAI_DURUMLAR:
        teslim.done_at = simdi


def _spool_yolu(teslim: models.AktarimTeslimi) -> Optional[Path]:
    if not teslim.spool_path:
        return None
    yol = Path(teslim.spool_path)
    return yol if yol.is_file() else None


def _rapor_dizini(teslim: models.AktarimTeslimi) -> Path:
    dizin = Path(teslim.rapor_dizini) if teslim.rapor_dizini else (
        get_teslim_spool_dir() / f"{teslim.id}_raporlar"
    )
    dizin.mkdir(parents=True, exist_ok=True)
    return dizin


def _uygulandi_var(db: Session, dosya_adi: Optional[str] = None, *, haric: Optional[int] = None) -> bool:
    """Defterde `uygulandi` teslim var mı (ada göre ya da herhangi biri)."""
    sorgu = db.query(models.AktarimTeslimi.id).filter(
        models.AktarimTeslimi.durum == DURUM_UYGULANDI
    )
    if dosya_adi is not None:
        sorgu = sorgu.filter(models.AktarimTeslimi.dosya_adi == dosya_adi.strip())
    if haric is not None:
        sorgu = sorgu.filter(models.AktarimTeslimi.id != haric)
    return sorgu.first() is not None


def _sayaclari_yaz(teslim: models.AktarimTeslimi, sonuc: hukdok_aktarim.AktarimSonucu) -> None:
    teslim.okunan = sonuc.okunan
    teslim.islenen = sonuc.islenen
    teslim.atlanan = sonuc.atlanan
    teslim.hata_sayisi = len(sonuc.hatalar)
    teslim.alan_degisikligi = sonuc.alan_degisikligi
    teslim.kart_degisen = sonuc.kart_degisen
    teslim.envanter_denk = not sonuc.envanter_farki


def _ozet_yaz(yol: Path, sonuc: hukdok_aktarim.AktarimSonucu) -> None:
    """Koşu özetini rapor dizinine bırakır; yazılamazsa WARNING (rapor yan üründür)."""
    try:
        yol.write_text(hukdok_aktarim.ozet_metni(sonuc) + "\n", encoding="utf-8")
    except OSError as exc:
        logger.warning("Teslim özeti yazılamadı (%s): %s", yol, exc)


def _kapi_satiri(karar: Optional[str], gerekce: Optional[str]) -> str:
    """`ozet.txt`'nin son satırı: kapı kararı (+ gerekçe); karar yoksa "henüz değerlendirilmedi"."""
    metin = str(karar) if karar else _KAPI_HENUZ_YOK
    if gerekce:
        metin += f" — {gerekce}"
    return _KAPI_SATIRI_ONEKI + metin


def _ozet_dosyasi_yaz(rapor: Path, sonuc: hukdok_aktarim.AktarimSonucu,
                      karar: Optional[str], gerekce: Optional[str]) -> None:
    """`<rapor>/ozet.txt` = ozet_metni + kapı satırı (G110 cevap paketinin `ozet_<teslim>.txt`i)."""
    try:
        (rapor / OZET_DOSYASI).write_text(
            hukdok_aktarim.ozet_metni(sonuc) + "\n" + _kapi_satiri(karar, gerekce) + "\n",
            encoding="utf-8",
        )
    except OSError as exc:
        logger.warning("Teslim özeti yazılamadı (%s): %s", rapor / OZET_DOSYASI, exc)


def _ozet_kapi_guncelle(rapor_dizini: Optional[str], karar: Optional[str], gerekce: Optional[str]) -> None:
    """`ozet.txt`'deki kapı satırını tazeler (kapı kuru koşudan SONRA değerlendirilir); dosya yoksa sessiz."""
    if not rapor_dizini:
        return
    yol = Path(rapor_dizini) / OZET_DOSYASI
    try:
        if not yol.is_file():
            return
        satirlar = [s for s in yol.read_text(encoding="utf-8").splitlines()
                    if not s.startswith(_KAPI_SATIRI_ONEKI)]
        satirlar.append(_kapi_satiri(karar, gerekce))
        yol.write_text("\n".join(satirlar) + "\n", encoding="utf-8")
    except OSError as exc:
        logger.warning("Teslim özetinde kapı satırı güncellenemedi (%s): %s", yol, exc)


def _cevap_dene(teslim_id: int, *, db: Session) -> bool:
    """`teslim_cevap.cevap_yukle` — her istisna WARNING ile yutulur (cevap yan üründür, durum değişmez)."""
    from services import teslim_cevap

    try:
        return teslim_cevap.cevap_yukle(teslim_id, db=db)
    except Exception as exc:
        db.rollback()
        logger.warning("Teslim #%s cevap yüklemesi yapılamadı (%s): %s", teslim_id, type(exc).__name__, exc)
        return False


def _basarisiz(db: Session, teslim: models.AktarimTeslimi, mesaj: str, *, error_log: bool = True) -> str:
    """Nihai başarısızlık: durum + hata_mesaji + (varsayılan) TEK ERROR."""
    kirpik = _kirp(mesaj)
    teslim.hata_mesaji = kirpik
    _durum_gecir(teslim, DURUM_BASARISIZ, not_=kirpik)
    db.commit()
    if error_log:
        logger.error("Teslim #%s (%s) BAŞARISIZ: %s", teslim.id, teslim.dosya_adi, kirpik)
    return DURUM_BASARISIZ


def _reddet(db: Session, teslim: models.AktarimTeslimi, mesaj: str) -> str:
    kirpik = _kirp(mesaj)
    teslim.hata_mesaji = kirpik
    _durum_gecir(teslim, DURUM_REDDEDILDI, not_=kirpik)
    db.commit()
    logger.warning("Teslim #%s (%s) reddedildi: %s", teslim.id, teslim.dosya_adi, kirpik)
    return DURUM_REDDEDILDI


# ═══════════════════════════════════════════════════════════════════════════
# Bildirim (G108) — ADMIN_EMAILS'e uygulama içi, dedupe'lu, hatası yutulur
# ═══════════════════════════════════════════════════════════════════════════

def bildirim_dedupe_key(teslim_id: int, durum: str, email: str) -> str:
    """`teslim:<id>:<durum>:<alıcı>` — alıcı SONDA (G082 dersi: anahtar global tekil)."""
    from services.notifications import normalize_email

    return f"teslim:{teslim_id}:{durum}:{normalize_email(email)}"


def _bildirim_govdesi(teslim: models.AktarimTeslimi, olay: str) -> str:
    """Kapı gerekçesi (inceleme) / hata mesajı (red, başarısız) / sayaç özeti (uygulandı)."""
    if olay == DURUM_INCELEME and teslim.kapi_gerekcesi:
        return f"Kapı: {teslim.kapi_gerekcesi}"
    if olay in (DURUM_REDDEDILDI, DURUM_BASARISIZ) and teslim.hata_mesaji:
        return str(teslim.hata_mesaji)
    parcalar = [
        f"okunan {teslim.okunan if teslim.okunan is not None else '-'}",
        f"işlenen {teslim.islenen if teslim.islenen is not None else '-'}",
        f"atlanan {teslim.atlanan if teslim.atlanan is not None else '-'}",
        f"hata {teslim.hata_sayisi if teslim.hata_sayisi is not None else '-'}",
        f"alan değişikliği {teslim.alan_degisikligi if teslim.alan_degisikligi is not None else '-'}",
    ]
    if teslim.uygulayan:
        parcalar.append(f"uygulayan {teslim.uygulayan}")
    return ", ".join(parcalar)


def bildir(teslim_id: int, olay: str, *, db: Optional[Session] = None) -> List[int]:
    """Teslim geçişi için ADMIN_EMAILS kümesine bildirim yazar; yazılan/mevcut id'leri döner.

    Yalnız `BILDIRIM_OLAYLARI`ndaki geçişler bildirim üretir (diğerleri boş
    liste). Alıcı kümesi boşsa WARNING + boş liste. Her türlü hata (DB, kimlik,
    import) WARNING ile YUTULUR — bildirim yan üründür, durum makinesi ve
    çağıran uç bozulmaz. Aynı geçiş ikinci kez çağrılırsa dedupe mevcut satırı
    döner, yeni satır açmaz.
    """
    tanim = BILDIRIM_OLAYLARI.get(olay)
    if tanim is None:
        return []
    fiil, severity = tanim
    try:
        from routes.config import _admin_emails
        from services.notifications import create_notification

        alicilar = sorted(_admin_emails())
        if not alicilar:
            logger.warning("Teslim #%s bildirimi hedefsiz: ADMIN_EMAILS boş (olay=%s)", teslim_id, olay)
            return []
        with _oturum(db) as session:
            teslim = _teslim_getir(session, teslim_id)
            title = f"Veri teslimi {fiil}: {teslim.dosya_adi}"
            body = _bildirim_govdesi(teslim, olay)
            ids: List[int] = []
            for email in alicilar:
                try:
                    ids.append(create_notification(
                        session,
                        recipient_email=email,
                        type=BILDIRIM_TURU,
                        title=title,
                        body=body,
                        severity=severity,
                        dedupe_key=bildirim_dedupe_key(int(teslim.id), olay, email),
                    ))
                except Exception as exc:
                    session.rollback()
                    logger.warning(
                        "Teslim #%s bildirimi yazılamadı (alıcı=%s, olay=%s): %s",
                        teslim_id, email, olay, exc,
                    )
            logger.info("Teslim #%s bildirimi (%s): %s alıcı", teslim_id, olay, len(ids))
            return ids
    except Exception as exc:
        logger.warning("Teslim #%s bildirimi atlandı (olay=%s): %s", teslim_id, olay, exc)
        return []


# ═══════════════════════════════════════════════════════════════════════════
# Aktarım koşusu (ayrı bağlantı)
# ═══════════════════════════════════════════════════════════════════════════

def _engine_of(db: Session):
    bind = db.get_bind()
    return getattr(bind, "engine", bind) or _varsayilan_engine


def _timeout_sifirla(conn) -> None:
    """`aktarimi_kos`'un oturum boyu yükselttiği statement_timeout'u geri alır (yalnız PG)."""
    if conn.dialect.name != "postgresql":
        return
    try:
        conn.rollback()
        conn.execute(text("RESET statement_timeout"))
        conn.commit()
    except Exception as exc:
        logger.warning("statement_timeout sıfırlanamadı (bağlantı kapatılıyor): %s", exc)


def _aktarimi_calistir(db: Session, *, yol: Path, dosya_adi: str, rapor: Path,
                       dry_run: bool) -> hukdok_aktarim.AktarimSonucu:
    """`aktarimi_kos`'u kendine ait bağlantıda koşturur (modül şerhi: oturum modeli).

    ÇAĞIRAN defter oturumunu ÖNCE commit etmiş olmalıdır — sqlite StaticPool'da
    iki oturum aynı DBAPI bağlantısını paylaşır.
    """
    conn = _engine_of(db).connect()
    try:
        fabrika = sessionmaker(bind=conn, autocommit=False, autoflush=False)
        return hukdok_aktarim.aktarimi_kos(
            fabrika,
            girdi=yol,
            sheet=VERI_SAYFASI,
            dry_run=dry_run,
            source=f"{AKTARIM_SOURCE_PREFIX}_{dosya_adi}",
            rapor_dizini=rapor,
        )
    finally:
        _timeout_sifirla(conn)
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════
# Yapı doğrulama
# ═══════════════════════════════════════════════════════════════════════════

class _YapiHatasi(Exception):
    """Teslim paketi yapısal olarak kabul edilemez → `reddedildi`."""


def onceki_teslim_adi_oku(ws) -> Optional[str]:
    """`DEGISIKLIK_OZETI` sayfasından "Önceki teslim" dosya adını çıkarır.

    Etiket hücresi ("Önceki teslim", "Önceki teslim:" …) aksan/boşluk
    duyarsız aranır; değer aynı hücrede `:` sonrasında ya da satırın sonraki
    dolu hücresindedir. `· ` sonrası (satır/sütun sayısı) atılır. Yer tutucu
    ("—", "yok") ya da boş değer None döner.
    """
    for satir in ws.iter_rows(min_row=1, max_row=_OZET_TARAMA_SATIRI, values_only=True):
        hucreler = list(satir or ())
        for i, hucre in enumerate(hucreler):
            metin = str(hucre).strip() if hucre is not None else ""
            if not metin:
                continue
            etiket, ayrac, kalan = metin.partition(":")
            if not _anahtar(etiket).startswith(_ONCEKI_ETIKET):
                continue
            deger = kalan.strip() if ayrac else ""
            if not deger:
                deger = next(
                    (str(h).strip() for h in hucreler[i + 1:] if h is not None and str(h).strip()),
                    "",
                )
            ad = deger.split("·")[0].strip()
            if not ad or ad.upper() in _YER_TUTUCULAR:
                return None
            return ad
    return None


def _yapi_dogrula(yol: Path) -> Tuple[Optional[str], bool]:
    """(önceki teslim adı, özet sayfası var mı) — kabul edilemezse `_YapiHatasi`."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    except Exception as exc:
        raise _YapiHatasi(f"dosya açılamadı ({type(exc).__name__}): {exc}") from exc
    try:
        if VERI_SAYFASI not in wb.sheetnames:
            raise _YapiHatasi(
                f"'{VERI_SAYFASI}' sayfası yok (mevcut: {', '.join(wb.sheetnames) or '-'})"
            )
        ozet_var = OZET_SAYFASI in wb.sheetnames
        onceki = onceki_teslim_adi_oku(wb[OZET_SAYFASI]) if ozet_var else None
    finally:
        wb.close()

    try:
        _, bulunan = hukdok_aktarim.xlsx_oku(yol, sheet=VERI_SAYFASI, limit=0)
    except hukdok_aktarim.AktarimHatasi as exc:
        raise _YapiHatasi(str(exc)) from exc
    except Exception as exc:
        raise _YapiHatasi(f"başlıklar okunamadı ({type(exc).__name__}): {exc}") from exc
    eksik = [alan for alan in ZORUNLU_BASLIKLAR if alan not in bulunan]
    if eksik:
        raise _YapiHatasi(
            f"'{VERI_SAYFASI}' sayfasında zorunlu başlık(lar) yok: {', '.join(eksik)} "
            f"(bulunan: {', '.join(sorted(bulunan)) or '-'})"
        )
    return onceki, ozet_var


# ═══════════════════════════════════════════════════════════════════════════
# Durum makinesi
# ═══════════════════════════════════════════════════════════════════════════

def teslim_kaydet(*, icerik: bytes, dosya_adi: str, kaynak: str,
                  sharepoint_item_id: Optional[str] = None, db: Optional[Session] = None) -> int:
    """İçeriği deftere (ve yeni içerikse spool'a) yazar; yeni satırın id'sini döner.

    Aynı sha256 daha önce alınmışsa mevcut satıra DOKUNULMAZ: yeni satır
    `yinelenen` olarak açılır, notu ilk id'ye işaret eder, spool'a yazılmaz.
    """
    if kaynak not in KAYNAKLAR:
        raise ValueError(f"kaynak {kaynak!r} tanınmıyor (izinli: {', '.join(KAYNAKLAR)})")
    ad = _guvenli_ad(dosya_adi)
    ozet = hashlib.sha256(icerik).hexdigest()

    with _oturum(db) as session:
        mevcut = _ayni_icerik(session, ozet)
        if mevcut is not None:
            return _yinelenen_kaydet(session, ad, ozet, kaynak, sharepoint_item_id, mevcut)

        teslim = models.AktarimTeslimi(
            dosya_adi=ad, sha256=ozet, kaynak=kaynak,
            sharepoint_item_id=sharepoint_item_id, durum_gecmisi=[],
        )
        _durum_gecir(teslim, DURUM_ALINDI, not_=f"{kaynak} kaynağından alındı ({len(icerik)} bayt)")
        session.add(teslim)
        try:
            session.flush()
        except IntegrityError:
            # Yarış: aynı içerik aynı anda ikinci kez — kısmi UNIQUE index yakaladı.
            session.rollback()
            mevcut = _ayni_icerik(session, ozet)
            if mevcut is None:
                raise
            return _yinelenen_kaydet(session, ad, ozet, kaynak, sharepoint_item_id, mevcut)

        spool = get_teslim_spool_dir() / f"{teslim.id}_{ad}"
        spool.write_bytes(icerik)
        teslim.spool_path = str(spool)
        session.commit()
        logger.info("Teslim #%s alındı: %s (%s, %s bayt) → %s", teslim.id, ad, kaynak, len(icerik), spool)
        return int(teslim.id)


def _ayni_icerik(db: Session, ozet: str) -> Optional[models.AktarimTeslimi]:
    return (
        db.query(models.AktarimTeslimi)
        .filter(models.AktarimTeslimi.sha256 == ozet,
                models.AktarimTeslimi.durum != DURUM_YINELENEN)
        .order_by(models.AktarimTeslimi.id)
        .first()
    )


def _yinelenen_kaydet(db: Session, ad: str, ozet: str, kaynak: str,
                      sharepoint_item_id: Optional[str], mevcut: models.AktarimTeslimi) -> int:
    teslim = models.AktarimTeslimi(
        dosya_adi=ad, sha256=ozet, kaynak=kaynak,
        sharepoint_item_id=sharepoint_item_id, durum_gecmisi=[],
    )
    _durum_gecir(
        teslim, DURUM_YINELENEN,
        not_=f"aynı içerik daha önce alındı: teslim #{mevcut.id} ({mevcut.dosya_adi}, {mevcut.durum})",
    )
    db.add(teslim)
    db.commit()
    logger.info("Teslim #%s yinelenen: %s — ilk kayıt #%s", teslim.id, ad, mevcut.id)
    return int(teslim.id)


def teslim_dogrula(teslim_id: int, *, db: Optional[Session] = None) -> str:
    """Yapı doğrulaması + zincir kontrolü → 'dogrulandi' | 'reddedildi'."""
    with _oturum(db) as session:
        teslim = _teslim_getir(session, teslim_id)
        _durum_kontrol(teslim, ISLENEBILIR_DURUMLAR, "doğrulama")
        yol = _spool_yolu(teslim)
        if yol is None:
            return _reddet(session, teslim, f"teslim dosyası spool'da yok: {teslim.spool_path or '-'}")
        try:
            onceki, ozet_var = _yapi_dogrula(yol)
        except _YapiHatasi as exc:
            return _reddet(session, teslim, str(exc))

        teslim.onceki_teslim_adi = onceki
        if not ozet_var:
            teslim.zincir_tamam = None
            zincir_notu = f"'{OZET_SAYFASI}' sayfası yok — zincir bilinmiyor"
        else:
            teslim.zincir_tamam = bool(onceki) and _uygulandi_var(session, onceki, haric=int(teslim.id))
            zincir_notu = (
                f"önceki teslim {onceki!r} defterde uygulandı" if teslim.zincir_tamam
                else f"önceki teslim {onceki!r} defterde uygulanmış değil"
            )
        teslim.hata_mesaji = None
        _durum_gecir(teslim, DURUM_DOGRULANDI, not_=f"yapı tamam; {zincir_notu}")
        session.commit()
        logger.info("Teslim #%s doğrulandı: %s", teslim.id, zincir_notu)
        return DURUM_DOGRULANDI


def teslim_kuru_kos(teslim_id: int, *, db: Optional[Session] = None) -> str:
    """`aktarimi_kos(dry_run=True)`; sayaçlar deftere, raporlar spool'a → 'kuru_kosuldu' | 'basarisiz'."""
    with _oturum(db) as session:
        teslim = _teslim_getir(session, teslim_id)
        _durum_kontrol(teslim, _KURU_KOS_DURUMLARI, "kuru koşu")
        yol = _spool_yolu(teslim)
        if yol is None:
            return _basarisiz(session, teslim, f"teslim dosyası spool'da yok: {teslim.spool_path or '-'}")
        rapor = _rapor_dizini(teslim)
        teslim.rapor_dizini = str(rapor)
        dosya_adi = str(teslim.dosya_adi)
        session.commit()                     # defter transaction'ı KAPALI — aktarım kendi bağlantısında

        try:
            sonuc = _aktarimi_calistir(session, yol=yol, dosya_adi=dosya_adi, rapor=rapor, dry_run=True)
        except Exception as exc:
            teslim = _teslim_getir(session, teslim_id)
            return _basarisiz(session, teslim, f"kuru koşu istisnası — {type(exc).__name__}: {exc}")

        teslim = _teslim_getir(session, teslim_id)
        _sayaclari_yaz(teslim, sonuc)
        _ozet_yaz(rapor / "kuru-kosu-ozeti.txt", sonuc)
        _ozet_dosyasi_yaz(rapor, sonuc, None, None)      # kapı henüz yok; kapi_degerlendir tazeler
        _durum_gecir(teslim, DURUM_KURU_KOSULDU, not_=_sayac_notu(sonuc))
        session.commit()
        logger.info("Teslim #%s kuru koşuldu: %s", teslim.id, _sayac_notu(sonuc))
        return DURUM_KURU_KOSULDU


def _sayac_notu(sonuc: hukdok_aktarim.AktarimSonucu) -> str:
    return (
        f"okunan {sonuc.okunan}, işlenen {sonuc.islenen}, atlanan {sonuc.atlanan}, "
        f"hata {len(sonuc.hatalar)}, alan değişikliği {sonuc.alan_degisikligi} "
        f"({sonuc.kart_degisen} kart), envanter {'denk' if not sonuc.envanter_farki else 'DENK DEĞİL'}"
    )


def kapi_ihlalleri(db: Session, teslim: models.AktarimTeslimi, esikler: Optional[dict] = None) -> List[str]:
    """Plan §2.2 kurallarının HEPSİNİ değerlendirir; ihlal listesini döner (boş = otomatik)."""
    esik = esikler or kapi_esikleri()
    ihlaller: List[str] = []
    if teslim.envanter_denk is not True:
        ihlaller.append("envanter_denk_degil (belge envanteri kuru koşuda denk çıkmadı)")
    if not _uygulandi_var(db, haric=int(teslim.id)):
        ihlaller.append("ilk_teslim (defterde uygulanmış teslim yok)")
    if teslim.zincir_tamam is False:
        ihlaller.append(
            f"zincir_eksik (önceki teslim {teslim.onceki_teslim_adi or '?'!r} defterde uygulandı değil)"
        )
    okunan = teslim.okunan or 0
    if okunan == 0:
        ihlaller.append("bos_teslim (okunan satır 0)")
    else:
        hata_orani = (teslim.hata_sayisi or 0) / okunan
        if hata_orani > esik["hata_orani"]:
            ihlaller.append(f"hata_orani {hata_orani:.4f} > {esik['hata_orani']}")
        eslesmeyen_orani = (teslim.atlanan or 0) / okunan
        if eslesmeyen_orani > esik["eslesmeyen_orani"]:
            ihlaller.append(f"eslesmeyen_orani {eslesmeyen_orani:.4f} > {esik['eslesmeyen_orani']}")
    if (teslim.alan_degisikligi or 0) > esik["alan_degisikligi"]:
        ihlaller.append(f"alan_degisikligi {teslim.alan_degisikligi} > {esik['alan_degisikligi']}")
    return ihlaller


def kapi_degerlendir(teslim_id: int, *, db: Optional[Session] = None) -> str:
    """Kuru koşu sayaçlarını eşiklere vurur → 'otomatik' | 'inceleme' (defteri de yazar).

    'inceleme' durumu `inceleme_bekliyor`a geçirir; 'otomatik' durumu
    DEĞİŞTİRMEZ (`kuru_kosuldu` kalır, uygulama ayrı adımdır). `inceleme_bekliyor`
    satırında yeniden değerlendirme geriye gitmez — yalnız karar/gerekçe tazelenir.
    """
    with _oturum(db) as session:
        teslim = _teslim_getir(session, teslim_id)
        _durum_kontrol(teslim, _KAPI_DURUMLARI, "kapı değerlendirmesi")
        ihlaller = kapi_ihlalleri(session, teslim)
        karar = KAPI_INCELEME if ihlaller else KAPI_OTOMATIK
        gerekce = "; ".join(ihlaller) or None
        teslim.kapi_karari = karar
        teslim.kapi_gerekcesi = gerekce
        if ihlaller and teslim.durum != DURUM_INCELEME:
            _durum_gecir(teslim, DURUM_INCELEME, not_=gerekce)
        session.commit()
        _ozet_kapi_guncelle(str(teslim.rapor_dizini) if teslim.rapor_dizini else None, karar, gerekce)
        logger.info("Teslim #%s kapı: %s%s", teslim.id, karar, f" — {gerekce}" if gerekce else "")
        return karar


def teslim_uygula(teslim_id: int, *, uygulayan: str, db: Optional[Session] = None) -> str:
    """Gerçek yazım → 'uygulandi' | 'basarisiz' (+ admin bildirimi). Yalnız
    kuru_kosuldu / inceleme_bekliyor'dan."""
    with _oturum(db) as session:
        durum = _teslim_uygula(session, teslim_id, uygulayan=uygulayan)
        bildir(teslim_id, durum, db=session)
        if durum == DURUM_UYGULANDI:
            _cevap_dene(teslim_id, db=session)       # G110: bir deneme; kalanı gece turu
        return durum


def _teslim_uygula(session: Session, teslim_id: int, *, uygulayan: str) -> str:
    """`teslim_uygula` gövdesi — durum makinesi (bildirim dış sarmalayıcıda)."""
    teslim = _teslim_getir(session, teslim_id)
    _durum_kontrol(teslim, _UYGULA_DURUMLARI, "uygulama")
    yol = _spool_yolu(teslim)
    if yol is None:
        return _basarisiz(session, teslim, f"teslim dosyası spool'da yok: {teslim.spool_path or '-'}")
    rapor = _rapor_dizini(teslim)
    teslim.rapor_dizini = str(rapor)
    teslim.uygulayan = uygulayan
    dosya_adi = str(teslim.dosya_adi)
    _durum_gecir(teslim, DURUM_UYGULANIYOR, not_=f"uygulayan: {uygulayan}")
    session.commit()                     # çökme izi: açılışta acilis_toparla bunu görür

    try:
        sonuc = _aktarimi_calistir(session, yol=yol, dosya_adi=dosya_adi, rapor=rapor, dry_run=False)
    except Exception as exc:
        teslim = _teslim_getir(session, teslim_id)
        return _basarisiz(session, teslim, f"uygulama istisnası — {type(exc).__name__}: {exc}")

    teslim = _teslim_getir(session, teslim_id)
    _sayaclari_yaz(teslim, sonuc)
    _ozet_yaz(rapor / "uygulama-ozeti.txt", sonuc)
    _ozet_dosyasi_yaz(
        rapor, sonuc,
        str(teslim.kapi_karari) if teslim.kapi_karari else None,
        str(teslim.kapi_gerekcesi) if teslim.kapi_gerekcesi else None,
    )
    if sonuc.yazildi:
        _durum_gecir(teslim, DURUM_UYGULANDI, not_=_sayac_notu(sonuc))
        session.commit()
        logger.info("Teslim #%s uygulandı (%s): %s", teslim.id, uygulayan, _sayac_notu(sonuc))
        return DURUM_UYGULANDI
    if sonuc.envanter_farki:
        # ERROR'u aktarimi_kos zaten bastı ("Aktarım GERİ ALINDI") — ikincisi yazılmaz.
        return _basarisiz(
            session, teslim,
            "belge envanteri denk değil, koşu geri alındı: "
            + belge_envanteri.bicimle(sonuc.envanter_farki).replace("\n", " | "),
            error_log=False,
        )
    return _basarisiz(session, teslim, "aktarım commit etmedi (sebep bildirilmedi)")


def teslimi_isle(teslim_id: int, *, otomatik_uygula: bool, db: Optional[Session] = None) -> str:
    """doğrula → kuru koş → kapı → (otomatik_uygula ve kapı 'otomatik' ise) uygula; son durumu döner.

    İşlenebilir olmayan (nihai ya da `uygulaniyor`) satıra DOKUNMAZ, mevcut
    durumu döner ve bildirim ÜRETMEZ — G108 yükleme ucu `yinelenen` satırı da
    buradan geçirir. Otomatik uygulama `uygulayan="gece-job"` imzasıyla yapılır.
    Eşik dışı / nihai sonuçta admin bildirimi (`bildir`) buradan düşer;
    uygulama yolu kendi bildirimini `teslim_uygula` içinde verir.
    """
    with _oturum(db) as session:
        teslim = _teslim_getir(session, teslim_id)
        if teslim.durum not in ISLENEBILIR_DURUMLAR:
            logger.info("Teslim #%s '%s' durumunda — işlenmedi", teslim.id, teslim.durum)
            return str(teslim.durum)
        durum = teslim_dogrula(teslim_id, db=session)
        if durum != DURUM_DOGRULANDI:
            bildir(teslim_id, durum, db=session)
            return durum
        durum = teslim_kuru_kos(teslim_id, db=session)
        if durum != DURUM_KURU_KOSULDU:
            bildir(teslim_id, durum, db=session)
            return durum
        karar = kapi_degerlendir(teslim_id, db=session)
        if karar == KAPI_OTOMATIK and otomatik_uygula:
            return teslim_uygula(teslim_id, uygulayan=GECE_UYGULAYAN, db=session)
        durum = str(_teslim_getir(session, teslim_id).durum)
        if durum == DURUM_INCELEME:
            bildir(teslim_id, durum, db=session)
        return durum


def acilis_toparla(*, db: Optional[Session] = None) -> int:
    """Açılışta `uygulaniyor`da kalmış satırları `inceleme_bekliyor`a düşürür; sayısını döner.

    Aktarım TEK transaction'dır: kesilen uygulama ya tamamen yazdı ya hiç —
    hangisi olduğunu insan raporlardan/defterden görür ve yeniden "Uygula" der.
    """
    with _oturum(db) as session:
        kalanlar = (
            session.query(models.AktarimTeslimi)
            .filter(models.AktarimTeslimi.durum == DURUM_UYGULANIYOR)
            .order_by(models.AktarimTeslimi.id)
            .all()
        )
        for teslim in kalanlar:
            parcalar: List[str] = [str(teslim.kapi_gerekcesi)] if teslim.kapi_gerekcesi else []
            parcalar.append("uygulama_kesildi (açılışta 'uygulaniyor' bulundu)")
            teslim.kapi_karari = KAPI_INCELEME
            teslim.kapi_gerekcesi = "; ".join(parcalar)
            _durum_gecir(
                teslim, DURUM_INCELEME,
                not_="açılışta 'uygulaniyor' durumunda bulundu — kesilmiş uygulama, insan kararı gerek",
            )
            logger.warning(
                "Teslim #%s (%s) açılışta 'uygulaniyor' durumundaydı → inceleme_bekliyor",
                teslim.id, teslim.dosya_adi,
            )
        if kalanlar:
            session.commit()
        return len(kalanlar)


# ═══════════════════════════════════════════════════════════════════════════
# SharePoint gözcüsü + gece turu + boot telafisi (G109)
# ═══════════════════════════════════════════════════════════════════════════

def teslim_gelen_klasoru() -> str:
    """`<SHAREPOINT_FOLDER_TESLIM_NAME | 03_VERI_TESLIM>/gelen` — env çağrı anında okunur."""
    kok = os.getenv("SHAREPOINT_FOLDER_TESLIM_NAME", "").strip().strip("/") or TESLIM_KLASORU_VARSAYILAN
    return f"{kok}/{TESLIM_GELEN_ALT_KLASORU}"


def sharepoint_item_anahtari(item: dict) -> str:
    """`sharepoint_item_id` kolonuna yazılan ucuz-eleme anahtarı: `<driveItem id>@<eTag>`
    (eTag'in tırnakları atılır; modül şerhi: SharePoint gözcüsü)."""
    etag = str(item.get("eTag") or "").strip().strip('"')
    return f"{str(item.get('id') or '').strip()}@{etag}"


def _bilinen_sp_anahtarlari(db: Session) -> set:
    satirlar = (
        db.query(models.AktarimTeslimi.sharepoint_item_id)
        .filter(models.AktarimTeslimi.sharepoint_item_id.isnot(None))
        .all()
    )
    return {str(anahtar) for (anahtar,) in satirlar}


def sharepoint_tara(*, db: Optional[Session] = None) -> dict:
    """Gelen klasörünü listeler, yeni teslim dosyalarını indirip deftere yazar.

    Döner: `{"yeni", "yinelenen", "atlanan"}` — `yinelenen` hem ucuz elemeyi
    (aynı id+eTag defterde) hem indirme sonrası sha256 yinelenenini sayar;
    `atlanan` ad kalıbına uymayanlardır. Anahtar kapalıysa listelemez, sıfır
    döner. Tek dosyanın indirme/kayıt hatası WARNING'dir, tur sürer; LİSTELEME
    hatası yükselir (tur düzeyinde ele alınır — modül şerhi).
    """
    sayac = {"yeni": 0, "yinelenen": 0, "atlanan": 0}
    with _oturum(db) as session:
        if not app_settings.veri_teslim_otomasyonu_etkin(db=session):
            logger.info("SharePoint teslim taraması atlandı: veri_teslim_otomasyonu kapalı")
            return sayac
        klasor = teslim_gelen_klasoru()
        dosyalar = _spu.list_folder_children(klasor)
        bilinen = _bilinen_sp_anahtarlari(session)
        for item in dosyalar:
            ad = str(item.get("name") or "").strip()
            if not TESLIM_AD_KALIBI.match(ad):
                sayac["atlanan"] += 1
                logger.info("SharePoint teslim klasöründe kalıp dışı dosya atlandı: %s", ad)
                continue
            anahtar = sharepoint_item_anahtari(item)
            if anahtar in bilinen:
                sayac["yinelenen"] += 1
                continue
            try:
                icerik, _ctype = _spu.download_file_from_sharepoint(klasor, ad)
                teslim_id = teslim_kaydet(
                    icerik=icerik, dosya_adi=ad, kaynak="sharepoint",
                    sharepoint_item_id=anahtar, db=session,
                )
            except Exception as exc:
                session.rollback()
                logger.warning("SharePoint teslim dosyası alınamadı (%s/%s): %s", klasor, ad, exc)
                continue
            bilinen.add(anahtar)
            if _teslim_getir(session, teslim_id).durum == DURUM_YINELENEN:
                sayac["yinelenen"] += 1
            else:
                sayac["yeni"] += 1
        logger.info(
            "SharePoint teslim taraması (%s): %s dosya — yeni %s, yinelenen %s, atlanan %s",
            klasor, len(dosyalar), sayac["yeni"], sayac["yinelenen"], sayac["atlanan"],
        )
        return sayac


def _bekleyen_idler(db: Session, durumlar: Tuple[str, ...]) -> List[int]:
    satirlar = (
        db.query(models.AktarimTeslimi.id)
        .filter(models.AktarimTeslimi.durum.in_(durumlar))
        .order_by(models.AktarimTeslimi.created_at, models.AktarimTeslimi.id)
        .all()
    )
    return [int(tid) for (tid,) in satirlar]


def _tek_uygulama_incelemeye(session: Session, teslim_id: int, uygulanan_id: int) -> str:
    """Aynı turda ikinci uygulanabilir teslim: kapı "otomatik" dese de insana bırakılır."""
    teslim = _teslim_getir(session, teslim_id)
    parcalar: List[str] = [str(teslim.kapi_gerekcesi)] if teslim.kapi_gerekcesi else []
    parcalar.append(
        f"{KAPI_TEK_UYGULAMA} (aynı gece turunda teslim #{uygulanan_id} uygulandı — "
        "ikincisi insan kararına bırakıldı)"
    )
    teslim.kapi_karari = KAPI_INCELEME
    teslim.kapi_gerekcesi = "; ".join(parcalar)
    _durum_gecir(teslim, DURUM_INCELEME, not_=parcalar[-1])
    session.commit()
    logger.info("Teslim #%s gece turunda incelemeye alındı: %s", teslim.id, parcalar[-1])
    bildir(teslim_id, DURUM_INCELEME, db=session)
    return DURUM_INCELEME


def _bekleyenleri_isle(session: Session, durumlar: Tuple[str, ...], *,
                       otomatik_uygula: bool) -> Tuple[dict, Optional[int]]:
    """Bekleyenleri `created_at` sırasıyla işler; (id → son durum, uygulanan id) döner.

    `otomatik_uygula=True` iken bile en fazla BİR teslim uygulanır; sonrakiler
    kuru koşuda kalırsa `tek_uygulama` gerekçesiyle incelemeye alınır. Satır
    düzeyi beklenmedik istisna WARNING'dir (nihai `basarisiz` ERROR'unu
    `teslimi_isle` zaten basar), tur sürer.
    """
    sonuc: dict = {}
    uygulanan_id: Optional[int] = None
    for teslim_id in _bekleyen_idler(session, durumlar):
        try:
            durum = teslimi_isle(
                teslim_id, otomatik_uygula=otomatik_uygula and uygulanan_id is None, db=session,
            )
            if durum == DURUM_UYGULANDI:
                uygulanan_id = teslim_id
            elif otomatik_uygula and uygulanan_id is not None and durum == DURUM_KURU_KOSULDU:
                durum = _tek_uygulama_incelemeye(session, teslim_id, uygulanan_id)
        except Exception as exc:
            session.rollback()
            logger.warning("Teslim #%s turda işlenemedi (%s): %s", teslim_id, type(exc).__name__, exc)
            durum = "hata"
        sonuc[teslim_id] = durum
    return sonuc, uygulanan_id


def gece_turu() -> dict:
    """04:00 TR gece turu (lider worker, APScheduler `veri_teslim` job'ı) — modül şerhi.

    Döner: `{"etkin", "toparlanan", "tara", "durumlar", "uygulanan"}`. Tarama
    başarısızlığı tur başına TEK ERROR'dur ve bekleyenlerin işlenmesini
    engellemez (dün indirilen paket bugün yine uygulanabilir).
    """
    ozet: dict = {"etkin": False, "toparlanan": 0, "tara": None, "durumlar": {}, "uygulanan": None}
    with _oturum(None) as session:
        if not app_settings.veri_teslim_otomasyonu_etkin(db=session):
            logger.info("Gece veri teslim turu atlandı: veri_teslim_otomasyonu kapalı")
            return ozet
        ozet["etkin"] = True
        ozet["toparlanan"] = acilis_toparla(db=session)
        try:
            ozet["tara"] = sharepoint_tara(db=session)
        except Exception as exc:
            session.rollback()
            logger.error("Gece veri teslim turu: SharePoint taraması başarısız — %s: %s",
                         type(exc).__name__, exc)
        ozet["durumlar"], ozet["uygulanan"] = _bekleyenleri_isle(
            session, GECE_ISLENEN_DURUMLAR, otomatik_uygula=True,
        )
        _bekleyen_cevaplari_yukle(session, haric=ozet["uygulanan"])
    logger.info("Gece veri teslim turu bitti: %s", ozet)
    return ozet


def _bekleyen_cevaplari_yukle(session: Session, *, haric: Optional[int]) -> dict:
    """G110: `uygulandi` + `cevap_yuklendi=False` teslimlerin cevap paketini yeniden dener.

    `haric` bu turda uygulanan teslimdir — `teslim_uygula` az önce denedi, aynı gece
    ikinci deneme boşuna; retry ertesi geceye kalır. Sonuç `ozet`e EKLENMEZ (dönüş
    şekli G109 sözleşmesi), INFO satırıyla loglanır.
    """
    from services import teslim_cevap

    sonuc = teslim_cevap.bekleyen_cevaplari_yukle(session, haric=haric)
    if sonuc:
        logger.info(
            "Gece veri teslim turu: cevap paketi %s teslimde denendi — yüklenen %s: %s",
            len(sonuc), sum(1 for v in sonuc.values() if v), sonuc,
        )
    return sonuc


def boot_catch_up() -> Optional[dict]:
    """Lider boot'unda bir kerelik telafi (modül şerhi: Boot telafisi) — UYGULAMA YOK.

    `api.py` daemon thread'de çağırır; her istisna burada yutulur, TEK WARNING
    loglanır (thread'den taşan istisna kimseye ulaşmaz; 04:00 cron'u asıl iştir).
    """
    ozet: dict = {"etkin": False, "toparlanan": 0, "tara": None, "durumlar": {}}
    try:
        with _oturum(None) as session:
            ozet["toparlanan"] = acilis_toparla(db=session)
            if not app_settings.veri_teslim_otomasyonu_etkin(db=session):
                logger.info("Veri teslim boot telafisi: anahtar kapalı — yalnız açılış toparlaması yapıldı")
                return ozet
            ozet["etkin"] = True
            ozet["tara"] = sharepoint_tara(db=session)
            ozet["durumlar"], _ = _bekleyenleri_isle(session, BOOT_ISLENEN_DURUMLAR, otomatik_uygula=False)
        logger.info("Veri teslim boot telafisi bitti: %s", ozet)
        return ozet
    except Exception as exc:
        logger.warning("Veri teslim boot telafisi yapılamadı (04:00 turu yeniden dener): %s", exc)
        return None

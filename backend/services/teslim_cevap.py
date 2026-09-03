"""Veri teslim cevap paketi — eşleşme CSV'si + raporların SharePoint'e geri yüklenmesi (G110).

Plan `docs/plan/veri-teslim-otomasyonu-plani-2026-09-03.md` §2.4: veri ekibine FAZ F §3
ve teslim paketinin `HUKDOK_TALEPLERI` #9'unda borçlu olduğumuz çıktı ("hangi SistemNo
hangi `cases.id`/`tracking_no` ile eşleşti, hangileri eşleşmedi") her `uygulandi` teslim
için kendiliğinden `<SHAREPOINT_FOLDER_TESLIM_NAME>/cevap/<teslim>/` altına gider.
Durum makinesi `teslim_kutusu`'ndadır; burası yalnız onun `uygulandi` satırı üzerinde
çalışır ve HİÇBİR durum geçişi yapmaz.

Cevap klasörünün içeriği:

* `eslesme_<teslim>.csv` — teslimin `Sheet` sayfasındaki her satır için `case_foys`
  üzerinden `cases.id` / `tracking_no` / `klasor_no_2` / `tku_no` / `case_party_id`;
  eşleşmeyenlerde `case_id` boş, `sebep` satır raporundaki ATLANDI/HATA sebebi
  (`ESLESME_BASLIKLARI`). Dosya rapor dizinine yazılır (spool'da kalıcı, admin rapor
  uçları da listeler) ve oradan yüklenir.
* `ozet_<teslim>.txt` — `teslim_kutusu`'nun rapor dizinine bıraktığı `ozet.txt`
  (`ozet_metni(sonuc)` + kapı kararı satırı) yüklenirken bu adı alır.
* `deger-havuzu-farki_<teslim>.csv` — teslimin `DEGER_HAVUZLARI` sayfasındaki kapalı liste
  değerleri ile bizim referans listelerimizin farkı, İKİ yönlü (`teslimde var / bizde yok`
  + `bizde var / teslimde yok`; `HAVUZ_FARKI_BASLIKLARI`). Fark YOKSA dosya üretilmez (varsa
  bayat kopya silinir). Üretim `teslim_kutusu` kuru koşu/uygulama adımındadır
  (`havuz_farki_csv_uret`, G112); fark varsa admin bildirimi de oradan düşer. Referans
  listesine YAZMA YOK — tahmin yasağı, `alleged_faults` seed'lenmez kararı korunur.
* rapor dizinindeki diğer CSV/TXT'ler (`satir-raporu_*.csv`, `kardes-foy-celiskileri_*.csv`,
  `kuru-kosu-ozeti.txt`, `uygulama-ozeti.txt`) kendi adlarıyla.

Tasarım kararları
-----------------
* **Yükleme SharePoint alışverişinin ikinci yönüdür; iki kapısı vardır.** (1) Otomasyon
  anahtarı (`veri_teslim_otomasyonu`) kapalıysa SharePoint'e ne bakılır (G109 gözcüsü)
  ne yazılır — elle "Uygula" yine çalışır, yalnız cevap dosyaları spool'da kalır ve
  anahtar açılınca ertesi gece turu yükler. (2) Yazma hedefinin kökü
  `SHAREPOINT_FOLDER_TESLIM_NAME` env'inden AÇIKÇA gelir; okuma tarafındaki
  `03_VERI_TESLIM` varsayılanı yazma için türetilmez (env yoksa INFO + atlanır, defter
  değişmez). Gerekçe: cevap dosyaları ortak arşive YAZILIR; kurulumu yapılmamış bir hedefe
  varsayılanla yazmak istemiyoruz. Aynı iki kapı, gerçek Graph kimlik bilgisi taşıyan
  konteynerde koşan mevcut G107/G108/G109 testlerinin (upload'ı sahtelemeden `uygulandi`ya
  ulaşırlar) prod SharePoint'e dosya bırakmasını da engeller — bu görevin test dosyası
  env'i açıkça kurar ve `upload_file_to_sharepoint`'i sahteler.
* **Kısmi başarısızlık teslimi `basarisiz` YAPMAZ** (plan §2.4): yazım zaten commit'li.
  Dosya başına WARNING, tur sonunda TEK özet WARNING; `cevap_yuklendi=False` kalır, her
  deneme `durum_gecmisi`ne (durum değişmeden) "cevap yükleme denemesi #N" notu düşer;
  ertesi gece turu (`teslim_kutusu.gece_turu`) yeniden dener. ERROR yazılmaz.
* **Klasör yaratma çağrısı yoktur**: `upload_file_to_sharepoint` küçük dosyada
  `PUT /drives/{drive}/root:/{klasör}/{ad}:/content` atar (`_upload_with_token`);
  `cevap/<teslim>/` ara klasörlerinin açılması Graph'ın yol-adresli PUT davranışına
  dayanır (Graph belgesi: eksik üst klasörler oluşturulur). Kod tabanında bu davranışa
  yaslanan başka bir çağrı YOK (mevcut yüklemeler düz klasöre gider;
  `use_date_subfolder` yolu hiçbir çağıranda açık değil) — gerçek SharePoint'te ilk
  cevap yüklemesi insan gözüyle doğrulanmalı (görev raporu "İzlenecekler").
* **CSV biçimi `hukdok_aktarim._csv_yaz` ile aynı** (UTF-8 BOM + `;`): fonksiyon private
  olduğu için deseni kopyalandı (`_csv_yaz`), byte eşitliği testle kilitli.
* **Sebep satır numarasıyla eşlenir** (`satir_no`), SistemNo ile değil: aynı SistemNo
  dosyada iki kez geçebilir. Uygulama koşusunun satır raporu rapor dizinindeki EN YENİ
  `satir-raporu_*.csv`'dir (ad zaman damgalı, sözlük sırası = zaman sırası).
"""
from __future__ import annotations

import csv
import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from sqlalchemy.orm import Session

from managers import reference_lists
import models
from scripts import hukdok_aktarim
from services import app_settings
from services import teslim_kutusu as tk
from sharepoint import sharepoint_uploader_graph as _spu

logger = logging.getLogger(__name__)

#: `<SHAREPOINT_FOLDER_TESLIM_NAME>/cevap/<teslim>` — cevap alt klasörü.
CEVAP_ALT_KLASORU = "cevap"
#: Eşleşme dosyasının sütunları (görev sözleşmesi, sıra sabit).
ESLESME_BASLIKLARI: Tuple[str, ...] = (
    "sistem_no", "dosya_no", "case_id", "tracking_no", "klasor_no_2",
    "tku_no", "case_party_id", "durum", "sebep",
)
DURUM_ESLESTI = "ESLESTI"
DURUM_ESLESMEDI = "ESLESMEDI"
#: Yüklenen dosya türleri → Graph content-type.
CEVAP_TURLERI: Dict[str, str] = {".csv": "text/csv", ".txt": "text/plain"}
#: `durum_gecmisi` notunun öneki — deneme sayısı buradan sayılır.
DENEME_NOTU_ONEKI = "cevap yükleme denemesi #"
_SATIR_RAPORU_KALIBI = "satir-raporu_*.csv"
_IN_PARCASI = 500

# ─── DEGER_HAVUZLARI (G112) ──────────────────────────────────────────────────
#: Teslim paketinin kapalı liste değerleri sayfası.
HAVUZ_SAYFASI = "DEGER_HAVUZLARI"
#: Havuz adındaki parantezli EK: gerçek paket (HUKDOK_TESLIM_PAKETI_2026-08-18,
#: `DEGER_HAVUZLARI` 3. satırdan itibaren) havuzu "İddia Edilen Kusur (kapalı, 7)",
#: "Uygulanan Yöntem (kadın doğum)" gibi yazar. Ek, etiket temizliğiyle atılır
#: (değer tahmini değil); sözlük anahtarı ve fark satırındaki havuz adı sayfadaki
#: yazımı KORUR.
_HAVUZ_EKI = re.compile(r"\([^)]*\)")


def _havuz_anahtari(ad: Any) -> str:
    """Havuz adı → eşleme anahtarı: parantezli ek atılır, sonra `tk._anahtar`
    ("İddia Edilen Kusur (kapalı, 7)" = "İDDİA EDİLEN KUSUR")."""
    return tk._anahtar(_HAVUZ_EKI.sub(" ", str(ad or "")))


#: "Havuz / Sütun" (başlık anahtarı) → `reference_lists.LIST_REGISTRY` anahtarı.
#: Sabit KÜÇÜK sözlük (görev sözleşmesi); eşlemesi olmayan havuz atlanır.
#: "Temyiz Onama Durumu": sözleşmedeki "Yargıtay Onama Durumu" havuzunun gerçek
#: paketteki adı (aynı liste — temyiz mercii Yargıtay); ikisi de tanınır.
HAVUZ_LISTE_ESLEMESI: Dict[str, str] = {
    _havuz_anahtari("İddia Edilen Kusur"):         "alleged_faults",
    _havuz_anahtari("İstinaf Karar Durumu"):       "appeal_decisions",
    _havuz_anahtari("Yargıtay Onama Durumu"):      "cassation_decisions",
    _havuz_anahtari("Temyiz Onama Durumu"):        "cassation_decisions",
    _havuz_anahtari("Yerel Mahkeme Karar Durumu"): "local_decisions",
    _havuz_anahtari("Olay Türü"):                  "event_types",
    _havuz_anahtari("Hükümdeki Rol"):              "judgment_roles",
}
#: Uzun biçim başlık adayları (havuz adı sütunu + değer sütunu). Uzun biçim
#: bulunamazsa GENİŞ biçim denenir: eşlemesi olan her başlık bir havuzdur,
#: değerler o sütunda aşağı iner (sayfa düzeni karşı tarafla yazılı
#: sabitlenmedi — G114 sözleşmesi).
HAVUZ_SUTUNLARI: Dict[str, Tuple[str, ...]] = {
    "havuz": ("Havuz / Sütun", "Havuz/Sütun", "Havuz", "Sütun"),
    "deger": ("Değer", "Değerler", "Havuz Değeri", "Havuz Değerleri", "İzinli Değerler"),
}
#: Başlık satırı ilk satır OLMAYABİLİR: gerçek paket 1. satırda "RESMİ DEĞER
#: HAVUZLARI" sayfa başlığı, 2. satır boş, 3. satırda ("Havuz / Sütun", "Değer")
#: taşır. Başlık ilk bu kadar satırda aranır; bulunamazsa sayfa havuzsuz sayılır.
HAVUZ_BASLIK_TARAMA = 10
#: Fark CSV'sinin sütunları (sıra sabit).
HAVUZ_FARKI_BASLIKLARI: Tuple[str, ...] = ("havuz", "liste", "yon", "deger")
YON_TESLIMDE_VAR = "teslimde var / bizde yok"
YON_BIZDE_VAR = "bizde var / teslimde yok"
#: Bir hücrede birden çok değer: `;`, satır sonu ya da `|` ile (değerlerin kendisi
#: `/` içerir — "Red/Esastan" — o yüzden `/` ayraç DEĞİLDİR).
_HAVUZ_AYRAC = re.compile(r"[;\r\n|]+")


# ═══════════════════════════════════════════════════════════════════════════
# Yardımcılar
# ═══════════════════════════════════════════════════════════════════════════

def teslim_adi_uzantisiz(dosya_adi: str) -> str:
    """`HUKDOK_TESLIM_X.xlsx` → `HUKDOK_TESLIM_X` (cevap klasörü ve dosya adlarının gövdesi)."""
    return Path(str(dosya_adi or "")).stem or "teslim"


def cevap_klasoru(dosya_adi: str) -> Optional[str]:
    """`<SHAREPOINT_FOLDER_TESLIM_NAME>/cevap/<teslim>`; env tanımsızsa None (modül şerhi)."""
    kok = os.getenv("SHAREPOINT_FOLDER_TESLIM_NAME", "").strip().strip("/")
    if not kok:
        return None
    return f"{kok}/{CEVAP_ALT_KLASORU}/{teslim_adi_uzantisiz(dosya_adi)}"


def _metin(deger: Any) -> Optional[str]:
    """Boşluk-normalize metin (hukdok_aktarim._metin / foy_map anahtarı ile aynı kural)."""
    if deger is None:
        return None
    metin = " ".join(str(deger).split())
    return metin or None


def _csv_yaz(yol: Path, basliklar: Sequence[str], satirlar: Iterable[Sequence[Any]]) -> Path:
    """`hukdok_aktarim._csv_yaz` deseni: UTF-8 BOM + `;` (Türkçe Excel doğrudan açar)."""
    with open(yol, "w", newline="", encoding="utf-8-sig") as dosya:
        yazici = csv.writer(dosya, delimiter=";")
        yazici.writerow(basliklar)
        yazici.writerows(satirlar)
    return yol


def _son_satir_raporu(rapor_dizini: Path) -> Optional[Path]:
    adaylar = sorted(p for p in rapor_dizini.glob(_SATIR_RAPORU_KALIBI) if p.is_file())
    return adaylar[-1] if adaylar else None


def _sebepleri_oku(rapor_dizini: Path) -> Dict[int, List[str]]:
    """En yeni satır raporundan {satir_no: [sebep, ...]} (okunamazsa WARNING + boş)."""
    yol = _son_satir_raporu(rapor_dizini)
    if yol is None:
        return {}
    sebepler: Dict[int, List[str]] = {}
    try:
        with open(yol, "r", newline="", encoding="utf-8-sig") as dosya:
            for kayit in csv.DictReader(dosya, delimiter=";"):
                try:
                    satir_no = int(str(kayit.get("satir_no") or "").strip())
                except ValueError:
                    continue
                sebep = str(kayit.get("sebep") or "").strip()
                if sebep:
                    sebepler.setdefault(satir_no, []).append(sebep)
    except (OSError, csv.Error) as exc:
        logger.warning("Satır raporu okunamadı (%s): %s — sebep sütunu boş kalacak", yol, exc)
        return {}
    return sebepler


def _foy_kayitlari(db: Session, sistem_nolar: Sequence[str]) -> Dict[str, Tuple[Any, ...]]:
    """{sistem_no: (case_id, tracking_no, klasor_no_2, tku_no, case_party_id)} — parçalı IN."""
    sonuc: Dict[str, Tuple[Any, ...]] = {}
    anahtarlar = sorted({s for s in sistem_nolar if s})
    for i in range(0, len(anahtarlar), _IN_PARCASI):
        parca = anahtarlar[i:i + _IN_PARCASI]
        satirlar = (
            db.query(
                models.CaseFoy.sistem_no, models.CaseFoy.case_id, models.Case.tracking_no,
                models.Case.klasor_no_2, models.CaseFoy.tku_no, models.CaseFoy.case_party_id,
            )
            .join(models.Case, models.Case.id == models.CaseFoy.case_id)
            .filter(models.CaseFoy.sistem_no.in_(parca))
            .all()
        )
        for sistem_no, *kalan in satirlar:
            sonuc[str(sistem_no)] = tuple(kalan)
    return sonuc


def _deneme_sayisi(teslim: models.AktarimTeslimi) -> int:
    gecmis: List[Any] = list(teslim.durum_gecmisi or [])
    return sum(1 for g in gecmis if str((g or {}).get("not") or "").startswith(DENEME_NOTU_ONEKI))


def _gecmis_notu_ekle(teslim: models.AktarimTeslimi, not_: str) -> None:
    """Durum DEĞİŞMEDEN `durum_gecmisi`ne not düşer (liste yeniden atanır — JSON kolonu)."""
    gecmis = list(teslim.durum_gecmisi or [])
    gecmis.append({"durum": teslim.durum, "at": tk._simdi().isoformat(), "not": not_})
    teslim.durum_gecmisi = gecmis


# ═══════════════════════════════════════════════════════════════════════════
# Eşleşme dosyası (Talep #9)
# ═══════════════════════════════════════════════════════════════════════════

def eslesme_csv_uret(teslim_id: int, hedef: Path, *, db: Optional[Session] = None) -> Path:
    """Teslimin `Sheet` satırları için SistemNo → kart eşleşme CSV'sini `hedef`e yazar.

    Her sayfa satırı bir CSV satırıdır (aynı SistemNo iki kez geçerse iki satır).
    Eşleşen: `case_foys` + `cases` kolonları, `durum=ESLESTI`. Eşleşmeyen: `case_id`
    boş, `durum=ESLESMEDI`, `sebep` = satır raporundaki kayıt (yoksa "satır raporunda
    kayıt yok"). Eşleşen satırın alan-düzeyi HATA notu varsa (`X yazılmadı: …`) o da
    `sebep`e düşer — veri ekibi düzeltme listesini buradan kurar.
    """
    hedef = Path(hedef)
    with tk._oturum(db) as session:
        teslim = tk._teslim_getir(session, teslim_id)
        yol = tk._spool_yolu(teslim)
        if yol is None:
            raise ValueError(f"Teslim #{teslim.id} dosyası spool'da yok: {teslim.spool_path or '-'}")
        satirlar, _ = hukdok_aktarim.xlsx_oku(yol, sheet=tk.VERI_SAYFASI)
        sebepler = _sebepleri_oku(Path(teslim.rapor_dizini)) if teslim.rapor_dizini else {}
        foyler = _foy_kayitlari(
            session, [_metin(s.degerler.get("sistem_no")) or "" for s in satirlar],
        )

    kayitlar: List[Tuple[Any, ...]] = []
    eslesen = 0
    for satir in satirlar:
        sistem_no = _metin(satir.degerler.get("sistem_no")) or ""
        dosya_no = _metin(satir.degerler.get("dosya_no")) or ""
        sebep = " | ".join(sebepler.get(satir.satir_no, []))
        foy = foyler.get(sistem_no)
        if foy is None:
            kayitlar.append((
                sistem_no, dosya_no, "", "", "", "", "",
                DURUM_ESLESMEDI, sebep or "satır raporunda kayıt yok",
            ))
            continue
        case_id, tracking_no, klasor_no_2, tku_no, case_party_id = foy
        eslesen += 1
        kayitlar.append((
            sistem_no, dosya_no, case_id, tracking_no or "", klasor_no_2 or "",
            tku_no or "", case_party_id if case_party_id is not None else "",
            DURUM_ESLESTI, sebep,
        ))

    hedef.parent.mkdir(parents=True, exist_ok=True)
    _csv_yaz(hedef, ESLESME_BASLIKLARI, kayitlar)
    logger.info(
        "Teslim #%s eşleşme dosyası yazıldı: %s (%s satır, %s eşleşen, %s eşleşmeyen)",
        teslim_id, hedef, len(kayitlar), eslesen, len(kayitlar) - eslesen,
    )
    return hedef


# ═══════════════════════════════════════════════════════════════════════════
# DEGER_HAVUZLARI fark raporu (G112) — yalnız RAPOR, listeye yazma YOK
# ═══════════════════════════════════════════════════════════════════════════

def _havuz_parcalari(deger: Any) -> List[str]:
    """Hücreyi değerlere böler (boşluk-normalize, mükerrersiz, boşlar atılır)."""
    parcalar: List[str] = []
    for ham in _HAVUZ_AYRAC.split(str(deger) if deger is not None else ""):
        metin = _metin(ham)
        if metin and metin not in parcalar:
            parcalar.append(metin)
    return parcalar


def _havuz_basligi_bul(satirlar: Sequence[Sequence[Any]]) -> Tuple[Optional[int], Dict[str, int]]:
    """İlk `HAVUZ_BASLIK_TARAMA` satırda başlık satırını bulur → (satır indeksi, uzun
    biçim sütun indeksleri). Önce UZUN biçim (havuz + değer başlığı aynı satırda);
    yoksa GENİŞ biçim (eşlemesi olan en az bir başlık taşıyan ilk satır; indeksler
    boş). Hiçbiri yoksa (None, {}). Sayfa başlığı ("RESMİ DEĞER HAVUZLARI") ve boş
    satırlar hiçbir adaya uymadığı için doğal olarak atlanır."""
    genis: Optional[int] = None
    for si, satir in enumerate(satirlar[:HAVUZ_BASLIK_TARAMA]):
        dosyadaki: Dict[str, int] = {}
        for i, ham in enumerate(satir):
            anahtar = tk._anahtar(ham)
            if anahtar and anahtar not in dosyadaki:
                dosyadaki[anahtar] = i
        indeksler: Dict[str, int] = {}
        for alan, adaylar in HAVUZ_SUTUNLARI.items():
            for aday in adaylar:
                if tk._anahtar(aday) in dosyadaki:
                    indeksler[alan] = dosyadaki[tk._anahtar(aday)]
                    break
        if "havuz" in indeksler and "deger" in indeksler:
            return si, indeksler
        if genis is None and any(_havuz_anahtari(h) in HAVUZ_LISTE_ESLEMESI for h in satir if h is not None):
            genis = si
    return genis, {}


def havuz_degerlerini_oku(yol: Path, *, sheet: str = HAVUZ_SAYFASI) -> Dict[str, List[str]]:
    """`DEGER_HAVUZLARI` → {havuz adı (sayfadaki yazım): [değer, ...]}; sayfa yoksa {}.

    Başlık satırı ilk `HAVUZ_BASLIK_TARAMA` satırda ARANIR (gerçek paket 3. satırda
    taşır — `_havuz_basligi_bul`). Uzun biçim ("Havuz / Sütun" + "Değer" sütunları,
    satır başına bir değer ya da `;` ile birleşik) önce denenir; yoksa geniş biçim
    (eşlemesi olan her başlık bir havuz, değerler aşağı iner). İkisi de havuz
    vermiyorsa INFO + {} (hata değil — sayfa isteğe bağlıdır). Eşleme
    `_havuz_anahtari` ile (parantezli ek atılır); eşlemesi olmayan havuzlar sözlüğe
    GİRMEZ, girenler sayfadaki yazımıyla ("İddia Edilen Kusur (kapalı, 7)") anahtarlanır.
    """
    import openpyxl

    wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            logger.info("%r sayfası yok — değer havuzu farkı bakılmadı", sheet)
            return {}
        satirlar = [list(s or ()) for s in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    baslik_i, indeksler = _havuz_basligi_bul(satirlar)

    havuzlar: Dict[str, List[str]] = {}
    if baslik_i is not None and "havuz" in indeksler and "deger" in indeksler:
        hi, di = indeksler["havuz"], indeksler["deger"]
        for ham in satirlar[baslik_i + 1:]:
            havuz = _metin(ham[hi]) if hi < len(ham) else None
            if not havuz or _havuz_anahtari(havuz) not in HAVUZ_LISTE_ESLEMESI:
                continue
            liste = havuzlar.setdefault(havuz, [])
            for deger in _havuz_parcalari(ham[di] if di < len(ham) else None):
                if deger not in liste:
                    liste.append(deger)
    elif baslik_i is not None:
        for i, ham_baslik in enumerate(satirlar[baslik_i]):
            havuz = _metin(ham_baslik)
            if not havuz or _havuz_anahtari(havuz) not in HAVUZ_LISTE_ESLEMESI:
                continue
            liste = havuzlar.setdefault(havuz, [])
            for ham in satirlar[baslik_i + 1:]:
                for deger in _havuz_parcalari(ham[i] if i < len(ham) else None):
                    if deger not in liste:
                        liste.append(deger)
    if not havuzlar:
        logger.info("%r sayfasında eşlemesi olan havuz yok — değer havuzu farkı bakılmadı", sheet)
    return havuzlar


def havuz_farki(db: Session, havuzlar: Dict[str, List[str]]) -> List[Tuple[str, str, str, str]]:
    """(havuz, liste, yön, değer) satırları — iki yön; karşılaştırma `tk._anahtar` ile
    (aksan/büyük-küçük/noktalama duyarsız: "RED/ESASTAN" = "Red/Esastan").

    Referans listesi `LIST_REGISTRY` üzerinden yalnız OKUNUR; `active` filtresi yok
    (dropdown'dan kaldırılmış değer de "bizde var"dır — stage_decisions ile aynı okuma).
    """
    farklar: List[Tuple[str, str, str, str]] = []
    for havuz, degerler in havuzlar.items():
        liste = HAVUZ_LISTE_ESLEMESI.get(_havuz_anahtari(havuz))
        if liste is None:
            continue
        model: Any = reference_lists.LIST_REGISTRY[liste].model
        bizdekiler: Dict[str, str] = {}
        for (ad,) in db.query(model.name).order_by(model.id):
            metin = _metin(ad)
            if metin:
                bizdekiler.setdefault(tk._anahtar(metin), metin)
        teslimdekiler = {tk._anahtar(d): d for d in degerler}
        for anahtar, deger in teslimdekiler.items():
            if anahtar not in bizdekiler:
                farklar.append((havuz, liste, YON_TESLIMDE_VAR, deger))
        for anahtar, ad in bizdekiler.items():
            if anahtar not in teslimdekiler:
                farklar.append((havuz, liste, YON_BIZDE_VAR, ad))
    return farklar


def teslim_havuz_farki(teslim_id: int, *, db: Optional[Session] = None) -> List[Tuple[str, str, str, str]]:
    """Teslimin spool dosyasındaki `DEGER_HAVUZLARI` ile referans listelerinin farkı
    (boş liste = fark yok ya da sayfa yok). Referans listelerine YAZMAZ."""
    with tk._oturum(db) as session:
        teslim = tk._teslim_getir(session, teslim_id)
        yol = tk._spool_yolu(teslim)
        if yol is None:
            raise ValueError(f"Teslim #{teslim.id} dosyası spool'da yok: {teslim.spool_path or '-'}")
        havuzlar = havuz_degerlerini_oku(yol)
        farklar = havuz_farki(session, havuzlar) if havuzlar else []
    if havuzlar and not farklar:
        logger.info("Teslim #%s değer havuzları referans listeleriyle örtüşüyor (%s havuz)",
                    teslim_id, len(havuzlar))
    return farklar


def havuz_farki_csv_yaz(farklar: Sequence[Tuple[str, str, str, str]], hedef: Path) -> Optional[Path]:
    """Fark varsa CSV'yi `hedef`e yazar ve yolu döner; yoksa dosya ÜRETİLMEZ (bayat
    kopya varsa silinir — listeye değer eklendiyse eski fark cevap paketine girmesin), None."""
    hedef = Path(hedef)
    if not farklar:
        if hedef.is_file():
            hedef.unlink()
        return None
    hedef.parent.mkdir(parents=True, exist_ok=True)
    _csv_yaz(hedef, HAVUZ_FARKI_BASLIKLARI, farklar)
    logger.info("Değer havuzu farkı yazıldı: %s (%s satır)", hedef, len(farklar))
    return hedef


def havuz_farki_csv_uret(teslim_id: int, hedef: Path, *, db: Optional[Session] = None) -> Optional[Path]:
    """`teslim_havuz_farki` + `havuz_farki_csv_yaz`: fark yoksa None (dosya yok)."""
    return havuz_farki_csv_yaz(teslim_havuz_farki(teslim_id, db=db), hedef)


def havuz_farki_ozeti(farklar: Sequence[Tuple[str, str, str, str]]) -> str:
    """Bildirim gövdesi: yön başına sayı + havuz adları."""
    teslimde = [f for f in farklar if f[2] == YON_TESLIMDE_VAR]
    bizde = [f for f in farklar if f[2] == YON_BIZDE_VAR]
    havuzlar = sorted({f[0] for f in farklar})
    return (
        f"{len(farklar)} fark — {YON_TESLIMDE_VAR}: {len(teslimde)}, {YON_BIZDE_VAR}: {len(bizde)} "
        f"(havuz: {', '.join(havuzlar)}). Listeye yazılmadı; rapor cevap paketinde."
    )


# ═══════════════════════════════════════════════════════════════════════════
# Geri yükleme
# ═══════════════════════════════════════════════════════════════════════════

def _cevap_dosyalari(rapor_dizini: Path, teslim_adi: str) -> List[Tuple[Path, str]]:
    """(yerel dosya, SharePoint'teki ad) listesi — CSV/TXT, ad sırasıyla; `ozet.txt` → `ozet_<teslim>.txt`."""
    dosyalar: List[Tuple[Path, str]] = []
    for yol in sorted(rapor_dizini.iterdir(), key=lambda p: p.name):
        if not yol.is_file() or yol.suffix.lower() not in CEVAP_TURLERI:
            continue
        hedef_ad = f"ozet_{teslim_adi}.txt" if yol.name == tk.OZET_DOSYASI else yol.name
        dosyalar.append((yol, hedef_ad))
    return dosyalar


def cevap_yukle(teslim_id: int, *, db: Optional[Session] = None) -> bool:
    """`uygulandi` teslimin cevap paketini SharePoint'e yükler; hepsi gittiyse True.

    Yalnız `uygulandi` durumundan çağrılır (aksi ValueError). Zaten yüklenmişse
    True döner, yeniden yüklemez. Anahtar kapalı ya da cevap klasörü env'de tanımsızsa
    INFO + False, deftere dokunulmaz (deneme sayılmaz). Aksi hâlde eşleşme dosyası
    üretilir, rapor dizinindeki CSV/TXT'ler tek tek yüklenir; her dosyanın hatası
    WARNING, sonuç `durum_gecmisi` notuna işlenir; hepsi başarılıysa
    `cevap_yuklendi=True`. Teslim durumu HİÇBİR koşulda değişmez.
    """
    with tk._oturum(db) as session:
        teslim = tk._teslim_getir(session, teslim_id)
        tk._durum_kontrol(teslim, frozenset({tk.DURUM_UYGULANDI}), "cevap yükleme")
        if teslim.cevap_yuklendi:
            logger.info("Teslim #%s cevap paketi zaten yüklü — atlandı", teslim.id)
            return True
        if not app_settings.veri_teslim_otomasyonu_etkin(db=session):
            logger.info("Teslim #%s cevap yüklemesi atlandı: veri_teslim_otomasyonu kapalı", teslim.id)
            return False
        klasor = cevap_klasoru(str(teslim.dosya_adi))
        if klasor is None:
            logger.info(
                "Teslim #%s cevap yüklemesi atlandı: SHAREPOINT_FOLDER_TESLIM_NAME tanımsız "
                "(cevap klasörü kurulmadan yazılmaz)", teslim.id,
            )
            return False
        rapor = Path(str(teslim.rapor_dizini)) if teslim.rapor_dizini else None
        deneme = _deneme_sayisi(teslim) + 1
        if rapor is None or not rapor.is_dir():
            _gecmis_notu_ekle(teslim, f"{DENEME_NOTU_ONEKI}{deneme}: rapor dizini yok ({teslim.rapor_dizini or '-'})")
            session.commit()
            logger.warning("Teslim #%s cevap yüklenemedi: rapor dizini yok (%s)", teslim.id, teslim.rapor_dizini)
            return False
        teslim_adi = teslim_adi_uzantisiz(str(teslim.dosya_adi))

        try:
            eslesme_csv_uret(int(teslim.id), rapor / f"eslesme_{teslim_adi}.csv", db=session)
        except Exception as exc:
            _gecmis_notu_ekle(
                teslim, f"{DENEME_NOTU_ONEKI}{deneme}: eşleşme dosyası üretilemedi — {type(exc).__name__}: {exc}",
            )
            session.commit()
            logger.warning(
                "Teslim #%s cevap yüklenemedi: eşleşme dosyası üretilemedi (%s): %s",
                teslim.id, type(exc).__name__, exc,
            )
            return False

        dosyalar = _cevap_dosyalari(rapor, teslim_adi)
        hatalar: List[str] = []
        for yol, hedef_ad in dosyalar:
            try:
                _spu.upload_file_to_sharepoint(
                    str(yol), hedef_ad, target_folder_name=klasor,
                    content_type=CEVAP_TURLERI[yol.suffix.lower()],
                )
            except Exception as exc:
                hatalar.append(f"{hedef_ad}: {type(exc).__name__}: {exc}")
                logger.warning("Teslim #%s cevap dosyası yüklenemedi (%s/%s): %s", teslim.id, klasor, hedef_ad, exc)

        basarili = len(dosyalar) - len(hatalar)
        not_ = f"{DENEME_NOTU_ONEKI}{deneme}: {basarili}/{len(dosyalar)} dosya → {klasor}"
        if hatalar:
            not_ += "; hatalar: " + " | ".join(hatalar)
        _gecmis_notu_ekle(teslim, tk._kirp(not_))
        if not hatalar:
            teslim.cevap_yuklendi = True
        session.commit()
        if hatalar:
            logger.warning(
                "Teslim #%s cevap paketi eksik yüklendi (%s/%s dosya, deneme #%s) — ertesi gece turu yeniden dener",
                teslim.id, basarili, len(dosyalar), deneme,
            )
            return False
        logger.info("Teslim #%s cevap paketi yüklendi: %s dosya → %s (deneme #%s)", teslim.id, basarili, klasor, deneme)
        return True


def cevap_bekleyen_idler(db: Session, *, haric: Optional[int] = None) -> List[int]:
    """`uygulandi` + `cevap_yuklendi=False` teslimler (id sırası); `haric` bu turda zaten denenen."""
    sorgu = db.query(models.AktarimTeslimi.id).filter(
        models.AktarimTeslimi.durum == tk.DURUM_UYGULANDI,
        models.AktarimTeslimi.cevap_yuklendi.is_(False),
    )
    if haric is not None:
        sorgu = sorgu.filter(models.AktarimTeslimi.id != haric)
    return [int(tid) for (tid,) in sorgu.order_by(models.AktarimTeslimi.id).all()]


def bekleyen_cevaplari_yukle(db: Session, *, haric: Optional[int] = None) -> Dict[int, bool]:
    """Bekleyen her teslim için `cevap_yukle`; {id: sonuç}. Satır istisnası WARNING, tur sürer."""
    sonuc: Dict[int, bool] = {}
    for teslim_id in cevap_bekleyen_idler(db, haric=haric):
        try:
            sonuc[teslim_id] = cevap_yukle(teslim_id, db=db)
        except Exception as exc:
            db.rollback()
            logger.warning("Teslim #%s cevap yüklemesi turda yapılamadı (%s): %s", teslim_id, type(exc).__name__, exc)
            sonuc[teslim_id] = False
    return sonuc

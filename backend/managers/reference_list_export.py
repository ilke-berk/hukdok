"""Referans listelerinin Excel çıktısı (yönetim paneli → Excel butonu).

Liste tanımını LIST_REGISTRY'den okur; yeni bir liste eklendiğinde burada
değişiklik gerekmez. Başlık biçimi takvim raporuyla aynı (report_builder).
"""
import io
from datetime import date

from managers.reference_lists import (
    COLUMN_TITLES, LIST_REGISTRY, LIST_TITLES, get_items, resolve_list_type,
)

_HEADER_COLOR = "4A1530"   # kurumsal bordo — takvim raporuyla aynı

# Kolon adına göre makul genişlikler; listede olmayanlar için varsayılan kullanılır
_WIDTHS = {
    "code": 16, "name": 38, "tc_no": 14, "sicil_no": 14, "gorev": 14,
    "email": 32, "phone": 18, "address": 55, "city": 16,
    "description": 24, "parent_code": 18, "role_type": 12,
}
_DEFAULT_WIDTH = 22


def build_filename(list_type: str, today: date = None) -> str:
    """hukdok-avukatlar-2026-07-28 — uzantısız."""
    key = resolve_list_type(list_type) or list_type
    slug = LIST_TITLES.get(key, key).lower()
    for src, dst in (("ı", "i"), ("ğ", "g"), ("ü", "u"), ("ş", "s"), ("ö", "o"), ("ç", "c"), (" ", "-")):
        slug = slug.replace(src, dst)
    return f"hukdok-{slug}-{(today or date.today()).isoformat()}"


def list_to_excel(list_type: str, today: date = None) -> bytes:
    """Listeyi tek sayfalık .xlsx olarak üretir."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    key = resolve_list_type(list_type)
    if key is None:
        # Route zaten 404 döndürüyor; burası doğrudan çağrılara karşı koruma
        raise ValueError(f"Bilinmeyen liste türü: {list_type}")
    spec = LIST_REGISTRY[key]
    title = LIST_TITLES.get(key, key)
    columns = list(spec.fields)
    items = get_items(key)
    stamp = today or date.today()

    wb = Workbook()
    ws = wb.active
    # Sayfa adı 31 karakterle sınırlı ve bazı işaretlere izin vermez
    ws.title = title[:31]

    ws.append([f"{title} — {stamp.strftime('%d.%m.%Y')} ({len(items)} kayıt)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    header = ws.cell(row=1, column=1)
    header.font = Font(bold=True, size=13)
    header.alignment = Alignment(horizontal="left")

    ws.append([])

    header_row = 3
    ws.append([COLUMN_TITLES.get(c, c) for c in columns])
    fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for item in items:
        ws.append([item.get(c) or "" for c in columns])

    for i, column in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = _WIDTHS.get(column, _DEFAULT_WIDTH)

    ws.freeze_panes = f"A{header_row + 1}"
    if items:
        last_col = ws.cell(row=header_row, column=len(columns)).column_letter
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(items)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

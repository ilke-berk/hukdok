"""Takvim raporu üretimi — belirli bir tarih aralığındaki duruşmaları ve elle
eklenen tarih işaretlerini toplar; PDF veya Excel olarak çıktı verir.

Bir işaret bir davaya bağlıysa (duruşma kayıtları), rapora müvekkil, karşı taraf,
mahkeme, esas no ve sorumlu avukat detayları da eklenir.
"""

import io
import logging
import os
from datetime import date

from sqlalchemy import or_

import models

logger = logging.getLogger(__name__)

# --- Türkçe karakter destekli font (DejaVu) ---
_PDF_FONT = "Helvetica"
_PDF_FONT_BOLD = "Helvetica-Bold"

try:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    _DEJAVU_CANDIDATES = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]
    if os.path.exists(_DEJAVU_CANDIDATES[0]):
        pdfmetrics.registerFont(TTFont("DejaVuSans", _DEJAVU_CANDIDATES[0]))
        _PDF_FONT = "DejaVuSans"
        if os.path.exists(_DEJAVU_CANDIDATES[1]):
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _DEJAVU_CANDIDATES[1]))
            _PDF_FONT_BOLD = "DejaVuSans-Bold"
        else:
            _PDF_FONT_BOLD = "DejaVuSans"
except Exception as e:  # pragma: no cover - font yoksa Helvetica'ya düşer
    logger.warning("DejaVu font kaydı başarısız, Helvetica kullanılacak: %s", e)


# Rapor sütun başlıkları (sıra ile)
COLUMNS = [
    "Tarih", "Saat", "Tür", "Açıklama", "Esas No",
    "Mahkeme", "Müvekkil", "Karşı Taraf", "Sorumlu Avukat",
]


def _fmt_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def build_report_rows(db, tenant_id: str, start: date, end: date):
    """Tarih aralığındaki tüm işaretleri (duruşma + elle) toplar, davaya bağlı
    olanları detaylandırır. Tarihe (ve saate) göre sıralı liste döndürür."""
    rows = []

    # --- Duruşmalar (davaya bağlı) ---
    hearings = (
        db.query(models.HearingDate)
        .join(models.Case, models.HearingDate.case_id == models.Case.id)
        .filter(
            models.HearingDate.hearing_date >= start,
            models.HearingDate.hearing_date <= end,
            or_(models.Case.tenant_id == tenant_id, models.Case.tenant_id.is_(None)),
        )
        .all()
    )
    for h in hearings:
        case = h.case
        clients, counters = [], []
        if case and case.parties:
            for p in case.parties:
                if p.party_type == "CLIENT":
                    clients.append(p.name)
                elif p.party_type == "COUNTER":
                    counters.append(p.name)
        rows.append({
            "date": h.hearing_date,
            "date_str": _fmt_date(h.hearing_date),
            "time": h.hearing_time or "",
            "type": "Duruşma",
            "title": h.note or "Duruşma",
            "esas_no": (case.esas_no if case else "") or "",
            "court": (case.court if case else "") or "",
            "client": ", ".join(clients),
            "counter": ", ".join(counters),
            "lawyer": h.lawyer_name or (case.responsible_lawyer_name if case else "") or "",
            "case_id": case.id if case else None,
        })

    # --- Elle eklenen işaretler (davaya bağlı değil) ---
    events = (
        db.query(models.CalendarEvent)
        .filter(
            models.CalendarEvent.event_date >= start,
            models.CalendarEvent.event_date <= end,
            or_(models.CalendarEvent.tenant_id == tenant_id, models.CalendarEvent.tenant_id.is_(None)),
        )
        .all()
    )
    for e in events:
        rows.append({
            "date": e.event_date,
            "date_str": _fmt_date(e.event_date),
            "time": e.event_time or "",
            "type": "İşaret",
            "title": e.title,
            "esas_no": "",
            "court": "",
            "client": "",
            "counter": "",
            "lawyer": "",
            "case_id": None,
        })

    rows.sort(key=lambda r: (r["date"], r["time"] or ""))
    return rows


def _row_values(r):
    return [
        r["date_str"], r["time"], r["type"], r["title"], r["esas_no"],
        r["court"], r["client"], r["counter"], r["lawyer"],
    ]


def rows_to_excel(rows, start: date, end: date) -> bytes:
    """Rapor satırlarını .xlsx olarak üretir."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Takvim Raporu"

    # Başlık satırı (aralık bilgisi)
    ws.append([f"Takvim Raporu — {_fmt_date(start)} – {_fmt_date(end)}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left")

    ws.append([])  # boş satır

    # Sütun başlıkları
    header_row = 3
    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="4A1530", end_color="4A1530", fill_type="solid")
    for col in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Veri satırları
    for r in rows:
        ws.append(_row_values(r))

    # Sütun genişlikleri
    widths = [12, 7, 9, 38, 14, 28, 28, 28, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=header_row, column=len(COLUMNS)).column_letter}{header_row + len(rows)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_pdf(rows, start: date, end: date) -> bytes:
    """Rapor satırlarını yatay A4 PDF olarak üretir."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Takvim Raporu",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName=_PDF_FONT_BOLD,
        fontSize=15, textColor=colors.HexColor("#4A1530"), alignment=0, spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "ReportSub", parent=styles["Normal"], fontName=_PDF_FONT,
        fontSize=9, textColor=colors.HexColor("#666666"), spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontName=_PDF_FONT, fontSize=7.5, leading=9,
    )
    header_style = ParagraphStyle(
        "Head", parent=styles["Normal"], fontName=_PDF_FONT_BOLD, fontSize=8,
        textColor=colors.white, leading=10,
    )

    elements = [
        Paragraph("Takvim Raporu", title_style),
        Paragraph(f"{_fmt_date(start)} – {_fmt_date(end)} · {len(rows)} kayıt", sub_style),
    ]

    # Tablo verisi (Paragraph ile sarılır → uzun metinler satır kırar)
    data = [[Paragraph(c, header_style) for c in COLUMNS]]
    for r in rows:
        data.append([Paragraph(_esc(v), cell_style) for v in _row_values(r)])

    # mm cinsinden sütun genişlikleri (toplam ~273mm yatay A4 içeriği)
    col_widths = [20, 12, 16, 52, 24, 42, 38, 38, 31]
    table = Table(data, colWidths=[w * mm for w in col_widths], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A1530")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8CFC4")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF6F0")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))

    if rows:
        elements.append(table)
    else:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("Bu aralıkta kayıt bulunamadı.", cell_style))

    doc.build(elements)
    return buf.getvalue()


def _esc(v) -> str:
    """Paragraph içine güvenli yerleştirme için minimal XML kaçışı."""
    s = "" if v is None else str(v)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
